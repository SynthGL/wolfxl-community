#!/usr/bin/env python3
"""Generate the fidelity corpus.

Every workbook here is generated on your machine. The harness ships no
third-party workbook bytes, so there is nothing to license and nothing to
take on trust: run the generator and diff the output yourself.

Two tiers, and the difference matters when reading results.

Tier A, `openpyxl-authored`
    Written by openpyxl. Tables, conditional formatting, data validation,
    charts, comments, defined names, hyperlinks, page setup, styles.

Tier B, `synthetic-injection`
    OOXML parts that no Python library can author from scratch: pivot caches,
    slicers, timelines, VBA, custom XML, embedded media. These are assembled
    by writing the parts, relationships, and content-type overrides directly
    into the package.

    Tier B parts are minimal and hand-built. They are NOT Excel-authored, and
    Excel's own equivalents are richer. They are validated structurally, by
    checking the zip entries, XML wellformedness, content-type overrides, and
    relationship targets, and not by opening them in Excel. What they test is
    narrow and real: whether a save carries a part and its relationship across
    or quietly drops them. Read any Tier B result as a part-retention result,
    not as a statement about Excel round-trip behaviour.
"""

from __future__ import annotations

import argparse
import json
import re
import struct
import zipfile
import zlib
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Callable
from xml.etree import ElementTree

if TYPE_CHECKING:  # openpyxl is only needed to build the corpus, not to read it
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MANIFEST_NOTE = (
    "Generated locally by corpus.py. No third-party workbook bytes are distributed."
)
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
XR10_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
DRAW_SS_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
DRAW_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

CONTENT_TYPES = {
    "pivotTable": "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml",
    "pivotCacheDefinition": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"
    ),
    "pivotCacheRecords": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"
    ),
    "slicer": "application/vnd.ms-excel.slicer+xml",
    "slicerCache": "application/vnd.ms-excel.slicerCache+xml",
    "timeline": "application/vnd.ms-excel.timeline+xml",
    "timelineCache": "application/vnd.ms-excel.timelineCache+xml",
    "vbaProject": "application/vnd.ms-office.vbaProject",
    "customXmlProps": "application/vnd.openxmlformats-officedocument.customXmlProperties+xml",
    "drawing": "application/vnd.openxmlformats-officedocument.drawing+xml",
    "macroWorkbook": "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
}

REL_TYPES = {
    "pivotTable": f"{REL_NS}/pivotTable",
    "pivotCacheDefinition": f"{REL_NS}/pivotCacheDefinition",
    "pivotCacheRecords": f"{REL_NS}/pivotCacheRecords",
    "slicer": "http://schemas.microsoft.com/office/2007/relationships/slicer",
    "slicerCache": "http://schemas.microsoft.com/office/2007/relationships/slicerCache",
    "timeline": "http://schemas.microsoft.com/office/2011/relationships/timeline",
    "timelineCache": "http://schemas.microsoft.com/office/2011/relationships/timelineCache",
    "vbaProject": "http://schemas.microsoft.com/office/2006/relationships/vbaProject",
    "customXml": f"{REL_NS}/customXml",
    "drawing": f"{REL_NS}/drawing",
    "image": f"{REL_NS}/image",
}


@dataclass(frozen=True)
class Fixture:
    """One corpus workbook and the parts a faithful save must keep."""

    name: str
    tier: str
    description: str
    expected_parts: tuple[str, ...]
    expected_features: tuple[str, ...]


def _normalize_package(path: Path) -> None:
    """Make generated package bytes reproducible across clean runs."""
    parts = _read_parts(path)
    core = parts.get("docProps/core.xml")
    if core is not None:
        core = re.sub(
            rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
            rb"\g<1>2020-01-01T00:00:00Z\g<2>",
            core,
        )
        parts["docProps/core.xml"] = core
    _write_parts(path, parts)

def _read_parts(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def _write_parts(path: Path, parts: dict[str, bytes]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as archive:
        for name, payload in parts.items():
            info = zipfile.ZipInfo(name, date_time=(2020, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            archive.writestr(info, payload)


def _add_content_types(
    parts: dict[str, bytes],
    *,
    overrides: dict[str, str] | None = None,
    defaults: dict[str, str] | None = None,
    replace: dict[str, str] | None = None,
) -> None:
    """Add or replace entries in [Content_Types].xml."""
    ElementTree.register_namespace("", CT_NS)
    root = ElementTree.fromstring(parts["[Content_Types].xml"].decode("utf-8"))

    existing_defaults = {
        element.get("Extension", "").lower()
        for element in root.findall(f"{{{CT_NS}}}Default")
    }
    for extension, content_type in (defaults or {}).items():
        if extension.lower() in existing_defaults:
            continue
        element = ElementTree.SubElement(root, f"{{{CT_NS}}}Default")
        element.set("Extension", extension)
        element.set("ContentType", content_type)

    by_partname = {
        element.get("PartName"): element for element in root.findall(f"{{{CT_NS}}}Override")
    }
    for part_name, content_type in (replace or {}).items():
        element = by_partname.get(part_name)
        if element is not None:
            element.set("ContentType", content_type)

    for part_name, content_type in (overrides or {}).items():
        if part_name in by_partname:
            by_partname[part_name].set("ContentType", content_type)
            continue
        element = ElementTree.SubElement(root, f"{{{CT_NS}}}Override")
        element.set("PartName", part_name)
        element.set("ContentType", content_type)

    parts["[Content_Types].xml"] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _add_relationships(
    parts: dict[str, bytes], rels_path: str, entries: list[tuple[str, str, str]]
) -> None:
    """Append (id, type, target) relationships, creating the part if needed."""
    ElementTree.register_namespace("", PKG_REL_NS)
    if rels_path in parts:
        root = ElementTree.fromstring(parts[rels_path].decode("utf-8"))
    else:
        root = ElementTree.Element(f"{{{PKG_REL_NS}}}Relationships")
    for rel_id, rel_type, target in entries:
        element = ElementTree.SubElement(root, f"{{{PKG_REL_NS}}}Relationship")
        element.set("Id", rel_id)
        element.set("Type", rel_type)
        element.set("Target", target)
    parts[rels_path] = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _inject_into_element(payload: bytes, marker: str, insert: str, *, before: bool) -> bytes:
    """Insert raw XML next to a marker tag, keeping schema element order.

    OOXML validates child order, so appending blindly to the end of
    worksheet or workbook XML produces a file Excel rejects. Anchoring on a
    known sibling keeps the sequence legal.
    """
    text = payload.decode("utf-8")
    index = text.find(marker)
    if index == -1:
        closing = text.rfind("</")
        return (text[:closing] + insert + text[closing:]).encode("utf-8")
    if before:
        return (text[:index] + insert + text[index:]).encode("utf-8")
    end = text.find(">", index)
    return (text[: end + 1] + insert + text[end + 1 :]).encode("utf-8")


def _ensure_r_namespace(payload: bytes) -> bytes:
    """Declare xmlns:r on the root element when it is missing.

    openpyxl only emits the relationship namespace on sheets that already
    need it, so injecting an `r:id` attribute into a plain sheet would
    otherwise produce an unbound prefix and unreadable XML.
    """
    text = payload.decode("utf-8")
    end = text.find(">")
    if end == -1 or 'xmlns:r=' in text[:end]:
        return payload
    return (text[:end] + f' xmlns:r="{REL_NS}"' + text[end:]).encode("utf-8")


def _minimal_png(width: int = 8, height: int = 8) -> bytes:
    """Build a valid PNG with zlib and struct, so Pillow is not required."""

    def chunk(tag: bytes, payload: bytes) -> bytes:
        return (
            struct.pack(">I", len(payload))
            + tag
            + payload
            + struct.pack(">I", zlib.crc32(tag + payload) & 0xFFFFFFFF)
        )

    raw = bytearray()
    for y in range(height):
        raw.append(0)  # no filter
        for x in range(width):
            raw += bytes((x * 32 % 256, y * 32 % 256, 128))
    header = struct.pack(">2I5B", width, height, 8, 2, 0, 0, 0)
    return (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", header)
        + chunk(b"IDAT", zlib.compress(bytes(raw), 9))
        + chunk(b"IEND", b"")
    )


# ---------------------------------------------------------------------------
# Tier A: openpyxl-authored
# ---------------------------------------------------------------------------

def _active(workbook: "Workbook") -> "Worksheet":
    """Return the active worksheet, refusing a workbook that has none.

    openpyxl types `active` as optional. Asserting here keeps the builders
    readable and turns a surprising None into an obvious failure.
    """
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - Workbook() always has a sheet
        raise RuntimeError("new workbook has no active worksheet")
    return sheet


def build_styles(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Styled"
    sheet["A1"] = "Region"
    sheet["B1"] = "Revenue"
    sheet["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="336699")
    sheet["B1"].font = Font(bold=True, italic=True, underline="single")
    thin = Side(style="thin", color="999999")
    sheet["B1"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row, (region, revenue) in enumerate(
        [("North", 1000.5), ("South", 2500.25), ("East", 1750), ("West", 3200.75)], start=2
    ):
        sheet[f"A{row}"] = region
        sheet[f"B{row}"] = revenue
        sheet[f"B{row}"].number_format = "#,##0.00"
    sheet["C1"] = "Share"
    for row in range(2, 6):
        sheet[f"C{row}"] = f"=B{row}/SUM($B$2:$B$5)"
        sheet[f"C{row}"].number_format = "0.0%"
    sheet.merge_cells("A7:C7")
    sheet["A7"] = "Merged footer"
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 22
    sheet.row_dimensions[1].height = 28
    workbook.create_sheet("Empty")
    workbook.save(path)


def build_tables_and_validation(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Data"
    sheet.append(["Item", "Qty", "Status", "Score"])
    rows = [
        ("Widget", 12, "Open", 45),
        ("Gadget", 7, "Closed", 88),
        ("Sprocket", 22, "Open", 61),
        ("Flange", 3, "Blocked", 12),
        ("Bearing", 18, "Closed", 97),
    ]
    for row in rows:
        sheet.append(row)

    table = Table(displayName="ItemTable", ref="A1:D6")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    sheet.add_table(table)

    status = DataValidation(
        type="list", formula1='"Open,Closed,Blocked"', allow_blank=True, showDropDown=False
    )
    status.error = "Pick a listed status"
    status.errorTitle = "Invalid status"
    sheet.add_data_validation(status)
    status.add("C2:C200")

    qty = DataValidation(type="whole", operator="between", formula1="0", formula2="1000")
    sheet.add_data_validation(qty)
    qty.add("B2:B200")

    sheet.conditional_formatting.add(
        "D2:D6", ColorScaleRule(start_type="min", start_color="FFF8696B", end_type="max", end_color="FF63BE7B")
    )
    sheet.conditional_formatting.add("B2:B6", DataBarRule(start_type="min", end_type="max", color="FF638EC6"))
    sheet.conditional_formatting.add("D2:D6", IconSetRule("3TrafficLights1", "percent", [0, 33, 67]))

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Structured reference"
    summary["B1"] = "=SUM(ItemTable[Qty])"
    summary["A2"] = "Cross sheet"
    summary["B2"] = "=SUM(Data!D2:D6)"
    workbook.save(path)


def build_charts(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Numbers"
    sheet.append(["Month", "Plan", "Actual"])
    for month, plan, actual in [
        ("Jan", 100, 92),
        ("Feb", 120, 133),
        ("Mar", 140, 128),
        ("Apr", 160, 171),
        ("May", 180, 164),
    ]:
        sheet.append([month, plan, actual])

    labels = Reference(sheet, min_col=1, min_row=2, max_row=6)
    values = Reference(sheet, min_col=2, min_row=1, max_col=3, max_row=6)

    bar = BarChart()
    bar.title = "Plan versus actual"
    bar.type = "col"
    bar.add_data(values, titles_from_data=True)
    bar.set_categories(labels)
    sheet.add_chart(bar, anchor="E2")

    line = LineChart()
    line.title = "Trend"
    line.add_data(values, titles_from_data=True)
    line.set_categories(labels)
    sheet.add_chart(line, anchor="E20")

    pie = PieChart()
    pie.title = "Actual split"
    pie.add_data(Reference(sheet, min_col=3, min_row=1, max_row=6), titles_from_data=True)
    pie.set_categories(labels)
    chart_sheet = workbook.create_sheet("Charts")
    chart_sheet.add_chart(pie, anchor="B2")
    workbook.save(path)


def build_names_links_comments(path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.pagebreak import Break

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Notes"
    sheet["A1"] = "Rate"
    sheet["B1"] = 0.075
    sheet["A2"] = "Docs"
    sheet["B2"] = "OOXML spec"
    sheet["B2"].hyperlink = "https://www.ecma-international.org/publications-and-standards/standards/ecma-376/"
    sheet["B2"].style = "Hyperlink"
    sheet["A4"] = "Reviewed"
    sheet["A4"].comment = Comment("Check this figure before sending.", "Harness", height=90, width=220)
    sheet["A5"] = "Second note"
    sheet["A5"].comment = Comment("Threaded comments are an Excel feature.", "Harness")

    workbook.defined_names.add(DefinedName("TaxRate", attr_text="Notes!$B$1"))
    workbook.defined_names.add(DefinedName("ReportRange", attr_text="Notes!$A$1:$B$5"))

    if sheet.oddHeader is None or sheet.oddHeader.center is None:
        raise RuntimeError("worksheet has no odd header center")
    if sheet.oddFooter is None or sheet.oddFooter.right is None:
        raise RuntimeError("worksheet has no odd footer right section")
    sheet.oddHeader.center.text = "Fidelity harness"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.print_title_rows = "1:1"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_options.horizontalCentered = True
    sheet.row_breaks.append(Break(id=3))
    workbook.save(path)


# ---------------------------------------------------------------------------
# Tier B: synthetic OOXML injection
# ---------------------------------------------------------------------------


def build_pivot(path: Path) -> None:
    """Inject a pivot cache, cache records, and a pivot table."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Source"
    sheet.append(["Region", "Amount"])
    for region, amount in [("North", 10), ("South", 20), ("North", 30), ("South", 40)]:
        sheet.append([region, amount])
    workbook.create_sheet("Pivot")
    workbook.save(path)

    parts = _read_parts(path)

    cache_definition = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheDefinition xmlns="{MAIN_NS}" xmlns:r="{REL_NS}" r:id="rId1" refreshedBy="harness" recordCount="4" createdVersion="6" refreshedVersion="6" minRefreshableVersion="3" refreshOnLoad="0">
  <cacheSource type="worksheet"><worksheetSource ref="A1:B5" sheet="Source"/></cacheSource>
  <cacheFields count="2">
    <cacheField name="Region" numFmtId="0">
      <sharedItems count="2"><s v="North"/><s v="South"/></sharedItems>
    </cacheField>
    <cacheField name="Amount" numFmtId="0">
      <sharedItems containsSemiMixedTypes="0" containsString="0" containsNumber="1" containsInteger="1" minValue="10" maxValue="40"/>
    </cacheField>
  </cacheFields>
</pivotCacheDefinition>""".encode("utf-8")

    cache_records = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotCacheRecords xmlns="{MAIN_NS}" xmlns:r="{REL_NS}" count="4">
  <r><x v="0"/><n v="10"/></r>
  <r><x v="1"/><n v="20"/></r>
  <r><x v="0"/><n v="30"/></r>
  <r><x v="1"/><n v="40"/></r>
</pivotCacheRecords>""".encode("utf-8")

    pivot_table = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<pivotTableDefinition xmlns="{MAIN_NS}" name="RegionPivot" cacheId="1" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" dataCaption="Values" updatedVersion="6" minRefreshableVersion="3" createdVersion="6" itemPrintTitles="1" useAutoFormatting="1" indent="0" outline="1" outlineData="1" multipleFieldFilters="0">
  <location ref="A3:B6" firstHeaderRow="1" firstDataRow="1" firstDataCol="1"/>
  <pivotFields count="2">
    <pivotField axis="axisRow" showAll="0">
      <items count="3"><item x="0"/><item x="1"/><item t="default"/></items>
    </pivotField>
    <pivotField dataField="1" showAll="0"/>
  </pivotFields>
  <rowFields count="1"><field x="0"/></rowFields>
  <rowItems count="3"><i><x/></i><i><x v="1"/></i><i t="grand"><x/></i></rowItems>
  <colItems count="1"><i/></colItems>
  <dataFields count="1"><dataField name="Sum of Amount" fld="1" baseField="0" baseItem="0"/></dataFields>
</pivotTableDefinition>""".encode("utf-8")

    parts["xl/pivotCache/pivotCacheDefinition1.xml"] = cache_definition
    parts["xl/pivotCache/pivotCacheRecords1.xml"] = cache_records
    parts["xl/pivotTables/pivotTable1.xml"] = pivot_table

    _add_content_types(
        parts,
        overrides={
            "/xl/pivotCache/pivotCacheDefinition1.xml": CONTENT_TYPES["pivotCacheDefinition"],
            "/xl/pivotCache/pivotCacheRecords1.xml": CONTENT_TYPES["pivotCacheRecords"],
            "/xl/pivotTables/pivotTable1.xml": CONTENT_TYPES["pivotTable"],
        },
    )
    _add_relationships(
        parts,
        "xl/_rels/workbook.xml.rels",
        [("rIdPivotCache1", REL_TYPES["pivotCacheDefinition"], "pivotCache/pivotCacheDefinition1.xml")],
    )
    _add_relationships(
        parts,
        "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
        [("rId1", REL_TYPES["pivotCacheRecords"], "pivotCacheRecords1.xml")],
    )
    _add_relationships(
        parts,
        "xl/pivotTables/_rels/pivotTable1.xml.rels",
        [("rId1", REL_TYPES["pivotCacheDefinition"], "../pivotCache/pivotCacheDefinition1.xml")],
    )
    _add_relationships(
        parts,
        "xl/worksheets/_rels/sheet2.xml.rels",
        [("rIdPivot1", REL_TYPES["pivotTable"], "../pivotTables/pivotTable1.xml")],
    )

    # pivotCaches must sit after definedNames and before calcPr.
    parts["xl/workbook.xml"] = _inject_into_element(
        parts["xl/workbook.xml"],
        "<calcPr",
        '<pivotCaches><pivotCache cacheId="1" r:id="rIdPivotCache1"/></pivotCaches>',
        before=True,
    )
    _write_parts(path, parts)


def build_slicers_and_timelines(path: Path) -> None:
    """Inject slicer and timeline parts, caches, and their extension lists."""
    build_pivot(path)
    parts = _read_parts(path)

    slicer_cache = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<slicerCacheDefinition xmlns="{X14_NS}" xmlns:mc="{MC_NS}" mc:Ignorable="x" name="Slicer_Region" sourceName="Region">
  <pivotTables><pivotTable tabId="1" name="RegionPivot"/></pivotTables>
  <data>
    <olap><levels count="0"/><selections count="0"/></olap>
  </data>
</slicerCacheDefinition>""".encode("utf-8")

    slicer = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<slicers xmlns="{X14_NS}">
  <slicer name="Region" cache="Slicer_Region" caption="Region" rowHeight="234950"/>
</slicers>""".encode("utf-8")

    timeline_cache = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<timelineCacheDefinition xmlns="{XR10_NS}" name="NativeTimeline_Date" sourceName="Date">
  <pivotTables><pivotTable tabId="1" name="RegionPivot"/></pivotTables>
  <state minimalRefreshVersion="6" lastRefreshVersion="6" pivotCacheId="1" filterType="unknown"/>
</timelineCacheDefinition>""".encode("utf-8")

    timeline = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<timelines xmlns="{XR10_NS}">
  <timeline name="Date" cache="NativeTimeline_Date" caption="Date" level="2" selectionLevel="2"/>
</timelines>""".encode("utf-8")

    parts["xl/slicerCaches/slicerCache1.xml"] = slicer_cache
    parts["xl/slicers/slicer1.xml"] = slicer
    parts["xl/timelineCaches/timelineCache1.xml"] = timeline_cache
    parts["xl/timelines/timeline1.xml"] = timeline

    _add_content_types(
        parts,
        overrides={
            "/xl/slicerCaches/slicerCache1.xml": CONTENT_TYPES["slicerCache"],
            "/xl/slicers/slicer1.xml": CONTENT_TYPES["slicer"],
            "/xl/timelineCaches/timelineCache1.xml": CONTENT_TYPES["timelineCache"],
            "/xl/timelines/timeline1.xml": CONTENT_TYPES["timeline"],
        },
    )
    _add_relationships(
        parts,
        "xl/_rels/workbook.xml.rels",
        [
            ("rIdSlicerCache1", REL_TYPES["slicerCache"], "slicerCaches/slicerCache1.xml"),
            ("rIdTimelineCache1", REL_TYPES["timelineCache"], "timelineCaches/timelineCache1.xml"),
        ],
    )
    _add_relationships(
        parts,
        "xl/worksheets/_rels/sheet2.xml.rels",
        [
            ("rIdSlicer1", REL_TYPES["slicer"], "../slicers/slicer1.xml"),
            ("rIdTimeline1", REL_TYPES["timeline"], "../timelines/timeline1.xml"),
        ],
    )

    workbook_ext = (
        f'<extLst><ext uri="{{BBE1A952-AA13-448e-AADC-164F8A28A991}}" xmlns:x14="{X14_NS}">'
        '<x14:slicerCaches><x14:slicerCache r:id="rIdSlicerCache1"/></x14:slicerCaches>'
        "</ext>"
        f'<ext uri="{{9260A510-F301-46a8-8635-F512D64BE5F5}}" xmlns:x15="{XR10_NS}">'
        '<x15:timelineCacheRefs><x15:timelineCacheRef r:id="rIdTimelineCache1"/></x15:timelineCacheRefs>'
        "</ext></extLst>"
    )
    parts["xl/workbook.xml"] = _inject_into_element(
        parts["xl/workbook.xml"], "</workbook>", workbook_ext, before=True
    )

    sheet_ext = (
        f'<extLst><ext uri="{{A8765BA9-456A-4dab-B4F3-ACF1C6B7E3CD}}" xmlns:x14="{X14_NS}">'
        '<x14:slicerList><x14:slicer r:id="rIdSlicer1"/></x14:slicerList>'
        "</ext>"
        f'<ext uri="{{7E03D99C-DC04-49d9-9315-930204A7B6E9}}" xmlns:x15="{XR10_NS}">'
        '<x15:timelineRefs><x15:timelineRef r:id="rIdTimeline1"/></x15:timelineRefs>'
        "</ext></extLst>"
    )
    parts["xl/worksheets/sheet2.xml"] = _inject_into_element(
        _ensure_r_namespace(parts["xl/worksheets/sheet2.xml"]),
        "</worksheet>",
        sheet_ext,
        before=True,
    )
    _write_parts(path, parts)


def build_macros(path: Path) -> None:
    """Inject a vbaProject part and switch the workbook to macro-enabled."""
    from openpyxl import Workbook

    staging = path.with_suffix(".staging.xlsx")
    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Macros"
    sheet["A1"] = "This workbook carries a vbaProject part."
    sheet["A2"] = "=1+1"
    workbook.save(staging)

    parts = _read_parts(staging)
    staging.unlink()

    # A labelled placeholder, not compiled VBA. The fidelity question is
    # whether the binary part and its relationship survive the save.
    parts["xl/vbaProject.bin"] = (
        b"WOLFXL-FIDELITY-HARNESS-SYNTHETIC-VBAPROJECT\x00"
        b"Not compiled VBA. Present to test binary part retention.\x00"
    ) + bytes(range(256)) * 4

    _add_content_types(
        parts,
        overrides={"/xl/vbaProject.bin": CONTENT_TYPES["vbaProject"]},
        replace={"/xl/workbook.xml": CONTENT_TYPES["macroWorkbook"]},
    )
    _add_relationships(
        parts, "xl/_rels/workbook.xml.rels", [("rIdVba1", REL_TYPES["vbaProject"], "vbaProject.bin")]
    )
    _write_parts(path, parts)


def build_custom_xml_and_media(path: Path) -> None:
    """Inject custom XML parts plus an embedded image with its drawing."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = _active(workbook)
    sheet.title = "Media"
    sheet["A1"] = "Carries customXml parts and an embedded PNG."
    workbook.save(path)

    parts = _read_parts(path)

    parts["customXml/item1.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<harnessMetadata xmlns="urn:wolfxl:fidelity-harness">'
        b"<generator>corpus.py</generator><tier>synthetic-injection</tier>"
        b"</harnessMetadata>"
    )
    parts["customXml/itemProps1.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<ds:datastoreItem xmlns:ds="http://schemas.openxmlformats.org/officeDocument/2006/customXml" '
        b'ds:itemID="{4F2E1B7C-0A3D-4E5F-9C81-2B6D7E8F1A34}">'
        b'<ds:schemaRefs><ds:schemaRef ds:uri="urn:wolfxl:fidelity-harness"/></ds:schemaRefs>'
        b"</ds:datastoreItem>"
    )
    parts["xl/media/image1.png"] = _minimal_png()

    parts["xl/drawings/drawing1.xml"] = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="{DRAW_SS_NS}" xmlns:a="{DRAW_NS}" xmlns:r="{REL_NS}">
  <xdr:oneCellAnchor>
    <xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
    <xdr:ext cx="914400" cy="914400"/>
    <xdr:pic>
      <xdr:nvPicPr>
        <xdr:cNvPr id="2" name="Harness image"/>
        <xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>
      </xdr:nvPicPr>
      <xdr:blipFill><a:blip r:embed="rIdImage1"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>
      <xdr:spPr>
        <a:xfrm><a:off x="0" y="0"/><a:ext cx="914400" cy="914400"/></a:xfrm>
        <a:prstGeom prst="rect"><a:avLst/></a:prstGeom>
      </xdr:spPr>
    </xdr:pic>
    <xdr:clientData/>
  </xdr:oneCellAnchor>
</xdr:wsDr>""".encode("utf-8")

    _add_content_types(
        parts,
        defaults={"png": "image/png"},
        overrides={
            "/customXml/itemProps1.xml": CONTENT_TYPES["customXmlProps"],
            "/xl/drawings/drawing1.xml": CONTENT_TYPES["drawing"],
        },
    )
    _add_relationships(
        parts, "_rels/.rels", [("rIdCustomXml1", REL_TYPES["customXml"], "customXml/item1.xml")]
    )
    _add_relationships(
        parts,
        "customXml/_rels/item1.xml.rels",
        [
            (
                "rIdProps1",
                f"{REL_NS}/customXmlProps",
                "itemProps1.xml",
            )
        ],
    )
    _add_relationships(
        parts, "xl/drawings/_rels/drawing1.xml.rels", [("rIdImage1", REL_TYPES["image"], "../media/image1.png")]
    )
    _add_relationships(
        parts, "xl/worksheets/_rels/sheet1.xml.rels", [("rIdDrawing1", REL_TYPES["drawing"], "../drawings/drawing1.xml")]
    )
    # <drawing> is the last child of worksheet, after any extLst-free content.
    parts["xl/worksheets/sheet1.xml"] = _inject_into_element(
        _ensure_r_namespace(parts["xl/worksheets/sheet1.xml"]),
        "</worksheet>",
        '<drawing r:id="rIdDrawing1"/>',
        before=True,
    )
    _write_parts(path, parts)


# ---------------------------------------------------------------------------
# registry
# ---------------------------------------------------------------------------

FIXTURES: tuple[tuple[Fixture, Callable[[Path], None]], ...] = (
    (
        Fixture(
            name="styles.xlsx",
            tier="openpyxl-authored",
            description="Fonts, fills, borders, number formats, merges, freeze panes, formulas.",
            expected_parts=("xl/styles.xml", "xl/worksheets/sheet1.xml"),
            expected_features=("styles",),
        ),
        build_styles,
    ),
    (
        Fixture(
            name="tables-and-validation.xlsx",
            tier="openpyxl-authored",
            description="Table with structured references, data validation, three conditional formats.",
            expected_parts=("xl/tables/table1.xml",),
            expected_features=("table", "data_validation", "conditional_formatting"),
        ),
        build_tables_and_validation,
    ),
    (
        Fixture(
            name="charts.xlsx",
            tier="openpyxl-authored",
            description="Bar, line, and pie charts across a worksheet and a second sheet.",
            expected_parts=("xl/charts/chart1.xml", "xl/drawings/drawing1.xml"),
            expected_features=("chart", "drawing"),
        ),
        build_charts,
    ),
    (
        Fixture(
            name="names-links-comments.xlsx",
            tier="openpyxl-authored",
            description="Defined names, hyperlink, comments with VML, page setup, print titles, page break.",
            expected_parts=("xl/comments/comment1.xml", "xl/drawings/commentsDrawing1.vml"),
            expected_features=("comment", "defined_names", "page_setup"),
        ),
        build_names_links_comments,
    ),
    (
        Fixture(
            name="pivot.xlsx",
            tier="synthetic-injection",
            description="Pivot table with a populated pivotCacheRecords part.",
            expected_parts=(
                "xl/pivotCache/pivotCacheDefinition1.xml",
                "xl/pivotCache/pivotCacheRecords1.xml",
                "xl/pivotTables/pivotTable1.xml",
            ),
            expected_features=("pivot",),
        ),
        build_pivot,
    ),
    (
        Fixture(
            name="slicers-and-timelines.xlsx",
            tier="synthetic-injection",
            description="Pivot plus slicer, slicerCache, timeline, and timelineCache parts.",
            expected_parts=(
                "xl/slicers/slicer1.xml",
                "xl/slicerCaches/slicerCache1.xml",
                "xl/timelines/timeline1.xml",
                "xl/timelineCaches/timelineCache1.xml",
            ),
            expected_features=("slicer", "timeline", "pivot"),
        ),
        build_slicers_and_timelines,
    ),
    (
        Fixture(
            name="macros.xlsm",
            tier="synthetic-injection",
            description="Macro-enabled workbook carrying a binary vbaProject part.",
            expected_parts=("xl/vbaProject.bin",),
            expected_features=("vba",),
        ),
        build_macros,
    ),
    (
        Fixture(
            name="custom-xml-and-media.xlsx",
            tier="synthetic-injection",
            description="Custom XML datastore parts plus an embedded PNG with its drawing.",
            expected_parts=(
                "customXml/item1.xml",
                "customXml/itemProps1.xml",
                "xl/media/image1.png",
                "xl/drawings/drawing1.xml",
            ),
            expected_features=("custom_xml", "media", "drawing"),
        ),
        build_custom_xml_and_media,
    ),
)


def _prepare_destination(destination: Path, *, clean: bool) -> None:
    """Prepare an output directory without recursively deleting user data."""
    if not destination.exists():
        destination.mkdir(parents=True)
        return
    if not destination.is_dir():
        raise RuntimeError(f"corpus destination is not a directory: {destination}")
    if not clean:
        return

    manifest_path = destination / "manifest.json"
    entries = list(destination.iterdir())
    if not entries:
        return
    if not manifest_path.is_file():
        raise RuntimeError(
            f"refusing to clean nonempty unmarked directory: {destination}; "
            "choose a new directory or pass --keep to overwrite fixture names"
        )
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise RuntimeError(f"refusing to clean directory with invalid manifest: {destination}") from error
    if manifest.get("note") != MANIFEST_NOTE:
        raise RuntimeError(f"refusing to clean directory with foreign manifest: {destination}")

    # Remove only filenames owned by this generator. Preserve every unrelated
    # file, even inside a marked directory.
    for fixture, _ in FIXTURES:
        fixture_path = destination / fixture.name
        if fixture_path.is_file():
            fixture_path.unlink()
    manifest_path.unlink()


def generate(destination: Path, *, clean: bool = True) -> list[Fixture]:
    """Write every fixture into `destination` and return the manifest."""
    _prepare_destination(destination, clean=clean)
    written: list[Fixture] = []
    for fixture, builder in FIXTURES:
        builder(destination / fixture.name)
        _normalize_package(destination / fixture.name)
        written.append(fixture)
    manifest = {
        "note": MANIFEST_NOTE,
        "fixtures": [asdict(fixture) for fixture in written],
    }
    (destination / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return written


def verify(destination: Path) -> list[str]:
    """Structurally validate the generated corpus. Returns a list of problems."""
    problems: list[str] = []
    for fixture, _ in FIXTURES:
        path = destination / fixture.name
        if not path.exists():
            problems.append(f"{fixture.name}: not generated")
            continue
        try:
            with zipfile.ZipFile(path) as archive:
                bad = archive.testzip()
                if bad is not None:
                    problems.append(f"{fixture.name}: corrupt zip entry {bad}")
                names = set(archive.namelist())
                for expected in fixture.expected_parts:
                    if expected not in names:
                        problems.append(f"{fixture.name}: missing expected part {expected}")
                overrides, defaults = _declared_content_types(archive)
                for expected in fixture.expected_parts:
                    extension = expected.rsplit(".", 1)[-1].lower()
                    if f"/{expected}" in overrides or extension in defaults:
                        continue
                    problems.append(f"{fixture.name}: {expected} has no declared content type")
                for name in names:
                    if not name.endswith((".xml", ".rels")):
                        continue
                    try:
                        ElementTree.fromstring(archive.read(name))
                    except ElementTree.ParseError as error:
                        problems.append(f"{fixture.name}: {name} is not wellformed XML ({error})")
                problems.extend(
                    f"{fixture.name}: {issue}" for issue in _dangling_relationships(archive, names)
                )
        except zipfile.BadZipFile as error:
            problems.append(f"{fixture.name}: not a zip ({error})")
    return problems


def _declared_content_types(archive: zipfile.ZipFile) -> tuple[set[str], set[str]]:
    """Return declared Override part names and Default extensions.

    A part is legally typed by either mechanism. Excel relies on the default
    `xml` extension for customXml items, so demanding an Override for every
    part would flag valid packages.
    """
    root = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    overrides = {
        element.get("PartName", "") for element in root.findall(f"{{{CT_NS}}}Override")
    }
    defaults = {
        (element.get("Extension") or "").lower()
        for element in root.findall(f"{{{CT_NS}}}Default")
    }
    return overrides, defaults


def _dangling_relationships(archive: zipfile.ZipFile, names: set[str]) -> list[str]:
    import posixpath

    problems: list[str] = []
    for name in sorted(names):
        if not name.endswith(".rels"):
            continue
        base = posixpath.dirname(posixpath.dirname(name)) or ""
        root = ElementTree.fromstring(archive.read(name))
        for element in root.findall(f"{{{PKG_REL_NS}}}Relationship"):
            if element.get("TargetMode") == "External":
                continue
            target = element.get("Target", "")
            if not target or target.startswith(("http://", "https://", "mailto:", "file:")):
                continue
            resolved = posixpath.normpath(posixpath.join(base, target)).lstrip("/")
            if resolved not in names:
                problems.append(f"{name} points at missing part {resolved}")
    return problems


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("destination", type=Path, nargs="?", default=Path("corpus"))
    parser.add_argument("--verify-only", action="store_true", help="Check an existing corpus")
    parser.add_argument("--keep", action="store_true", help="Do not wipe the destination first")
    args = parser.parse_args(argv)

    if not args.verify_only:
        written = generate(args.destination, clean=not args.keep)
        print(f"Generated {len(written)} workbooks in {args.destination}")
        for fixture in written:
            print(f"  {fixture.tier:<20} {fixture.name}")

    problems = verify(args.destination)
    if problems:
        print(f"\n{len(problems)} structural problem(s):")
        for problem in problems:
            print(f"  - {problem}")
        return 1
    print("\nStructural check passed: parts present, XML wellformed, no dangling relationships.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
