"""Benchmark WolfXL Community against other open-source Python Excel libraries.

Engines and their honest scope:

- wolfxl           full read/write (this project)
- openpyxl         full read/write (the de-facto standard, baseline)
- XlsxWriter       write-only
- PyExcelerate     write-only
- pylightxl        pure-Python read/write (no native datetime support)
- pandas           DataFrame to_excel/read_excel (openpyxl and calamine engines)
- polars           DataFrame write_excel/read_excel (wraps XlsxWriter/fastexcel)
- duckdb           SQL engine (excel extension: COPY xlsx / read_xlsx to Arrow)
- tablib           Dataset wrapper (openpyxl xlsx backend)
- pyexcel          wrapper (openpyxl backend via pyexcel-xlsx)
- python-calamine  read-only, returns Python values
- fastexcel        read-only, returns Arrow tables (not Python cell objects)
- xlsx2csv         read-only, transcodes to CSV text

Inclusion bar: xlsx-capable, no external application required, and broad real
adoption (roughly one million PyPI downloads per month). Format-specific
libraries (xlrd/xlwt for .xls, pyxlsb for .xlsb), xlwings (requires a running
Excel installation), and commercial SDKs are out of scope.

Write-only libraries appear only in write cases and read-only libraries only
in read cases. pandas, polars, and duckdb are timed from a prebuilt
DataFrame (duckdb reads a registered DataFrame view), which is their primary
use case; DataFrame construction is excluded from their timings. tablib is
timed from a prebuilt Dataset. duckdb, tablib, pyexcel, and xlsx2csv are
imported lazily inside their engine functions so the peak-RSS children of
other engines do not pay for their imports.
Read cases share one fixture written by openpyxl so no reader is parsing its
own writer's output. A separate peak-RSS pass runs each large case once per
engine in a fresh process; a grid-only baseline child shows how much of the
peak is the input data itself.

Usage:

    python benchmarks/benchmark_python_excel_ecosystem.py \
        --rounds 5 --output-dir /tmp/ecosystem-results --prefix my-run
"""

from __future__ import annotations

import argparse
import datetime as dt
import importlib.metadata
import json
import platform
import statistics
import subprocess
import sys
import tempfile
from pathlib import Path
from time import perf_counter
from typing import Any, Callable

import openpyxl
import pandas as pd
import polars as pl
import pyexcelerate
import pylightxl
import xlsxwriter
import wolfxl
from python_calamine import CalamineWorkbook

import fastexcel

BASE_DT = dt.datetime(2024, 1, 1, 9, 30, 0)


def plain_row(row_idx: int, cols: int) -> list[Any]:
    row: list[Any] = [f"r{row_idx}c0"]
    for col in range(1, cols):
        if col % 2 == 0:
            row.append(row_idx * 31 + col)
        else:
            row.append(row_idx + col / 7.0)
    return row


def mixed_row(row_idx: int) -> list[Any]:
    return [
        f"name-{row_idx}",
        row_idx * 17,
        row_idx / 3.0,
        row_idx % 2 == 0,
        BASE_DT + dt.timedelta(minutes=row_idx),
    ]


def unique_string_row(row_idx: int, cols: int) -> list[str]:
    return [f"u{row_idx}-{col}-{row_idx * 2654435761 % 100000:05d}" for col in range(cols)]


# --- write engines ---------------------------------------------------------


def write_wolfxl(path: Path, grid: list[list[Any]]) -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.write_rows(grid)
    wb.save(str(path))


def write_openpyxl(path: Path, grid: list[list[Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row in grid:
        ws.append(row)
    wb.save(str(path))


def write_openpyxl_wo(path: Path, grid: list[list[Any]]) -> None:
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    for row in grid:
        ws.append(row)
    wb.save(str(path))


def write_xlsxwriter(path: Path, grid: list[list[Any]]) -> None:
    wb = xlsxwriter.Workbook(str(path), {"default_date_format": "yyyy-mm-dd hh:mm:ss"})
    ws = wb.add_worksheet()
    for row_idx, row in enumerate(grid):
        ws.write_row(row_idx, 0, row)
    wb.close()


def write_xlsxwriter_cm(path: Path, grid: list[list[Any]]) -> None:
    wb = xlsxwriter.Workbook(
        str(path),
        {"constant_memory": True, "default_date_format": "yyyy-mm-dd hh:mm:ss"},
    )
    ws = wb.add_worksheet()
    for row_idx, row in enumerate(grid):
        ws.write_row(row_idx, 0, row)
    wb.close()


def write_pyexcelerate(path: Path, grid: list[list[Any]]) -> None:
    wb = pyexcelerate.Workbook()
    wb.new_sheet("Sheet1", data=grid)
    wb.save(str(path))


def write_pylightxl(path: Path, grid: list[list[Any]]) -> None:
    db = pylightxl.Database()
    db.add_ws("Sheet1")
    ws = db.ws("Sheet1")
    for row_idx, row in enumerate(grid, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.update_index(row_idx, col_idx, value)
    pylightxl.writexl(db, str(path))


def write_pandas(path: Path, frame: "pd.DataFrame") -> None:
    frame.to_excel(str(path), index=False, header=False, engine="openpyxl")


def write_pandas_xlsxwriter(path: Path, frame: "pd.DataFrame") -> None:
    frame.to_excel(str(path), index=False, header=False, engine="xlsxwriter")


def write_polars(path: Path, frame: "pl.DataFrame") -> None:
    frame.write_excel(str(path), include_header=False)


_DUCKDB_CON: Any = None


def _duckdb_connection() -> Any:
    global _DUCKDB_CON
    if _DUCKDB_CON is None:
        import duckdb

        _DUCKDB_CON = duckdb.connect()
        _DUCKDB_CON.execute("INSTALL excel; LOAD excel;")
    return _DUCKDB_CON


def _tablib_dataset(grid: list[list[Any]]) -> Any:
    import tablib

    dataset = tablib.Dataset()
    dataset.extend(grid)
    return dataset


def write_duckdb(con: Any, view: str, path: Path) -> None:
    if path.exists():
        path.unlink()
    con.execute(f"COPY (SELECT * FROM {view}) TO '{path}' (FORMAT xlsx)")


def write_tablib(path: Path, dataset: Any) -> None:
    path.write_bytes(dataset.export("xlsx"))


def write_pyexcel(path: Path, grid: list[list[Any]]) -> None:
    import pyexcel

    pyexcel.save_as(array=grid, dest_file_name=str(path))


# --- read engines ----------------------------------------------------------


def read_wolfxl(path: Path) -> int:
    wb = wolfxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    assert ws is not None
    count = 0
    for row in ws.iter_rows(values_only=True):
        count += len(row)
    wb.close()
    return count


def read_openpyxl(path: Path) -> int:
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    assert ws is not None
    count = 0
    for row in ws.iter_rows(values_only=True):
        count += len(row)
    wb.close()
    return count


def read_python_calamine(path: Path) -> int:
    wb = CalamineWorkbook.from_path(str(path))
    rows = wb.get_sheet_by_index(0).to_python(skip_empty_area=False)
    return sum(len(row) for row in rows)


def read_fastexcel(path: Path) -> int:
    reader = fastexcel.read_excel(str(path))
    table = reader.load_sheet(0, header_row=None).to_arrow()
    return table.num_rows * table.num_columns


def read_pylightxl(path: Path) -> int:
    db = pylightxl.readxl(str(path))
    ws = db.ws(db.ws_names[0])
    return sum(len(row) for row in ws.rows)


def read_pandas_openpyxl(path: Path) -> int:
    frame = pd.read_excel(str(path), header=None, engine="openpyxl")
    return int(frame.shape[0] * frame.shape[1])


def read_pandas_calamine(path: Path) -> int:
    frame = pd.read_excel(str(path), header=None, engine="calamine")
    return int(frame.shape[0] * frame.shape[1])


def read_polars(path: Path) -> int:
    frame = pl.read_excel(str(path), has_header=False)
    return frame.height * frame.width


def read_duckdb(path: Path) -> int:
    con = _duckdb_connection()
    table = con.execute(
        f"SELECT * FROM read_xlsx('{path}', header = false)"
    ).fetch_arrow_table()
    return table.num_rows * table.num_columns


def read_tablib(path: Path) -> int:
    import tablib

    dataset = tablib.Dataset().load(path.read_bytes(), format="xlsx", headers=False)
    return dataset.height * dataset.width


def read_pyexcel(path: Path) -> int:
    import pyexcel

    rows = pyexcel.get_array(file_name=str(path))
    count = sum(len(row) for row in rows)
    pyexcel.free_resources()
    return count


def read_xlsx2csv(path: Path) -> int:
    from xlsx2csv import Xlsx2csv

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
        csv_path = Path(tmp.name)
    try:
        Xlsx2csv(str(path), outputencoding="utf-8").convert(str(csv_path))
        with csv_path.open(encoding="utf-8") as fh:
            first = fh.readline()
            cols = first.count(",") + 1
            lines = 1 + sum(1 for _ in fh)
    finally:
        csv_path.unlink(missing_ok=True)
    return lines * cols


READ_ENGINES: dict[str, Callable[[Path], int]] = {
    "wolfxl": read_wolfxl,
    "openpyxl": read_openpyxl,
    "python_calamine": read_python_calamine,
    "fastexcel": read_fastexcel,
    "pylightxl": read_pylightxl,
    "pandas_openpyxl": read_pandas_openpyxl,
    "pandas_calamine": read_pandas_calamine,
    "polars": read_polars,
    "duckdb": read_duckdb,
    "tablib": read_tablib,
    "pyexcel": read_pyexcel,
    "xlsx2csv": read_xlsx2csv,
}

ENGINE_SCOPE = {
    "wolfxl": "full read/write",
    "openpyxl": "full read/write",
    "openpyxl_wo": "openpyxl write_only mode",
    "xlsxwriter": "write-only",
    "xlsxwriter_cm": "XlsxWriter constant_memory mode",
    "pyexcelerate": "write-only",
    "pylightxl": "pure-Python read/write",
    "pandas_openpyxl": "DataFrame I/O, openpyxl engine",
    "pandas_xlsxwriter": "DataFrame I/O, xlsxwriter engine",
    "pandas_calamine": "DataFrame read, calamine engine",
    "polars": "DataFrame I/O, wraps XlsxWriter/fastexcel",
    "python_calamine": "read-only, Python values",
    "fastexcel": "read-only, Arrow tables",
    "duckdb": "SQL engine; DataFrame in, Arrow out",
    "tablib": "Dataset wrapper, openpyxl backend",
    "pyexcel": "wrapper, openpyxl backend",
    "xlsx2csv": "read-only, transcodes to CSV text",
}

# pylightxl has no native datetime support, so it is excluded from the mixed
# case rather than measured on a workload it cannot represent. The mode
# variants (openpyxl write_only, XlsxWriter constant_memory, pandas with the
# xlsxwriter engine) are measured on the large plain write and the memory pass
# only: they exist to steelman the baselines on the workload where mode choice
# matters, not to add near-duplicate bars to every chart.
WRITE_CASE_ENGINES = {
    "write_plain_large": (
        "wolfxl",
        "openpyxl",
        "openpyxl_wo",
        "xlsxwriter",
        "xlsxwriter_cm",
        "pyexcelerate",
        "pylightxl",
        "pandas_openpyxl",
        "pandas_xlsxwriter",
        "polars",
        "duckdb",
        "tablib",
        "pyexcel",
    ),
    "write_mixed": (
        "wolfxl",
        "openpyxl",
        "xlsxwriter",
        "pyexcelerate",
        "pandas_openpyxl",
        "polars",
        "duckdb",
        "tablib",
        "pyexcel",
    ),
    "write_unique_strings": (
        "wolfxl",
        "openpyxl",
        "xlsxwriter",
        "pyexcelerate",
        "pylightxl",
        "pandas_openpyxl",
        "polars",
        "duckdb",
        "tablib",
        "pyexcel",
    ),
}

READ_CASE_ENGINES = {
    "read_values_large": (
        "wolfxl",
        "openpyxl",
        "python_calamine",
        "fastexcel",
        "pylightxl",
        "pandas_openpyxl",
        "pandas_calamine",
        "polars",
        "duckdb",
        "tablib",
        "pyexcel",
        "xlsx2csv",
    ),
}

MEMORY_WRITE_ENGINES = WRITE_CASE_ENGINES["write_plain_large"]
MEMORY_READ_ENGINES = READ_CASE_ENGINES["read_values_large"]


def build_write_engines(
    names: tuple[str, ...], grid: list[list[Any]], workdir: Path, case: str
) -> dict[str, Callable[[], None]]:
    needs_pandas = bool(
        {"pandas_openpyxl", "pandas_xlsxwriter", "duckdb"} & set(names)
    )
    pandas_frame = pd.DataFrame(grid) if needs_pandas else None
    polars_frame = pl.DataFrame(grid, orient="row") if "polars" in names else None
    engines: dict[str, Callable[[], None]] = {}
    for name in names:
        target = workdir / f"{case}-{name}.xlsx"
        if name == "pandas_openpyxl":
            engines[name] = lambda t=target, f=pandas_frame: write_pandas(t, f)
        elif name == "pandas_xlsxwriter":
            engines[name] = lambda t=target, f=pandas_frame: write_pandas_xlsxwriter(
                t, f
            )
        elif name == "polars":
            engines[name] = lambda t=target, f=polars_frame: write_polars(t, f)
        elif name == "duckdb":
            con = _duckdb_connection()
            view = f"grid_{case}"
            con.register(view, pandas_frame)
            engines[name] = lambda c=con, v=view, t=target: write_duckdb(c, v, t)
        elif name == "tablib":
            dataset = _tablib_dataset(grid)
            engines[name] = lambda t=target, d=dataset: write_tablib(t, d)
        elif name == "pyexcel":
            engines[name] = lambda t=target, g=grid: write_pyexcel(t, g)
        else:
            writer = {
                "wolfxl": write_wolfxl,
                "openpyxl": write_openpyxl,
                "openpyxl_wo": write_openpyxl_wo,
                "xlsxwriter": write_xlsxwriter,
                "xlsxwriter_cm": write_xlsxwriter_cm,
                "pyexcelerate": write_pyexcelerate,
                "pylightxl": write_pylightxl,
            }[name]
            engines[name] = lambda w=writer, t=target, g=grid: w(t, g)
    return engines


def cpu_brand() -> str:
    cpuinfo = Path("/proc/cpuinfo")
    if cpuinfo.exists():
        for line in cpuinfo.read_text(encoding="utf-8").splitlines():
            if line.lower().startswith("model name"):
                return line.split(":", 1)[1].strip()
    try:
        return subprocess.run(
            ["sysctl", "-n", "machdep.cpu.brand_string"],
            capture_output=True,
            text=True,
            check=True,
        ).stdout.strip()
    except Exception:
        return platform.processor() or "unknown"


def library_versions() -> dict[str, str]:
    versions = {}
    for dist in (
        "wolfxl",
        "openpyxl",
        "XlsxWriter",
        "PyExcelerate",
        "pylightxl",
        "pandas",
        "polars",
        "python-calamine",
        "fastexcel",
        "pyarrow",
        "duckdb",
        "tablib",
        "pyexcel",
        "pyexcel-xlsx",
        "xlsx2csv",
    ):
        try:
            versions[dist] = importlib.metadata.version(dist)
        except importlib.metadata.PackageNotFoundError:
            versions[dist] = "not installed"
    return versions


class RoundBudgetExceeded(Exception):
    """One benchmark round did not finish inside the wall-clock budget."""


_ALARM_ARMED = False


def _on_alarm(signum: int, frame: Any) -> None:
    if _ALARM_ARMED:
        raise RoundBudgetExceeded()
    # A late SIGALRM delivered after the timed window closed is ignored;
    # restoring the default handler instead would kill the whole process.


def install_round_budget_handler() -> None:
    import signal

    signal.signal(signal.SIGALRM, _on_alarm)


def measure(fn: Callable[[], Any], rounds: int, budget_seconds: int) -> list[float]:
    """Time ``rounds`` runs of ``fn`` after one discarded warmup.

    Every run (warmup included) must finish within ``budget_seconds`` of wall
    clock, enforced with SIGALRM so a pathologically slow pure-Python engine
    cannot stall the whole suite. Raises :class:`RoundBudgetExceeded`.
    ``install_round_budget_handler`` must have been called first.
    """
    import signal

    global _ALARM_ARMED
    samples = []
    for idx in range(rounds + 1):  # first run is the discarded warmup
        _ALARM_ARMED = True
        signal.alarm(budget_seconds)
        try:
            started = perf_counter()
            fn()
            elapsed = perf_counter() - started
        finally:
            _ALARM_ARMED = False
            signal.alarm(0)
        if idx:
            samples.append(elapsed)
    return samples


# --- peak-RSS child mode ----------------------------------------------------


def peak_rss_bytes() -> int:
    if sys.platform.startswith("linux"):
        # ru_maxrss is unusable here: the kernel folds the parent's resident
        # set at fork time into the child's accounting, so a child spawned by
        # a large parent inherits the parent's high-water mark even after
        # exec. VmHWM belongs to this process's own address space only.
        with open("/proc/self/status", encoding="ascii") as status:
            for line in status:
                if line.startswith("VmHWM:"):
                    return int(line.split()[1]) * 1024
        raise RuntimeError("VmHWM not found in /proc/self/status")
    import resource

    return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss  # bytes on macOS


def run_memory_child(spec: str, rows: int, cols: int, fixture: str | None) -> None:
    kind, engine = spec.split(":", 1)
    if kind == "write":
        grid = [plain_row(idx, cols) for idx in range(1, rows + 1)]
        if engine == "grid_baseline":
            pass
        else:
            with tempfile.TemporaryDirectory() as tmp:
                engines = build_write_engines(
                    (engine,), grid, Path(tmp), "memory_write"
                )
                engines[engine]()
    elif kind == "read":
        assert fixture is not None
        expected = rows * cols
        count = READ_ENGINES[engine](Path(fixture))
        if count != expected:
            raise AssertionError(f"read {count} cells, expected {expected}")
    else:
        raise ValueError(f"unknown memory child kind: {kind}")
    print(json.dumps({"peak_rss_bytes": peak_rss_bytes()}))


def run_memory_pass(
    rows: int,
    cols: int,
    fixture: Path,
    budget_seconds: int,
    engine_filter: set[str] | None = None,
) -> list[dict[str, Any]]:
    def selected(engine: str) -> bool:
        return engine_filter is None or engine in engine_filter

    records: list[dict[str, Any]] = []
    specs: list[tuple[str, str, str | None]] = []
    if selected("grid_baseline"):
        specs.append(("memory_write_large", "write:grid_baseline", None))
    specs += [
        ("memory_write_large", f"write:{e}", None)
        for e in MEMORY_WRITE_ENGINES
        if selected(e)
    ]
    specs += [
        ("memory_read_large", f"read:{e}", str(fixture))
        for e in MEMORY_READ_ENGINES
        if selected(e)
    ]
    for case, spec, fixture_arg in specs:
        engine = spec.split(":", 1)[1]
        print(f"[{case}] {engine} ...", flush=True)
        cmd = [
            sys.executable,
            __file__,
            "--memory-child",
            spec,
            "--rows",
            str(rows),
            "--cols",
            str(cols),
        ]
        if fixture_arg:
            cmd += ["--fixture", fixture_arg]
        base = {
            "case": case,
            "engine": engine,
            "scope": ENGINE_SCOPE.get(engine, "input grid only, no Excel library"),
            "rows": rows,
            "cols": cols,
        }
        try:
            proc = subprocess.run(
                cmd, capture_output=True, text=True, check=True,
                timeout=budget_seconds,
            )
        except subprocess.TimeoutExpired:
            print(
                f"[{case}] {engine} DID NOT FINISH within {budget_seconds} s",
                flush=True,
            )
            records.append(
                base | {"timed_out": True, "round_budget_seconds": budget_seconds}
            )
            continue
        payload = json.loads(proc.stdout.strip().splitlines()[-1])
        records.append(base | {"peak_rss_bytes": payload["peak_rss_bytes"]})
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--rows", type=int, default=200_000)
    parser.add_argument("--cols", type=int, default=8)
    parser.add_argument("--mixed-rows", type=int, default=10_000)
    parser.add_argument("--string-rows", type=int, default=100_000)
    parser.add_argument("--string-cols", type=int, default=5)
    parser.add_argument("--rounds", type=int, default=5)
    parser.add_argument(
        "--round-budget",
        type=int,
        default=240,
        help="max wall-clock seconds per round; slower engines are recorded as DNF",
    )
    parser.add_argument("--output-dir", type=Path, default=Path("/tmp/ecosystem-results"))
    parser.add_argument("--prefix", default="ecosystem")
    parser.add_argument("--case", default=None, help="comma-separated case filter")
    parser.add_argument(
        "--engines",
        default=None,
        help="comma-separated engine filter for incremental runs; results must "
        "be merged with merge_ecosystem_results.py, which enforces an "
        "identical environment",
    )
    parser.add_argument("--no-memory", action="store_true", help="skip the peak-RSS pass")
    parser.add_argument("--memory-child", default=None, help=argparse.SUPPRESS)
    parser.add_argument("--fixture", default=None, help=argparse.SUPPRESS)
    args = parser.parse_args()

    if args.memory_child:
        run_memory_child(args.memory_child, args.rows, args.cols, args.fixture)
        return

    case_filter = set(args.case.split(",")) if args.case else None
    engine_filter = set(args.engines.split(",")) if args.engines else None

    def include(case: str) -> bool:
        return case_filter is None or case in case_filter

    args.output_dir.mkdir(parents=True, exist_ok=True)
    workdir = Path(tempfile.mkdtemp(prefix="wolfxl-ecosystem-"))

    plain_grid = [plain_row(idx, args.cols) for idx in range(1, args.rows + 1)]

    cases: list[dict[str, Any]] = []
    if include("write_plain_large"):
        cases.append(
            {
                "case": "write_plain_large",
                "kind": "write",
                "rows": args.rows,
                "cols": args.cols,
                "grid": plain_grid,
            }
        )
    if include("write_mixed"):
        cases.append(
            {
                "case": "write_mixed",
                "kind": "write",
                "rows": args.mixed_rows,
                "cols": 5,
                "grid": [mixed_row(idx) for idx in range(1, args.mixed_rows + 1)],
            }
        )
    if include("write_unique_strings"):
        cases.append(
            {
                "case": "write_unique_strings",
                "kind": "write",
                "rows": args.string_rows,
                "cols": args.string_cols,
                "grid": [
                    unique_string_row(idx, args.string_cols)
                    for idx in range(1, args.string_rows + 1)
                ],
            }
        )

    fixture = workdir / "read-fixture.xlsx"
    needs_fixture = include("read_values_large") or not args.no_memory
    if needs_fixture:
        write_openpyxl(fixture, plain_grid)
    if include("read_values_large"):
        cases.append(
            {
                "case": "read_values_large",
                "kind": "read",
                "rows": args.rows,
                "cols": args.cols,
                "fixture": fixture,
            }
        )

    install_round_budget_handler()
    results: list[dict[str, Any]] = []
    def selected_engines(names: tuple[str, ...]) -> tuple[str, ...]:
        if engine_filter is None:
            return names
        return tuple(n for n in names if n in engine_filter)

    for spec in cases:
        units = spec["rows"] * spec["cols"]
        if spec["kind"] == "write":
            engines = build_write_engines(
                selected_engines(WRITE_CASE_ENGINES[spec["case"]]),
                spec["grid"],
                workdir,
                spec["case"],
            )
        else:
            expected = units
            engines = {}
            for name in selected_engines(READ_CASE_ENGINES[spec["case"]]):

                def run_read(r=READ_ENGINES[name], f=spec["fixture"], e=expected) -> None:
                    count = r(f)
                    if count != e:
                        raise AssertionError(f"read {count} cells, expected {e}")

                engines[name] = run_read

        for name, fn in engines.items():
            print(f"[{spec['case']}] {name} ...", flush=True)
            base = {
                "case": spec["case"],
                "engine": name,
                "scope": ENGINE_SCOPE[name],
                "rows": spec["rows"],
                "cols": spec["cols"],
                "units": units,
            }
            try:
                samples = measure(fn, args.rounds, args.round_budget)
            except RoundBudgetExceeded:
                print(
                    f"[{spec['case']}] {name} DID NOT FINISH one round "
                    f"within {args.round_budget} s",
                    flush=True,
                )
                results.append(
                    base
                    | {
                        "timed_out": True,
                        "round_budget_seconds": args.round_budget,
                    }
                )
                continue
            median = statistics.median(samples)
            results.append(
                base
                | {
                    "samples_seconds": [round(s, 6) for s in samples],
                    "median_seconds": round(median, 6),
                    "units_per_second": round(units / median, 2) if median else None,
                }
            )

    memory_records: list[dict[str, Any]] = []
    if not args.no_memory:
        memory_records = run_memory_pass(
            args.rows, args.cols, fixture, args.round_budget, engine_filter
        )

    comparisons = []
    by_case: dict[str, dict[str, float]] = {}
    for row in results:
        if row.get("timed_out"):
            continue
        by_case.setdefault(row["case"], {})[row["engine"]] = row["median_seconds"]
    for case, engines_medians in sorted(by_case.items()):
        baseline = engines_medians.get("openpyxl")
        if not baseline:
            continue
        for engine, median in sorted(engines_medians.items()):
            if engine == "openpyxl":
                continue
            comparisons.append(
                {
                    "case": case,
                    "engine": engine,
                    "speedup_vs_openpyxl": round(baseline / median, 4),
                }
            )

    payload = {
        "metadata": {
            "timestamp_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
            "python": platform.python_version(),
            "platform": platform.platform(),
            "machine": platform.machine(),
            "cpu_brand": cpu_brand(),
            "rounds": args.rounds,
            "warmup_rounds": 1,
            "round_budget_seconds": args.round_budget,
            "read_fixture_writer": "openpyxl",
            "dataframe_timing": "pandas/polars/duckdb timed from a prebuilt DataFrame; tablib from a prebuilt Dataset",
            "memory_pass": "one fresh process per engine, single run, peak RSS",
            "versions": library_versions(),
            "engine_scope": ENGINE_SCOPE,
        },
        "results": results,
        "memory": memory_records,
        "comparisons": comparisons,
    }

    json_path = args.output_dir / f"{args.prefix}.json"
    json_path.write_text(json.dumps(payload, indent=1) + "\n", encoding="utf-8")

    lines = ["# Python Excel ecosystem benchmark", ""]
    meta = payload["metadata"]
    lines.append(
        f"{meta['cpu_brand']} | Python {meta['python']} | median of {meta['rounds']} rounds"
    )
    lines.append("")
    for case in by_case:
        lines.append(f"## {case}")
        lines.append("")
        lines.append("| engine | scope | median s | units/s |")
        lines.append("|---|---|---|---|")
        for row in results:
            if row["case"] != case:
                continue
            if row.get("timed_out"):
                lines.append(
                    f"| {row['engine']} | {row['scope']} | DNF "
                    f"(> {row['round_budget_seconds']} s/round) | - |"
                )
                continue
            lines.append(
                f"| {row['engine']} | {row['scope']} | {row['median_seconds']:.4f} "
                f"| {row['units_per_second']:,.0f} |"
            )
        lines.append("")
    if memory_records:
        for case in ("memory_write_large", "memory_read_large"):
            rows_for_case = [r for r in memory_records if r["case"] == case]
            if not rows_for_case:
                continue
            lines.append(f"## {case}")
            lines.append("")
            lines.append("| engine | peak RSS MiB |")
            lines.append("|---|---|")
            for row in rows_for_case:
                if row.get("timed_out"):
                    lines.append(
                        f"| {row['engine']} | DNF "
                        f"(> {row['round_budget_seconds']} s) |"
                    )
                    continue
                lines.append(
                    f"| {row['engine']} | {row['peak_rss_bytes'] / 1048576:,.0f} |"
                )
            lines.append("")
    md_path = args.output_dir / f"{args.prefix}.md"
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {json_path}")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    sys.exit(main())
