#!/usr/bin/env python3
"""Category-boundary benchmark for WolfXL's unqualified writer-speed claim.

This intentionally compares each library's fastest natural bulk input path. The values
and XLSX dimensions match, but the input containers differ: WolfXL receives Python rows
and Jetxl receives an Arrow table. It is therefore not an API-equivalence benchmark.
"""

from __future__ import annotations

import argparse
import gc
import importlib.metadata
import json
import platform
import statistics
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from time import perf_counter
from typing import Any, Callable

import jetxl
import openpyxl
import pyarrow as pa
import wolfxl


def make_row(index: int) -> list[Any]:
    return [
        f"acct-{index % 997}",
        index,
        index * 1.25,
        index % 2 == 0,
        f"region-{index % 8}",
        index % 31,
        index / 7.0,
        index * 17,
    ]


def cpu_brand() -> str:
    cpuinfo = Path("/proc/cpuinfo")
    if cpuinfo.exists():
        for line in cpuinfo.read_text(encoding="utf-8").splitlines():
            if line.lower().startswith("model name"):
                return line.split(":", 1)[1].strip()
    return platform.processor() or "unknown"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=200_000)
    parser.add_argument("--cols", type=int, default=8)
    parser.add_argument("--rounds", type=int, default=5)
    parser.add_argument("--output-dir", type=Path, default=Path("audit-results"))
    args = parser.parse_args()

    if args.cols != 8:
        raise ValueError("this benchmark currently defines exactly eight mixed columns")

    grid = [make_row(index) for index in range(1, args.rows + 1)]
    columns = {
        f"c{column}": [record[column] for record in grid]
        for column in range(args.cols)
    }
    table = pa.table(columns)

    with tempfile.TemporaryDirectory(prefix="wolfxl-sota-") as temp_dir:
        root = Path(temp_dir)
        targets = {
            "wolfxl_grid": root / "wolfxl.xlsx",
            "jetxl_arrow": root / "jetxl.xlsx",
        }

        def write_wolfxl() -> None:
            path = targets["wolfxl_grid"]
            path.unlink(missing_ok=True)
            workbook = wolfxl.Workbook()
            worksheet = workbook.active
            worksheet.write_rows(grid, copy=False, plain_values_only=True)
            workbook.save(str(path))
            workbook.close()

        def write_jetxl() -> None:
            path = targets["jetxl_arrow"]
            path.unlink(missing_ok=True)
            jetxl.write_sheet_arrow(table, str(path), write_header_row=False)

        writers: dict[str, Callable[[], None]] = {
            "wolfxl_grid": write_wolfxl,
            "jetxl_arrow": write_jetxl,
        }

        for writer in writers.values():
            writer()

        samples: dict[str, list[float]] = {name: [] for name in writers}
        for round_index in range(args.rounds):
            order = (
                ("wolfxl_grid", "jetxl_arrow")
                if round_index % 2 == 0
                else ("jetxl_arrow", "wolfxl_grid")
            )
            for name in order:
                gc.collect()
                enabled = gc.isenabled()
                if enabled:
                    gc.disable()
                started = perf_counter()
                try:
                    writers[name]()
                finally:
                    samples[name].append(perf_counter() - started)
                    if enabled:
                        gc.enable()

        validations: dict[str, dict[str, int]] = {}
        for name, path in targets.items():
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
            worksheet = workbook.active
            validations[name] = {
                "max_row": int(worksheet.max_row or 0),
                "max_column": int(worksheet.max_column or 0),
                "output_bytes": path.stat().st_size,
            }
            workbook.close()
            if validations[name]["max_row"] < args.rows:
                raise AssertionError(f"{name} emitted too few rows: {validations[name]}")
            if validations[name]["max_column"] < args.cols:
                raise AssertionError(f"{name} emitted too few columns: {validations[name]}")

        results: list[dict[str, Any]] = []
        for name in writers:
            values = samples[name]
            results.append(
                {
                    "engine": name,
                    "input_shape": (
                        "prebuilt Python list-of-rows"
                        if name == "wolfxl_grid"
                        else "prebuilt Arrow table"
                    ),
                    "samples_seconds": values,
                    "median_seconds": statistics.median(values),
                    "min_seconds": min(values),
                    "max_seconds": max(values),
                    "validation": validations[name],
                }
            )

        by_name = {result["engine"]: result for result in results}
        jetxl_speed_vs_wolfxl = (
            by_name["wolfxl_grid"]["median_seconds"]
            / by_name["jetxl_arrow"]["median_seconds"]
        )
        metadata = {
            "timestamp_utc": datetime.now(timezone.utc).isoformat(),
            "python": platform.python_version(),
            "platform": platform.platform(),
            "cpu_brand": cpu_brand(),
            "rows": args.rows,
            "cols": args.cols,
            "rounds": args.rounds,
            "timing_boundary": "prebuilt natural input object to closed XLSX",
            "comparability_note": (
                "The values and output dimensions match, but the input containers differ. "
                "This tests the boundary of an unqualified fastest-library claim; it is "
                "not an openpyxl-API equivalence benchmark."
            ),
            "versions": {
                package: importlib.metadata.version(package)
                for package in ("wolfxl", "jetxl", "pyarrow", "openpyxl")
            },
        }
        payload = {
            "metadata": metadata,
            "jetxl_speed_vs_wolfxl": jetxl_speed_vs_wolfxl,
            "winner": "jetxl_arrow" if jetxl_speed_vs_wolfxl > 1.0 else "wolfxl_grid",
            "results": results,
        }

        args.output_dir.mkdir(parents=True, exist_ok=True)
        (args.output_dir / "fresh-jetxl-boundary.json").write_text(
            json.dumps(payload, indent=2) + "\n",
            encoding="utf-8",
        )
        lines = [
            "# WolfXL vs Jetxl category-boundary benchmark",
            "",
            f"- CPU: `{metadata['cpu_brand']}`",
            f"- Workload: `{args.rows:,} x {args.cols}` mixed primitive cells",
            f"- Timing boundary: {metadata['timing_boundary']}",
            f"- Caveat: {metadata['comparability_note']}",
            "",
            "| Engine | Natural input | Median s | Min-Max s | Output dimensions |",
            "|---|---|---:|---:|---:|",
        ]
        for result in results:
            validation = result["validation"]
            lines.append(
                "| {engine} | {shape} | {median:.6f} | {minimum:.6f}-{maximum:.6f} | "
                "{rows}x{cols} |".format(
                    engine=result["engine"],
                    shape=result["input_shape"],
                    median=result["median_seconds"],
                    minimum=result["min_seconds"],
                    maximum=result["max_seconds"],
                    rows=validation["max_row"],
                    cols=validation["max_column"],
                )
            )
        lines.extend(
            [
                "",
                f"Jetxl speed relative to WolfXL: **{jetxl_speed_vs_wolfxl:.2f}x** "
                "(>1 means Jetxl is faster).",
                "",
            ]
        )
        (args.output_dir / "fresh-jetxl-boundary.md").write_text(
            "\n".join(lines),
            encoding="utf-8",
        )
        print(json.dumps(payload, indent=2))


if __name__ == "__main__":
    main()
