from __future__ import annotations

import argparse
import gc
import json
import os
import platform
import shutil
import statistics
import subprocess
import sys
import tempfile

try:
    import resource
except ImportError:  # pragma: no cover - Windows lacks the resource module
    resource = None  # type: ignore[assignment]
from collections.abc import Callable
from datetime import datetime, timezone
from pathlib import Path
from time import perf_counter
from typing import Any

import openpyxl
import wolfxl
from openpyxl.styles import Font as OpenpyxlFont
from openpyxl.styles import PatternFill as OpenpyxlPatternFill
from wolfxl.styles import Font as WolfxlFont
from wolfxl.styles import PatternFill as WolfxlPatternFill


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "docs" / "performance" / "baselines"


def plain_row(row_idx: int, cols: int) -> list[Any]:
    return [
        row_idx,
        f"customer-{row_idx % 997}",
        row_idx * 1.25,
        row_idx % 17,
        f"region-{row_idx % 4}",
    ][:cols]


def unique_string_row(row_idx: int, cols: int) -> list[Any]:
    return [
        row_idx,
        f"customer-{row_idx:08d}",
        f"invoice-{row_idx:08d}-{row_idx % 97:02d}",
        f"region-{row_idx:08d}-{row_idx % 4}",
        f"memo-{row_idx:08d}-{row_idx * 17}",
    ][:cols]


def plain_numeric_checksum(rows: int, cols: int) -> float:
    """Checksum every numeric value emitted by ``plain_row``."""
    checksum = 0.0
    for row_idx in range(1, rows + 1):
        for value in plain_row(row_idx, cols):
            value_type = type(value)
            if value_type is int or value_type is float:
                checksum += float(value)
    return checksum


def row_numeric_checksum(row: Any) -> float:
    """Checksum numeric values in an already materialized row."""
    checksum = 0.0
    for value in row:
        value_type = type(value)
        if value_type is int or value_type is float:
            checksum += float(value)
    return checksum


def mixed_row(row_idx: int, cols: int) -> list[Any]:
    row: list[Any] = [
        row_idx,
        f"customer-{row_idx % 997}",
        row_idx % 2 == 0,
        f"=A{row_idx + 1}*2",
        "#N/A" if row_idx % 23 == 0 else f"status-{row_idx % 5}",
        None if row_idx % 7 == 0 else row_idx * 0.75,
    ]
    return row[:cols]


def wide_row(row_idx: int, cols: int) -> list[Any]:
    row: list[Any] = []
    for col_idx in range(1, cols + 1):
        row.append(f"tag-{row_idx % 31}-{col_idx}" if col_idx % 5 == 0 else row_idx * col_idx)
    return row


def formula_row(row_idx: int, cols: int) -> list[Any]:
    row: list[Any] = [
        row_idx,
        row_idx * 2,
        f"=A{row_idx}+B{row_idx}",
        f"=SUM(A{row_idx}:B{row_idx})",
        f"region-{row_idx % 4}",
        f"=IF(A{row_idx}>0,\"ok\",\"bad\")",
    ]
    return row[:cols]


def time_samples(
    fn: Callable[[], Any], rounds: int, warmup: bool = True
) -> tuple[list[float], Any]:
    samples: list[float] = []
    last_result: Any = None
    if warmup:
        fn()
    for _ in range(rounds):
        gc.collect()
        gc_was_enabled = gc.isenabled()
        if gc_was_enabled:
            gc.disable()
        started = perf_counter()
        try:
            last_result = fn()
            samples.append(perf_counter() - started)
        finally:
            if gc_was_enabled:
                gc.enable()
    return samples, last_result


def build_openpyxl_plain(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(plain_row(row_idx, cols))
    wb.save(path)
    wb.close()
    return rows * cols


def build_wolfxl_append_plain(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(plain_row(row_idx, cols))
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_write_rows_plain(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_rows(grid)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_write_rows_nocopy_plain(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_rows(grid, copy=False)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_write_rows_fast_plain(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_rows(grid, copy=False, plain_values_only=True)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_write_rows_fast_unique_strings(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    grid = [unique_string_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_rows(grid, copy=False, plain_values_only=True)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_openpyxl_mixed(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(mixed_row(row_idx, cols))
    wb.save(path)
    wb.close()
    return rows * cols


def build_wolfxl_append_mixed(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(mixed_row(row_idx, cols))
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_openpyxl_wide(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(wide_row(row_idx, cols))
    wb.save(path)
    wb.close()
    return rows * cols


def build_wolfxl_wide(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(wide_row(row_idx, cols))
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_openpyxl_formula(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(formula_row(row_idx, cols))
    wb.save(path)
    wb.close()
    return rows * cols


def build_wolfxl_formula(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        ws.append(formula_row(row_idx, cols))
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_formula_rows_fast(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    grid = [formula_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_rows(grid, copy=False, plain_values_only=True)
    wb.save(str(path))
    wb.close()
    return rows * cols


def _openpyxl_style_cell(cell: Any, row_idx: int, col_idx: int) -> None:
    if col_idx == 1:
        cell.font = OpenpyxlFont(bold=True)
    if col_idx == 2 and row_idx % 2 == 0:
        cell.fill = OpenpyxlPatternFill(patternType="solid", fgColor="FFD966")
    if col_idx == 3:
        cell.number_format = "#,##0.00"


def _wolfxl_style_cell(cell: Any, row_idx: int, col_idx: int) -> None:
    if col_idx == 1:
        cell.font = WolfxlFont(bold=True)
    if col_idx == 2 and row_idx % 2 == 0:
        cell.fill = WolfxlPatternFill(patternType="solid", fgColor="FFD966")
    if col_idx == 3:
        cell.number_format = "#,##0.00"


def build_openpyxl_styled(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _openpyxl_style_cell(cell, row_idx, col_idx)
    wb.save(path)
    wb.close()
    return rows * cols


def build_openpyxl_styled_phased(path: Path, rows: int, cols: int) -> dict[str, Any]:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    started = perf_counter()
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _openpyxl_style_cell(cell, row_idx, col_idx)
    populate_seconds = perf_counter() - started

    started = perf_counter()
    wb.save(path)
    save_seconds = perf_counter() - started
    wb.close()

    return {
        "units": rows * cols,
        "phase_seconds": {
            "populate_cells": populate_seconds,
            "save": save_seconds,
        },
    }


def build_wolfxl_styled(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _wolfxl_style_cell(cell, row_idx, col_idx)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_wolfxl_styled_phased(path: Path, rows: int, cols: int) -> dict[str, Any]:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None

    started = perf_counter()
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _wolfxl_style_cell(cell, row_idx, col_idx)
    populate_seconds = perf_counter() - started

    started = perf_counter()
    wb.save(str(path))
    save_seconds = perf_counter() - started
    wb.close()

    return {
        "units": rows * cols,
        "phase_seconds": {
            "populate_cells": populate_seconds,
            "save": save_seconds,
        },
    }


def _wolfxl_styled_rows_style(row_idx: int, col_idx: int, _value: Any) -> dict[str, Any] | None:
    if col_idx == 1:
        return {"font": WolfxlFont(bold=True)}
    if col_idx == 2 and row_idx % 2 == 0:
        return {"fill": WolfxlPatternFill(patternType="solid", fgColor="FFD966")}
    if col_idx == 3:
        return {"number_format": "#,##0.00"}
    return None


_WOLFXL_STYLED_ROWS_GRID_BOLD = {"bold": True}
_WOLFXL_STYLED_ROWS_GRID_FILL = {"bg_color": "FFD966"}
_WOLFXL_STYLED_ROWS_GRID_NUMBER = {"number_format": "#,##0.00"}


def _wolfxl_styled_rows_grid_styles(rows: int, cols: int) -> list[list[dict[str, Any] | None]]:
    odd_row = [
        _WOLFXL_STYLED_ROWS_GRID_BOLD
        if col_idx == 1
        else _WOLFXL_STYLED_ROWS_GRID_NUMBER
        if col_idx == 3
        else None
        for col_idx in range(1, cols + 1)
    ]
    even_row = [
        _WOLFXL_STYLED_ROWS_GRID_BOLD
        if col_idx == 1
        else _WOLFXL_STYLED_ROWS_GRID_FILL
        if col_idx == 2
        else _WOLFXL_STYLED_ROWS_GRID_NUMBER
        if col_idx == 3
        else None
        for col_idx in range(1, cols + 1)
    ]
    return [even_row if row_idx % 2 == 0 else odd_row for row_idx in range(1, rows + 1)]


def build_wolfxl_styled_rows_phased(path: Path, rows: int, cols: int) -> dict[str, Any]:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None

    started = perf_counter()
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    ws.write_styled_rows(
        grid,
        _wolfxl_styled_rows_style,
        copy=False,
        plain_values_only=True,
    )
    populate_seconds = perf_counter() - started

    started = perf_counter()
    wb.save(str(path))
    save_seconds = perf_counter() - started
    wb.close()

    return {
        "units": rows * cols,
        "phase_seconds": {
            "populate_cells": populate_seconds,
            "save": save_seconds,
        },
    }


def build_wolfxl_styled_rows_grid_phased(
    path: Path,
    rows: int,
    cols: int,
) -> dict[str, Any]:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None

    started = perf_counter()
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    styles = _wolfxl_styled_rows_grid_styles(rows, cols)
    ws.write_styled_rows(
        grid,
        styles,
        copy=False,
        plain_values_only=True,
        normalized_styles=True,
    )
    populate_seconds = perf_counter() - started

    started = perf_counter()
    wb.save(str(path))
    save_seconds = perf_counter() - started
    wb.close()

    return {
        "units": rows * cols,
        "phase_seconds": {
            "populate_cells": populate_seconds,
            "save": save_seconds,
        },
    }


def build_openpyxl_multi_sheet(path: Path, rows: int, cols: int, sheet_count: int) -> int:
    wb = openpyxl.Workbook()
    for sheet_idx in range(1, sheet_count + 1):
        if sheet_idx == 1:
            ws = wb.active
            assert ws is not None
            ws.title = f"Data {sheet_idx}"
        else:
            ws = wb.create_sheet(f"Data {sheet_idx}")
        for row_idx in range(1, rows + 1):
            ws.append(plain_row(row_idx, cols))
    wb.save(path)
    wb.close()
    return rows * cols * sheet_count


def build_wolfxl_multi_sheet(path: Path, rows: int, cols: int, sheet_count: int) -> int:
    wb = wolfxl.Workbook()
    for sheet_idx in range(1, sheet_count + 1):
        if sheet_idx == 1:
            ws = wb.active
            assert ws is not None
            ws.title = f"Data {sheet_idx}"
        else:
            ws = wb.create_sheet(f"Data {sheet_idx}")
        for row_idx in range(1, rows + 1):
            ws.append(plain_row(row_idx, cols))
    wb.save(str(path))
    wb.close()
    return rows * cols * sheet_count


def build_wolfxl_multi_sheet_write_rows_fast(
    path: Path,
    rows: int,
    cols: int,
    sheet_count: int,
) -> int:
    wb = wolfxl.Workbook()
    grid = [plain_row(row_idx, cols) for row_idx in range(1, rows + 1)]
    for sheet_idx in range(1, sheet_count + 1):
        if sheet_idx == 1:
            ws = wb.active
            assert ws is not None
            ws.title = f"Data {sheet_idx}"
        else:
            ws = wb.create_sheet(f"Data {sheet_idx}")
        ws.write_rows(grid, copy=False, plain_values_only=True)
    wb.save(str(path))
    wb.close()
    return rows * cols * sheet_count


def build_openpyxl_cell_by_cell(path: Path, rows: int, cols: int) -> int:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(path)
    wb.close()
    return rows * cols


def build_wolfxl_cell_by_cell(path: Path, rows: int, cols: int) -> int:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(str(path))
    wb.close()
    return rows * cols


def build_openpyxl_cell_by_cell_phased(path: Path, rows: int, cols: int) -> dict[str, Any]:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    started = perf_counter()
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    construct_seconds = perf_counter() - started
    started = perf_counter()
    wb.save(path)
    save_seconds = perf_counter() - started
    wb.close()
    return {
        "units": rows * cols,
        "phase_seconds": {
            "construct_loop": construct_seconds,
            "save_flush": save_seconds,
        },
    }


def build_wolfxl_cell_by_cell_phased(path: Path, rows: int, cols: int) -> dict[str, Any]:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    started = perf_counter()
    for row_idx in range(1, rows + 1):
        for col_idx, value in enumerate(plain_row(row_idx, cols), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    construct_seconds = perf_counter() - started
    started = perf_counter()
    wb.save(str(path))
    save_seconds = perf_counter() - started
    wb.close()
    return {
        "units": rows * cols,
        "phase_seconds": {
            "construct_loop": construct_seconds,
            "save_flush": save_seconds,
        },
    }


def read_openpyxl_values(path: Path) -> tuple[int, float]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        checksum += row_numeric_checksum(row)
    wb.close()
    return rows, checksum


def read_openpyxl_read_only_values(path: Path) -> tuple[int, float]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        checksum += row_numeric_checksum(row)
    wb.close()
    return rows, checksum


def read_wolfxl_values(path: Path) -> tuple[int, float]:
    wb = wolfxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        checksum += row_numeric_checksum(row)
    wb.close()
    return rows, checksum


def read_wolfxl_read_only_values(path: Path) -> tuple[int, float]:
    wb = wolfxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        checksum += row_numeric_checksum(row)
    wb.close()
    return rows, checksum


def read_wolfxl_read_only_value_chunks(path: Path) -> tuple[int, float]:
    wb = wolfxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    for chunk in ws.iter_value_chunks(chunk_size=1024):
        rows += len(chunk)
        for row in chunk:
            checksum += row_numeric_checksum(row)
    wb.close()
    return rows, checksum


def read_wolfxl_read_only_value_chunks_phased(path: Path) -> dict[str, Any]:
    started = perf_counter()
    wb = wolfxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    open_seconds = perf_counter() - started

    started = perf_counter()
    rows = 0
    checksum = 0.0
    for chunk in ws.iter_value_chunks(chunk_size=1024):
        rows += len(chunk)
        for row in chunk:
            checksum += row_numeric_checksum(row)
    iter_seconds = perf_counter() - started
    wb.close()
    return {
        "rows": rows,
        "checksum": checksum,
        "phase_seconds": {
            "open_workbook": open_seconds,
            "iterate_values": iter_seconds,
        },
    }


def read_wolfxl_read_only_values_phased(path: Path) -> dict[str, Any]:
    started = perf_counter()
    wb = wolfxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    open_seconds = perf_counter() - started

    started = perf_counter()
    rows = 0
    checksum = 0.0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        checksum += row_numeric_checksum(row)
    iter_seconds = perf_counter() - started
    wb.close()
    return {
        "rows": rows,
        "checksum": checksum,
        "phase_seconds": {
            "open_workbook": open_seconds,
            "iterate_values": iter_seconds,
        },
    }


def read_openpyxl_wide_values(path: Path) -> tuple[int, float, int]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    cells = 0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        for value in row:
            if value is not None:
                cells += 1
            if isinstance(value, (int, float)):
                checksum += float(value)
    wb.close()
    return rows, checksum, cells


def read_wolfxl_wide_values(path: Path) -> tuple[int, float, int]:
    wb = wolfxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    cells = 0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        for value in row:
            if value is not None:
                cells += 1
            if isinstance(value, (int, float)):
                checksum += float(value)
    wb.close()
    return rows, checksum, cells


def read_openpyxl_formulas(path: Path) -> tuple[int, float, int]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb.active
    rows = 0
    checksum = 0.0
    formula_cells = 0
    for row in ws.iter_rows():
        rows += 1
        first_value = row[0].value if row else None
        if isinstance(first_value, (int, float)):
            checksum += float(first_value)
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formula_cells += 1
    wb.close()
    return rows, checksum, formula_cells


def read_wolfxl_formulas(path: Path) -> tuple[int, float, int]:
    wb = wolfxl.load_workbook(str(path), data_only=False)
    ws = wb.active
    rows = 0
    checksum = 0.0
    formula_cells = 0
    for row in ws.iter_rows(values_only=True):
        rows += 1
        first_value = row[0] if row else None
        if isinstance(first_value, (int, float)):
            checksum += float(first_value)
        for value in row:
            if isinstance(value, str) and value.startswith("="):
                formula_cells += 1
    wb.close()
    return rows, checksum, formula_cells


def read_openpyxl_styled_cells(path: Path) -> tuple[int, float, int]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    styled_cells = 0
    for row in ws.iter_rows():
        rows += 1
        first_value = row[0].value if row else None
        if isinstance(first_value, (int, float)):
            checksum += float(first_value)
        for cell in row:
            fill_kind = getattr(cell.fill, "fill_type", None) or getattr(cell.fill, "patternType", None)
            if cell.font.bold or fill_kind or cell.number_format != "General":
                styled_cells += 1
    wb.close()
    return rows, checksum, styled_cells


def read_wolfxl_styled_cells(path: Path) -> tuple[int, float, int]:
    wb = wolfxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = 0
    checksum = 0.0
    styled_cells = 0
    for row in ws.iter_rows():
        rows += 1
        first_value = row[0].value if row else None
        if isinstance(first_value, (int, float)):
            checksum += float(first_value)
        for cell in row:
            fill_kind = getattr(cell.fill, "fill_type", None) or getattr(cell.fill, "patternType", None)
            if cell.font.bold or fill_kind or cell.number_format != "General":
                styled_cells += 1
    wb.close()
    return rows, checksum, styled_cells


def read_openpyxl_multi_sheet(path: Path) -> tuple[int, float]:
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    rows = 0
    checksum = 0.0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            rows += 1
            if row and isinstance(row[0], (int, float)):
                checksum += float(row[0])
    wb.close()
    return rows, checksum


def read_wolfxl_multi_sheet(path: Path) -> tuple[int, float]:
    wb = wolfxl.load_workbook(str(path), data_only=True)
    rows = 0
    checksum = 0.0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            rows += 1
            if row and isinstance(row[0], (int, float)):
                checksum += float(row[0])
    wb.close()
    return rows, checksum


def modify_openpyxl(path: Path) -> str:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws["B2"] = "changed"
    ws["C3"] = 12345
    wb.save(path)
    wb.close()
    return "changed"


def modify_wolfxl(path: Path) -> str:
    wb = wolfxl.load_workbook(str(path), modify=True)
    ws = wb.active
    ws["B2"] = "changed"
    ws["C3"] = 12345
    wb.save(str(path))
    wb.close()
    return "changed"


# ---------------------------------------------------------------------------
# Peak-RSS instrumentation
#
# Peak RSS (``ru_maxrss``) is a per-process high-water mark: once a workload
# touches its peak, the number never drops for the life of the process. To get
# a clean per-(engine, workload) figure we run each measured workload in its
# own subprocess so its high-water mark is not polluted by earlier workloads.
# Uses only the stdlib ``resource`` module — no psutil dependency.
# ---------------------------------------------------------------------------


def maxrss_bytes() -> int:
    """Peak resident-set size of this process, normalized to bytes.

    macOS reports ``ru_maxrss`` in bytes; Linux reports it in kibibytes.
    Returns -1 when the ``resource`` module is unavailable (Windows).
    """
    if resource is None:
        return -1
    raw = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if sys.platform == "darwin":
        return int(raw)
    return int(raw) * 1024


def _rss_workload(case: str, engine: str, path: Path, rows: int, cols: int, fixture: Path | None) -> None:
    """Execute exactly one workload for the RSS subprocess worker."""
    base = case[len("large_") :] if case.startswith("large_") else case
    if base == "write_append_plain":
        (build_openpyxl_plain if engine == "openpyxl" else build_wolfxl_append_plain)(path, rows, cols)
    elif base == "read_values_plain":
        assert fixture is not None
        (read_openpyxl_values if engine == "openpyxl" else read_wolfxl_values)(fixture)
    elif base == "read_only_values_plain":
        assert fixture is not None
        (read_openpyxl_read_only_values if engine == "openpyxl" else read_wolfxl_read_only_values)(fixture)
    elif base in ("modify_two_cells_plain", "modify"):
        assert fixture is not None
        shutil.copy2(fixture, path)
        (modify_openpyxl if engine == "openpyxl" else modify_wolfxl)(path)
    else:
        raise ValueError(f"unknown RSS workload case: {case}")


def run_rss_worker(args: argparse.Namespace) -> None:
    """Subprocess entrypoint: run one workload and emit its peak RSS as JSON."""
    fixture = (
        Path(args.rss_worker_fixture)
        if args.rss_worker_fixture and args.rss_worker_fixture != "-"
        else None
    )
    baseline = maxrss_bytes()
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / "rss-target.xlsx"
        _rss_workload(
            args.rss_worker_case,
            args.rss_worker_engine,
            path,
            args.rss_worker_rows,
            args.rss_worker_cols,
            fixture,
        )
    peak = maxrss_bytes()
    print(
        json.dumps(
            {
                "peak_rss_bytes": peak,
                "baseline_rss_bytes": baseline,
                "delta_rss_bytes": (peak - baseline) if peak >= 0 and baseline >= 0 else -1,
            }
        )
    )


def bench_rss(
    case: str,
    engine: str,
    rows: int,
    cols: int,
    fixture: Path | None,
) -> dict[str, Any]:
    """Spawn a clean subprocess that runs one workload and reports peak RSS."""
    cmd = [
        sys.executable,
        os.path.abspath(__file__),
        "--rss-worker-case",
        case,
        "--rss-worker-engine",
        engine,
        "--rss-worker-rows",
        str(rows),
        "--rss-worker-cols",
        str(cols),
        "--rss-worker-fixture",
        str(fixture) if fixture is not None else "-",
    ]
    out = subprocess.check_output(cmd, cwd=ROOT, text=True)
    data = json.loads(out.strip().splitlines()[-1])
    return {
        "case": case,
        "engine": engine,
        "rows": rows,
        "cols": cols,
        "peak_rss_bytes": data["peak_rss_bytes"],
        "baseline_rss_bytes": data["baseline_rss_bytes"],
        "delta_rss_bytes": data["delta_rss_bytes"],
    }


def memory_comparison_rows(memory: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """For each case, ratio of each wolfxl engine's peak RSS to openpyxl's.

    A ratio below 1.0 means wolfxl peaked at less memory (better).
    """
    by_case: dict[str, dict[str, dict[str, Any]]] = {}
    for entry in memory:
        by_case.setdefault(entry["case"], {})[entry["engine"]] = entry

    comparisons: list[dict[str, Any]] = []
    for case, engines in sorted(by_case.items()):
        openpyxl_entry = engines.get("openpyxl")
        if openpyxl_entry is None:
            continue
        op_peak = openpyxl_entry["peak_rss_bytes"]
        op_delta = openpyxl_entry.get("delta_rss_bytes")
        for engine, entry in sorted(engines.items()):
            if engine == "openpyxl":
                continue
            wx_peak = entry["peak_rss_bytes"]
            wx_delta = entry.get("delta_rss_bytes")
            comparisons.append(
                {
                    "case": case,
                    "engine": engine,
                    "openpyxl_peak_rss_bytes": op_peak,
                    "wolfxl_peak_rss_bytes": wx_peak,
                    "peak_rss_excess_bytes": wx_peak - op_peak,
                    "peak_rss_ratio_vs_openpyxl": (
                        wx_peak / op_peak if op_peak and wx_peak >= 0 and op_peak > 0 else None
                    ),
                    "openpyxl_delta_rss_bytes": op_delta,
                    "wolfxl_delta_rss_bytes": wx_delta,
                    "delta_rss_excess_bytes": (
                        wx_delta - op_delta
                        if isinstance(op_delta, int) and isinstance(wx_delta, int)
                        else None
                    ),
                    "delta_rss_ratio_vs_openpyxl": (
                        wx_delta / op_delta
                        if isinstance(op_delta, int)
                        and isinstance(wx_delta, int)
                        and op_delta > 0
                        and wx_delta >= 0
                        else None
                    ),
                }
            )
    return comparisons


def bench_write(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    writer: Callable[[Path, int, int], int],
    warmup: bool = True,
) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        def run_once() -> int:
            path = tmp / f"{name}-{engine}.xlsx"
            if path.exists():
                path.unlink()
            cells = writer(path, rows, cols)
            if not path.exists() or path.stat().st_size == 0:
                raise AssertionError(f"{name} did not create a workbook")
            return cells

        samples, cells = time_samples(run_once, rounds, warmup=warmup)
    return result_payload(name, engine, rows, cols, int(cells), samples)


def bench_write_phased(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    writer: Callable[[Path, int, int], dict[str, Any]],
) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        def run_once() -> dict[str, Any]:
            path = tmp / f"{name}-{engine}.xlsx"
            if path.exists():
                path.unlink()
            result = writer(path, rows, cols)
            if not path.exists() or path.stat().st_size == 0:
                raise AssertionError(f"{name} did not create a workbook")
            return result

        run_once()
        samples: list[float] = []
        phase_samples: dict[str, list[float]] = {}
        units = rows * cols
        for _ in range(rounds):
            started = perf_counter()
            result = run_once()
            samples.append(perf_counter() - started)
            units = int(result["units"])
            for phase, seconds in result["phase_seconds"].items():
                phase_samples.setdefault(phase, []).append(float(seconds))

    payload = result_payload(name, engine, rows, cols, units, samples)
    payload["phase_samples_seconds"] = phase_samples
    payload["phase_medians_seconds"] = {
        phase: statistics.median(values)
        for phase, values in sorted(phase_samples.items())
    }
    return payload


def bench_read(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    fixture: Path,
    reader: Callable[[Path], tuple[int, float]],
    warmup: bool = True,
) -> dict[str, Any]:
    def run_once() -> tuple[int, float]:
        return reader(fixture)

    samples, result = time_samples(run_once, rounds, warmup=warmup)
    read_rows, checksum = result
    expected_checksum = plain_numeric_checksum(rows, cols)
    if read_rows != rows or checksum != expected_checksum:
        raise AssertionError(
            f"{name} returned rows={read_rows}, checksum={checksum}; "
            f"expected rows={rows}, checksum={expected_checksum}"
        )
    return result_payload(name, engine, rows, cols, rows * cols, samples)


def bench_read_phased(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    fixture: Path,
    reader: Callable[[Path], dict[str, Any]],
    warmup: bool = True,
) -> dict[str, Any]:
    def run_once() -> dict[str, Any]:
        return reader(fixture)

    if warmup:
        run_once()

    samples: list[float] = []
    phase_samples: dict[str, list[float]] = {}
    last_result: dict[str, Any] | None = None
    for _ in range(rounds):
        started = perf_counter()
        last_result = run_once()
        samples.append(perf_counter() - started)
        for phase, seconds in last_result["phase_seconds"].items():
            phase_samples.setdefault(phase, []).append(float(seconds))

    assert last_result is not None
    read_rows = int(last_result["rows"])
    checksum = float(last_result["checksum"])
    expected_checksum = plain_numeric_checksum(rows, cols)
    if read_rows != rows or checksum != expected_checksum:
        raise AssertionError(
            f"{name} returned rows={read_rows}, checksum={checksum}; "
            f"expected rows={rows}, checksum={expected_checksum}"
        )

    payload = result_payload(name, engine, rows, cols, rows * cols, samples)
    payload["phase_samples_seconds"] = phase_samples
    payload["phase_medians_seconds"] = {
        phase: statistics.median(values)
        for phase, values in sorted(phase_samples.items())
    }
    return payload


def bench_read_checked(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    fixture: Path,
    reader: Callable[[Path], tuple[Any, ...]],
    expected: tuple[Any, ...],
    units: int,
    warmup: bool = True,
) -> dict[str, Any]:
    def run_once() -> tuple[Any, ...]:
        return reader(fixture)

    samples, result = time_samples(run_once, rounds, warmup=warmup)
    if result != expected:
        raise AssertionError(f"{name} returned {result}; expected {expected}")
    return result_payload(name, engine, rows, cols, units, samples)


def bench_modify(
    name: str,
    engine: str,
    rows: int,
    cols: int,
    rounds: int,
    fixture: Path,
    modifier: Callable[[Path], str],
    warmup: bool = True,
) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)

        def run_once() -> dict[str, Any]:
            path = tmp / f"{name}-{engine}.xlsx"
            phase_seconds: dict[str, float] = {}

            started = perf_counter()
            shutil.copy2(fixture, path)
            phase_seconds["setup_copy"] = perf_counter() - started

            started = perf_counter()
            result = modifier(path)
            phase_seconds["modify_save"] = perf_counter() - started

            started = perf_counter()
            wb = openpyxl.load_workbook(path, data_only=False)
            ws = wb.active
            assert ws["B2"].value == "changed"
            assert ws["C3"].value == 12345
            wb.close()
            phase_seconds["verify_load"] = perf_counter() - started
            return {"result": result, "phase_seconds": phase_seconds}

        if warmup:
            run_once()
        samples: list[float] = []
        phase_samples: dict[str, list[float]] = {}
        for _ in range(rounds):
            started = perf_counter()
            result = run_once()
            samples.append(perf_counter() - started)
            for phase, seconds in result["phase_seconds"].items():
                phase_samples.setdefault(phase, []).append(float(seconds))

    payload = result_payload(name, engine, rows, cols, 2, samples)
    payload["phase_samples_seconds"] = phase_samples
    payload["phase_medians_seconds"] = {
        phase: statistics.median(values)
        for phase, values in sorted(phase_samples.items())
    }
    return payload


def result_payload(
    case: str,
    engine: str,
    rows: int,
    cols: int,
    units: int,
    samples: list[float],
) -> dict[str, Any]:
    median = statistics.median(samples)
    return {
        "case": case,
        "engine": engine,
        "rows": rows,
        "cols": cols,
        "units": units,
        "samples_seconds": samples,
        "median_seconds": median,
        "units_per_second": units / median if median else None,
    }


def comparison_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_case: dict[str, dict[str, dict[str, Any]]] = {}
    for result in results:
        by_case.setdefault(result["case"], {})[result["engine"]] = result

    comparisons: list[dict[str, Any]] = []
    for case, engines in sorted(by_case.items()):
        openpyxl_result = engines.get("openpyxl")
        if openpyxl_result is None:
            continue
        for engine, result in sorted(engines.items()):
            if engine == "openpyxl":
                continue
            wolfxl_seconds = result["median_seconds"]
            comparisons.append(
                {
                    "case": case,
                    "engine": engine,
                    "speedup_vs_openpyxl": (
                        openpyxl_result["median_seconds"] / wolfxl_seconds
                        if wolfxl_seconds
                        else None
                    ),
                }
            )
    return comparisons


def phase_comparison_rows(results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Compare phase medians for phased workloads.

    Some benchmarks include setup or verification phases to keep the workload
    honest. This table exposes the sub-phases so the report can distinguish
    "end-to-end check" time from the actual operation being optimized.
    """
    by_case: dict[str, dict[str, dict[str, Any]]] = {}
    for result in results:
        if result.get("phase_medians_seconds"):
            by_case.setdefault(result["case"], {})[result["engine"]] = result

    comparisons: list[dict[str, Any]] = []
    for case, engines in sorted(by_case.items()):
        openpyxl_result = engines.get("openpyxl")
        if openpyxl_result is None:
            continue
        openpyxl_phases = openpyxl_result.get("phase_medians_seconds") or {}
        for engine, result in sorted(engines.items()):
            if engine == "openpyxl":
                continue
            wolfxl_phases = result.get("phase_medians_seconds") or {}
            shared_phases = sorted(openpyxl_phases.keys() & wolfxl_phases.keys())
            for phase in shared_phases:
                wolfxl_seconds = wolfxl_phases[phase]
                comparisons.append(
                    {
                        "case": case,
                        "engine": engine,
                        "phase": phase,
                        "openpyxl_seconds": openpyxl_phases[phase],
                        "wolfxl_seconds": wolfxl_seconds,
                        "speedup_vs_openpyxl": (
                            openpyxl_phases[phase] / wolfxl_seconds
                            if wolfxl_seconds
                            else None
                        ),
                    }
                )
    return comparisons


_OPERATION_PHASES = {
    "modify_two_cells_plain": "modify_save",
    "large_modify_two_cells_plain": "modify_save",
}


def operation_comparison_rows(
    phase_comparisons: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return claim-facing operation speedups for phased harness benchmarks.

    Some benchmark cases include setup or verification work so the output stays
    correct and comparable. For SOTA gating, those harness phases should stay
    visible, but the operation claim should use the phase that belongs to the
    library behavior being measured.
    """
    rows = []
    for comparison in phase_comparisons:
        phase = _OPERATION_PHASES.get(comparison["case"])
        if phase is None or comparison["phase"] != phase:
            continue
        rows.append({**comparison, "speedup_basis": "operation_phase"})
    return rows


def metadata(rounds: int) -> dict[str, Any]:
    return {
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "git_commit": git_value(["rev-parse", "HEAD"]),
        "git_branch": git_value(["rev-parse", "--abbrev-ref", "HEAD"]),
        "git_dirty": git_dirty(),
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "processor": platform.processor(),
        "machine": platform.machine(),
        "cpu_brand": cpu_brand(),
        "wolfxl_version": wolfxl.__version__,
        "openpyxl_version": openpyxl.__version__,
        "rounds": rounds,
        "pid": os.getpid(),
    }


def git_value(args: list[str]) -> str | None:
    try:
        return subprocess.check_output(["git", *args], cwd=ROOT, text=True).strip()
    except Exception:
        return None


def git_dirty() -> bool | None:
    try:
        return bool(
            subprocess.check_output(["git", "status", "--porcelain"], cwd=ROOT, text=True).strip()
        )
    except Exception:
        return None


def cpu_brand() -> str | None:
    try:
        return subprocess.check_output(
            ["sysctl", "-n", "machdep.cpu.brand_string"],
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip()
    except Exception:
        return None


def selected_cases(values: list[str] | None) -> set[str] | None:
    """Return the requested benchmark case names, or ``None`` for all cases."""
    if not values:
        return None

    cases: set[str] = set()
    for value in values:
        for case in value.split(","):
            case = case.strip()
            if case:
                cases.add(case)
    return cases or None


def _format_bytes(value: int) -> str:
    if value is None or value < 0:
        return "n/a"
    mib = value / (1024 * 1024)
    return f"{mib:,.1f} MiB"


def write_markdown(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        "# WolfXL vs openpyxl Benchmark Run",
        "",
        f"- Timestamp UTC: `{payload['metadata']['timestamp_utc']}`",
        f"- Git branch: `{payload['metadata']['git_branch']}`",
        f"- Git commit: `{payload['metadata']['git_commit']}`",
        f"- Git dirty: `{payload['metadata'].get('git_dirty')}`",
        f"- Python: `{payload['metadata']['python']}`",
        f"- WolfXL: `{payload['metadata']['wolfxl_version']}`",
        f"- openpyxl: `{payload['metadata']['openpyxl_version']}`",
        f"- Rounds: `{payload['metadata']['rounds']}`",
        "",
        "| Case | Engine | Rows | Cols | Median seconds | Units/sec | Phase medians |",
        "|---|---:|---:|---:|---:|---:|---|",
    ]
    for result in payload["results"]:
        units_per_second = result["units_per_second"]
        rate = "" if units_per_second is None else f"{units_per_second:,.0f}"
        phase_medians = result.get("phase_medians_seconds") or {}
        phases = ", ".join(
            f"{phase}={seconds:.6f}s"
            for phase, seconds in phase_medians.items()
        )
        lines.append(
            "| {case} | {engine} | {rows:,} | {cols:,} | {median:.6f} | {rate} | {phases} |".format(
                case=result["case"],
                engine=result["engine"],
                rows=result["rows"],
                cols=result["cols"],
                median=result["median_seconds"],
                rate=rate,
                phases=phases,
            )
        )
    lines.extend(
        [
            "",
            "_Speedup `>1.0x` means WolfXL is faster (higher is better)._",
            "",
            "| Case | WolfXL path | Speedup vs openpyxl |",
            "|---|---:|---:|",
        ]
    )
    for comparison in payload["comparisons"]:
        speedup = comparison["speedup_vs_openpyxl"]
        formatted = "" if speedup is None else f"{speedup:.2f}x"
        lines.append(
            f"| {comparison['case']} | {comparison['engine']} | {formatted} |"
        )

    phase_comparisons = payload.get("phase_comparisons") or []
    if phase_comparisons:
        lines.extend(
            [
                "",
                "## Phase speedups",
                "",
                "For phased workloads, this separates setup/verification time"
                " from the operation being measured. Speedup `>1.0x` means"
                " WolfXL is faster for that phase.",
                "",
                "| Case | WolfXL path | Phase | openpyxl seconds | WolfXL seconds | Speedup |",
                "|---|---|---|---:|---:|---:|",
            ]
        )
        for comparison in phase_comparisons:
            speedup = comparison["speedup_vs_openpyxl"]
            formatted = "" if speedup is None else f"{speedup:.2f}x"
            lines.append(
                "| {case} | {engine} | {phase} | {op:.6f} | {wx:.6f} | {speedup} |".format(
                    case=comparison["case"],
                    engine=comparison["engine"],
                    phase=comparison["phase"],
                    op=comparison["openpyxl_seconds"],
                    wx=comparison["wolfxl_seconds"],
                    speedup=formatted,
                )
            )

    operation_comparisons = payload.get("operation_comparisons") or []
    if operation_comparisons:
        lines.extend(
            [
                "",
                "## Operation speedups",
                "",
                "For benchmarks with setup or verification phases, this table"
                " shows the library operation phase used for claim gating. The"
                " full end-to-end timing above is still reported.",
                "",
                "| Case | WolfXL path | Operation phase | openpyxl seconds | WolfXL seconds | Speedup |",
                "|---|---|---|---:|---:|---:|",
            ]
        )
        for comparison in operation_comparisons:
            speedup = comparison["speedup_vs_openpyxl"]
            formatted = "" if speedup is None else f"{speedup:.2f}x"
            lines.append(
                "| {case} | {engine} | {phase} | {op:.6f} | {wx:.6f} | {speedup} |".format(
                    case=comparison["case"],
                    engine=comparison["engine"],
                    phase=comparison["phase"],
                    op=comparison["openpyxl_seconds"],
                    wx=comparison["wolfxl_seconds"],
                    speedup=formatted,
                )
            )

    memory = payload.get("memory") or []
    if memory:
        lines.extend(
            [
                "",
                "## Peak memory (RSS)",
                "",
                "Peak resident-set size per workload, each measured in its own"
                " subprocess (clean `ru_maxrss` high-water mark). Baseline is the"
                " interpreter with both libraries imported.",
                "",
                "| Case | Engine | Rows | Cols | Peak RSS | Delta over baseline |",
                "|---|---:|---:|---:|---:|---:|",
            ]
        )
        for entry in memory:
            lines.append(
                "| {case} | {engine} | {rows:,} | {cols:,} | {peak} | {delta} |".format(
                    case=entry["case"],
                    engine=entry["engine"],
                    rows=entry["rows"],
                    cols=entry["cols"],
                    peak=_format_bytes(entry["peak_rss_bytes"]),
                    delta=_format_bytes(entry["delta_rss_bytes"]),
                )
            )
        lines.extend(
            [
                "",
                "_Ratio `<1.0x` means WolfXL uses less memory (lower is better)."
                " A ratio `>1.0x` means WolfXL's peak RSS is higher than"
                " openpyxl's. Large-sheet WolfXL reads use the streaming path"
                " and skip eager worksheet-wide style hydration, so their peak"
                " RSS should stay bounded._",
                "",
                "| Case | WolfXL path | Peak ratio | Workload delta ratio | Workload delta diff |",
                "|---|---:|---:|---:|---:|",
            ]
        )
        for comparison in payload.get("memory_comparisons", []):
            peak_ratio = comparison["peak_rss_ratio_vs_openpyxl"]
            delta_ratio = comparison.get("delta_rss_ratio_vs_openpyxl")
            peak_formatted = "" if peak_ratio is None else f"{peak_ratio:.2f}x"
            delta_formatted = "" if delta_ratio is None else f"{delta_ratio:.2f}x"
            delta_diff = comparison.get("delta_rss_excess_bytes")
            lines.append(
                "| {case} | {engine} | {peak_ratio} | {delta_ratio} | {delta_diff} |".format(
                    case=comparison["case"],
                    engine=comparison["engine"],
                    peak_ratio=peak_formatted,
                    delta_ratio=delta_formatted,
                    delta_diff=_format_bytes(delta_diff)
                    if isinstance(delta_diff, int)
                    else "n/a",
                )
            )

    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def add_arguments(parser: argparse.ArgumentParser) -> None:
    parser.add_argument("--rows", type=int, default=10_000)
    parser.add_argument("--cols", type=int, default=5)
    parser.add_argument("--small-rows", type=int, default=2_000)
    parser.add_argument("--wide-rows", type=int, default=1_000)
    parser.add_argument("--wide-cols", type=int, default=50)
    parser.add_argument("--style-rows", type=int, default=2_000)
    parser.add_argument("--formula-rows", type=int, default=2_000)
    parser.add_argument("--sheet-rows", type=int, default=1_000)
    parser.add_argument("--sheet-count", type=int, default=5)
    parser.add_argument("--rounds", type=int, default=3)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--prefix", default=None)
    parser.add_argument(
        "--case",
        action="append",
        dest="cases",
        help=(
            "Run only a named benchmark case. Repeat the option or pass a "
            "comma-separated list, for example --case write_cell_by_cell_plain."
        ),
    )
    # Large-file pass (modify-mode + bounded-memory story).
    parser.add_argument("--large-rows", type=int, default=200_000)
    parser.add_argument("--large-cols", type=int, default=8)
    parser.add_argument("--large-rounds", type=int, default=1)
    parser.add_argument(
        "--no-large",
        dest="large",
        action="store_false",
        help="Skip the large-file workloads.",
    )
    parser.add_argument(
        "--no-memory",
        dest="memory",
        action="store_false",
        help="Skip the peak-RSS measurement pass.",
    )
    parser.set_defaults(large=True, memory=True)
    # Hidden RSS subprocess-worker mode (one workload, prints peak RSS JSON).
    parser.add_argument("--rss-worker-case", default=None, help=argparse.SUPPRESS)
    parser.add_argument("--rss-worker-engine", default=None, help=argparse.SUPPRESS)
    parser.add_argument("--rss-worker-rows", type=int, default=0, help=argparse.SUPPRESS)
    parser.add_argument("--rss-worker-cols", type=int, default=0, help=argparse.SUPPRESS)
    parser.add_argument("--rss-worker-fixture", default=None, help=argparse.SUPPRESS)


# Workloads measured in the peak-RSS pass. Each tuple is
# (case, [engines], needs_fixture). wolfxl engine label is "wolfxl" so the
# memory comparison pairs it against "openpyxl" cleanly.
RSS_CASES = (
    ("write_append_plain", ("openpyxl", "wolfxl"), False),
    ("read_values_plain", ("openpyxl", "wolfxl"), True),
    ("read_only_values_plain", ("openpyxl", "wolfxl"), True),
    ("modify_two_cells_plain", ("openpyxl", "wolfxl"), True),
)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Record a local WolfXL vs openpyxl performance baseline."
    )
    add_arguments(parser)
    args = parser.parse_args()

    # Subprocess RSS-worker mode: run exactly one workload and exit.
    if args.rss_worker_case is not None:
        run_rss_worker(args)
        return

    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    prefix = args.prefix or datetime.now(timezone.utc).strftime("%Y-%m-%d-baseline")
    case_filter = selected_cases(args.cases)

    def include_case(case: str) -> bool:
        return case_filter is None or case in case_filter

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        read_fixture = tmp / "read-fixture.xlsx"
        wide_read_fixture = tmp / "wide-read-fixture.xlsx"
        modify_fixture = tmp / "modify-fixture.xlsx"
        formula_fixture = tmp / "formula-fixture.xlsx"
        styled_fixture = tmp / "styled-fixture.xlsx"
        multi_sheet_fixture = tmp / "multi-sheet-fixture.xlsx"
        if include_case("read_values_plain") or include_case("read_only_values_plain"):
            build_openpyxl_plain(read_fixture, args.rows, args.cols)
        if include_case("read_values_wide"):
            build_openpyxl_wide(wide_read_fixture, args.wide_rows, args.wide_cols)
        if include_case("modify_two_cells_plain"):
            build_openpyxl_plain(modify_fixture, args.rows, args.cols)
        if include_case("read_formula_text"):
            build_openpyxl_formula(formula_fixture, args.formula_rows, min(args.cols, 6))
        if include_case("read_styled_cells"):
            build_openpyxl_styled(styled_fixture, args.style_rows, args.cols)
        if include_case("read_multi_sheet_values"):
            build_openpyxl_multi_sheet(
                multi_sheet_fixture,
                args.sheet_rows,
                args.cols,
                args.sheet_count,
            )

        results = []

        def add_result(case: str, payload_factory: Callable[[], dict[str, Any]]) -> None:
            if include_case(case):
                results.append(payload_factory())

        add_result(
            "write_append_plain",
            lambda: bench_write(
                "write_append_plain",
                "openpyxl",
                args.rows,
                args.cols,
                args.rounds,
                build_openpyxl_plain,
            ),
        )
        add_result(
            "write_append_plain",
            lambda: bench_write(
                "write_append_plain",
                "wolfxl_append",
                args.rows,
                args.cols,
                args.rounds,
                build_wolfxl_append_plain,
            ),
        )
        add_result(
            "write_append_plain",
            lambda: bench_write(
                "write_append_plain",
                "wolfxl_write_rows",
                args.rows,
                args.cols,
                args.rounds,
                build_wolfxl_write_rows_plain,
            ),
        )
        add_result(
            "write_append_mixed",
            lambda: bench_write(
                "write_append_mixed",
                "openpyxl",
                args.rows,
                min(args.cols, 6),
                args.rounds,
                build_openpyxl_mixed,
            ),
        )
        add_result(
            "write_append_mixed",
            lambda: bench_write(
                "write_append_mixed",
                "wolfxl_append",
                args.rows,
                min(args.cols, 6),
                args.rounds,
                build_wolfxl_append_mixed,
            ),
        )
        add_result(
            "write_append_wide",
            lambda: bench_write(
                "write_append_wide",
                "openpyxl",
                args.wide_rows,
                args.wide_cols,
                args.rounds,
                build_openpyxl_wide,
            ),
        )
        add_result(
            "write_append_wide",
            lambda: bench_write(
                "write_append_wide",
                "wolfxl_append",
                args.wide_rows,
                args.wide_cols,
                args.rounds,
                build_wolfxl_wide,
            ),
        )
        add_result(
            "write_append_formulas",
            lambda: bench_write(
                "write_append_formulas",
                "openpyxl",
                args.formula_rows,
                min(args.cols, 6),
                args.rounds,
                build_openpyxl_formula,
            ),
        )
        add_result(
            "write_append_formulas",
            lambda: bench_write(
                "write_append_formulas",
                "wolfxl_append",
                args.formula_rows,
                min(args.cols, 6),
                args.rounds,
                build_wolfxl_formula,
            ),
        )
        add_result(
            "write_append_formulas",
            lambda: bench_write(
                "write_append_formulas",
                "wolfxl_write_rows_fast_formula",
                args.formula_rows,
                min(args.cols, 6),
                args.rounds,
                build_wolfxl_formula_rows_fast,
            ),
        )
        add_result(
            "write_styled_cells",
            lambda: bench_write_phased(
                "write_styled_cells",
                "openpyxl",
                args.style_rows,
                args.cols,
                args.rounds,
                build_openpyxl_styled_phased,
            ),
        )
        add_result(
            "write_styled_cells",
            lambda: bench_write_phased(
                "write_styled_cells",
                "wolfxl_cell",
                args.style_rows,
                args.cols,
                args.rounds,
                build_wolfxl_styled_phased,
            ),
        )
        add_result(
            "write_styled_cells",
            lambda: bench_write_phased(
                "write_styled_cells",
                "wolfxl_write_styled_rows",
                args.style_rows,
                args.cols,
                args.rounds,
                build_wolfxl_styled_rows_phased,
            ),
        )
        add_result(
            "write_multi_sheet_plain",
            lambda: bench_write(
                "write_multi_sheet_plain",
                "openpyxl",
                args.sheet_rows,
                args.cols,
                args.rounds,
                lambda path, rows, cols: build_openpyxl_multi_sheet(
                    path, rows, cols, args.sheet_count
                ),
            ),
        )
        add_result(
            "write_multi_sheet_plain",
            lambda: bench_write(
                "write_multi_sheet_plain",
                "wolfxl_append",
                args.sheet_rows,
                args.cols,
                args.rounds,
                lambda path, rows, cols: build_wolfxl_multi_sheet(
                    path, rows, cols, args.sheet_count
                ),
            ),
        )
        add_result(
            "write_multi_sheet_plain",
            lambda: bench_write(
                "write_multi_sheet_plain",
                "wolfxl_write_rows_fast_multi_sheet",
                args.sheet_rows,
                args.cols,
                args.rounds,
                lambda path, rows, cols: build_wolfxl_multi_sheet_write_rows_fast(
                    path,
                    rows,
                    cols,
                    args.sheet_count,
                ),
            ),
        )
        add_result(
            "write_cell_by_cell_plain",
            lambda: bench_write_phased(
                "write_cell_by_cell_plain",
                "openpyxl",
                args.small_rows,
                args.cols,
                args.rounds,
                build_openpyxl_cell_by_cell_phased,
            ),
        )
        add_result(
            "write_cell_by_cell_plain",
            lambda: bench_write_phased(
                "write_cell_by_cell_plain",
                "wolfxl_cell",
                args.small_rows,
                args.cols,
                args.rounds,
                build_wolfxl_cell_by_cell_phased,
            ),
        )
        add_result(
            "read_values_plain",
            lambda: bench_read(
                "read_values_plain",
                "openpyxl",
                args.rows,
                args.cols,
                args.rounds,
                read_fixture,
                read_openpyxl_values,
            ),
        )
        add_result(
            "read_values_plain",
            lambda: bench_read(
                "read_values_plain",
                "wolfxl",
                args.rows,
                args.cols,
                args.rounds,
                read_fixture,
                read_wolfxl_values,
            ),
        )
        add_result(
            "read_only_values_plain",
            lambda: bench_read(
                "read_only_values_plain",
                "openpyxl",
                args.rows,
                args.cols,
                args.rounds,
                read_fixture,
                read_openpyxl_read_only_values,
            ),
        )
        add_result(
            "read_only_values_plain",
            lambda: bench_read(
                "read_only_values_plain",
                "wolfxl",
                args.rows,
                args.cols,
                args.rounds,
                read_fixture,
                read_wolfxl_read_only_values,
            ),
        )
        add_result(
            "read_values_wide",
            lambda: bench_read_checked(
                "read_values_wide",
                "openpyxl",
                args.wide_rows,
                args.wide_cols,
                args.rounds,
                wide_read_fixture,
                read_openpyxl_wide_values,
                (
                    args.wide_rows,
                    sum(
                        float(value)
                        for row_idx in range(1, args.wide_rows + 1)
                        for value in wide_row(row_idx, args.wide_cols)
                        if isinstance(value, (int, float))
                    ),
                    args.wide_rows * args.wide_cols,
                ),
                args.wide_rows * args.wide_cols,
            ),
        )
        add_result(
            "read_values_wide",
            lambda: bench_read_checked(
                "read_values_wide",
                "wolfxl",
                args.wide_rows,
                args.wide_cols,
                args.rounds,
                wide_read_fixture,
                read_wolfxl_wide_values,
                (
                    args.wide_rows,
                    sum(
                        float(value)
                        for row_idx in range(1, args.wide_rows + 1)
                        for value in wide_row(row_idx, args.wide_cols)
                        if isinstance(value, (int, float))
                    ),
                    args.wide_rows * args.wide_cols,
                ),
                args.wide_rows * args.wide_cols,
            ),
        )
        add_result(
            "read_formula_text",
            lambda: bench_read_checked(
                "read_formula_text",
                "openpyxl",
                args.formula_rows,
                min(args.cols, 6),
                args.rounds,
                formula_fixture,
                read_openpyxl_formulas,
                (
                    args.formula_rows,
                    args.formula_rows * (args.formula_rows + 1) / 2,
                    args.formula_rows
                    * sum(
                        1
                        for value in formula_row(1, min(args.cols, 6))
                        if isinstance(value, str) and value.startswith("=")
                    ),
                ),
                args.formula_rows * min(args.cols, 6),
            ),
        )
        add_result(
            "read_formula_text",
            lambda: bench_read_checked(
                "read_formula_text",
                "wolfxl",
                args.formula_rows,
                min(args.cols, 6),
                args.rounds,
                formula_fixture,
                read_wolfxl_formulas,
                (
                    args.formula_rows,
                    args.formula_rows * (args.formula_rows + 1) / 2,
                    args.formula_rows
                    * sum(
                        1
                        for value in formula_row(1, min(args.cols, 6))
                        if isinstance(value, str) and value.startswith("=")
                    ),
                ),
                args.formula_rows * min(args.cols, 6),
            ),
        )
        add_result(
            "read_styled_cells",
            lambda: bench_read_checked(
                "read_styled_cells",
                "openpyxl",
                args.style_rows,
                args.cols,
                args.rounds,
                styled_fixture,
                read_openpyxl_styled_cells,
                (
                    args.style_rows,
                    args.style_rows * (args.style_rows + 1) / 2,
                    args.style_rows
                    + (args.style_rows // 2 if args.cols >= 2 else 0)
                    + (args.style_rows if args.cols >= 3 else 0),
                ),
                args.style_rows * args.cols,
            ),
        )
        add_result(
            "read_styled_cells",
            lambda: bench_read_checked(
                "read_styled_cells",
                "wolfxl",
                args.style_rows,
                args.cols,
                args.rounds,
                styled_fixture,
                read_wolfxl_styled_cells,
                (
                    args.style_rows,
                    args.style_rows * (args.style_rows + 1) / 2,
                    args.style_rows
                    + (args.style_rows // 2 if args.cols >= 2 else 0)
                    + (args.style_rows if args.cols >= 3 else 0),
                ),
                args.style_rows * args.cols,
            ),
        )
        add_result(
            "read_multi_sheet_values",
            lambda: bench_read_checked(
                "read_multi_sheet_values",
                "openpyxl",
                args.sheet_rows * args.sheet_count,
                args.cols,
                args.rounds,
                multi_sheet_fixture,
                read_openpyxl_multi_sheet,
                (
                    args.sheet_rows * args.sheet_count,
                    (args.sheet_rows * (args.sheet_rows + 1) / 2)
                    * args.sheet_count,
                ),
                args.sheet_rows * args.sheet_count * args.cols,
            ),
        )
        add_result(
            "read_multi_sheet_values",
            lambda: bench_read_checked(
                "read_multi_sheet_values",
                "wolfxl",
                args.sheet_rows * args.sheet_count,
                args.cols,
                args.rounds,
                multi_sheet_fixture,
                read_wolfxl_multi_sheet,
                (
                    args.sheet_rows * args.sheet_count,
                    (args.sheet_rows * (args.sheet_rows + 1) / 2)
                    * args.sheet_count,
                ),
                args.sheet_rows * args.sheet_count * args.cols,
            ),
        )
        add_result(
            "modify_two_cells_plain",
            lambda: bench_modify(
                "modify_two_cells_plain",
                "openpyxl",
                args.rows,
                args.cols,
                args.rounds,
                modify_fixture,
                modify_openpyxl,
            ),
        )
        add_result(
            "modify_two_cells_plain",
            lambda: bench_modify(
                "modify_two_cells_plain",
                "wolfxl_modify",
                args.rows,
                args.cols,
                args.rounds,
                modify_fixture,
                modify_wolfxl,
            ),
        )

        # ---- Large-file pass: bounded-memory + modify-mode advantage ----
        large_read_fixture: Path | None = None
        large_modify_fixture: Path | None = None
        if args.large:
            large_read_fixture = tmp / "large-read-fixture.xlsx"
            large_modify_fixture = tmp / "large-modify-fixture.xlsx"
            if include_case("large_read_values_plain") or include_case("large_read_only_values_plain"):
                build_openpyxl_plain(large_read_fixture, args.large_rows, args.large_cols)
            if include_case("large_modify_two_cells_plain"):
                build_openpyxl_plain(large_modify_fixture, args.large_rows, args.large_cols)

            lr, lc, lrounds = args.large_rows, args.large_cols, args.large_rounds
            add_result(
                "large_write_append_plain",
                lambda: bench_write("large_write_append_plain", "openpyxl", lr, lc, lrounds, build_openpyxl_plain, warmup=False),
            )
            add_result(
                "large_write_append_plain",
                lambda: bench_write("large_write_append_plain", "wolfxl_append", lr, lc, lrounds, build_wolfxl_append_plain, warmup=False),
            )
            add_result(
                "large_write_append_plain",
                lambda: bench_write("large_write_append_plain", "wolfxl_write_rows", lr, lc, lrounds, build_wolfxl_write_rows_plain, warmup=False),
            )
            add_result(
                "large_read_values_plain",
                lambda: bench_read("large_read_values_plain", "openpyxl", lr, lc, lrounds, large_read_fixture, read_openpyxl_values, warmup=False),
            )
            add_result(
                "large_read_values_plain",
                lambda: bench_read("large_read_values_plain", "wolfxl", lr, lc, lrounds, large_read_fixture, read_wolfxl_values, warmup=False),
            )
            add_result(
                "large_read_only_values_plain",
                lambda: bench_read("large_read_only_values_plain", "openpyxl", lr, lc, lrounds, large_read_fixture, read_openpyxl_read_only_values, warmup=False),
            )
            add_result(
                "large_read_only_values_plain",
                lambda: bench_read("large_read_only_values_plain", "wolfxl", lr, lc, lrounds, large_read_fixture, read_wolfxl_read_only_values, warmup=False),
            )
            add_result(
                "large_modify_two_cells_plain",
                lambda: bench_modify("large_modify_two_cells_plain", "openpyxl", lr, lc, lrounds, large_modify_fixture, modify_openpyxl, warmup=False),
            )
            add_result(
                "large_modify_two_cells_plain",
                lambda: bench_modify("large_modify_two_cells_plain", "wolfxl_modify", lr, lc, lrounds, large_modify_fixture, modify_wolfxl, warmup=False),
            )

        # ---- Peak-RSS pass (subprocess-isolated high-water marks) ----
        memory: list[dict[str, Any]] = []
        if args.memory:

            def _fixture_for(case: str, standard: bool) -> Path | None:
                if "modify" in case:
                    return modify_fixture if standard else large_modify_fixture
                if "read" in case:
                    return read_fixture if standard else large_read_fixture
                return None

            for case, engines, _needs in RSS_CASES:
                if not include_case(case):
                    continue
                for engine in engines:
                    memory.append(
                        bench_rss(case, engine, args.rows, args.cols, _fixture_for(case, True))
                    )
            if args.large:
                for case, engines, _needs in RSS_CASES:
                    large_case = f"large_{case}"
                    if not include_case(large_case):
                        continue
                    for engine in engines:
                        memory.append(
                            bench_rss(
                                large_case,
                                engine,
                                args.large_rows,
                                args.large_cols,
                                _fixture_for(case, False),
                            )
                        )

    if not results and not memory:
        requested = ", ".join(sorted(case_filter or ()))
        raise SystemExit(f"no benchmark cases matched: {requested}")

    phase_comparisons = phase_comparison_rows(results)
    payload = {
        "metadata": metadata(args.rounds),
        "parameters": {
            "cases": sorted(case_filter) if case_filter else None,
            "rows": args.rows,
            "cols": args.cols,
            "small_rows": args.small_rows,
            "wide_rows": args.wide_rows,
            "wide_cols": args.wide_cols,
            "style_rows": args.style_rows,
            "formula_rows": args.formula_rows,
            "sheet_rows": args.sheet_rows,
            "sheet_count": args.sheet_count,
            "large_rows": args.large_rows if args.large else None,
            "large_cols": args.large_cols if args.large else None,
            "large_rounds": args.large_rounds if args.large else None,
        },
        "results": results,
        "comparisons": comparison_rows(results),
        "phase_comparisons": phase_comparisons,
        "operation_comparisons": operation_comparison_rows(phase_comparisons),
        "memory": memory,
        "memory_comparisons": memory_comparison_rows(memory),
    }
    json_path = output_dir / f"{prefix}.json"
    md_path = output_dir / f"{prefix}.md"
    json_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    write_markdown(md_path, payload)
    print(json.dumps(payload, indent=2))
    print(f"\nWrote {json_path}")
    print(f"Wrote {md_path}")


if __name__ == "__main__":
    main()
