"""Sprint Ι Pod-β — SAX streaming read tests.

Pin the contract added in Sprint Ι Pod-β: ``load_workbook(read_only=True)``
exposes ``Worksheet.iter_rows`` as a true streaming generator backed by
``wolfxl._rust.StreamingSheetReader``. Cells yielded in that path are
``StreamingCell`` proxies that surface the same value/style attributes as
the eager ``Cell`` API but reject mutation. The auto-trigger heuristic is
also covered: workbooks loaded without ``read_only=True`` but with
> ``AUTO_STREAM_ROW_THRESHOLD`` rows still flow through the streaming
path, transparently to the caller.
"""

from __future__ import annotations

import datetime as dt
import os
from pathlib import Path
import re
import zipfile

import openpyxl
import pytest

import wolfxl
from wolfxl._streaming import (
    AUTO_STREAM_ROW_THRESHOLD,
    StreamingCell,
    should_auto_stream,
)


# ---------------------------------------------------------------------------
# Fixture builders — generated on the fly via openpyxl so the fixture content
# stays in lockstep with the parity reference implementation.
# ---------------------------------------------------------------------------


def _build_basic(path: Path, n_rows: int = 1000, n_cols: int = 5) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c, value=r * 10 + c)
    wb.save(path)
    return path


def _build_styled(path: Path) -> Path:
    """Small fixture exercising bold/red/center alignment + named formats."""
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Styled"
    ws.cell(row=1, column=1, value="bold")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2, value="red-fill")
    ws.cell(row=1, column=2).fill = PatternFill(
        fill_type="solid", fgColor="FFFF0000"
    )
    ws.cell(row=2, column=1, value=3.14)
    ws.cell(row=2, column=1).number_format = "0.00"
    ws.cell(row=2, column=2, value="centered")
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=1, value="plain")
    wb.save(path)
    return path


def _build_mixed_types(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.cell(row=1, column=1, value=42)
    ws.cell(row=1, column=2, value=3.14)
    ws.cell(row=1, column=3, value="hello")
    ws.cell(row=1, column=4, value=True)
    ws.cell(row=1, column=5, value=False)
    ws.cell(row=2, column=1, value=dt.datetime(2024, 1, 15, 12, 30))
    ws.cell(row=2, column=2, value=dt.date(2024, 6, 1))
    ws.cell(row=2, column=3, value="=A1*2")  # formula
    # Row 3 is intentionally absent → sparse row test.
    ws.cell(row=4, column=2, value="middle")  # only B4 populated
    wb.save(path)
    return path


def _build_sparse(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws.cell(row=1, column=1, value="r1")
    ws.cell(row=5, column=3, value="r5c3")
    ws.cell(row=10, column=1, value="r10")
    wb.save(path)
    return path


def _build_formula_no_date_styles(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FormulaPlain"
    ws["A1"] = 2
    ws["B1"] = "=A1*2"
    wb.save(path)
    return path


def _build_synthetic_60k(path: Path) -> Path:
    """Build a sheet with > AUTO_STREAM_ROW_THRESHOLD rows to exercise the
    auto-trigger. We use openpyxl's write_only mode for speed.
    """
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Big")
    for r in range(1, AUTO_STREAM_ROW_THRESHOLD + 5_001):
        ws.append([r, r * 2, f"row{r}"])
    wb.save(path)
    return path


def _build_fake_high_dimension(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BigDim"
    ws["A1"] = "only-row"
    raw_path = path.with_name("raw-high-dimension.xlsx")
    wb.save(raw_path)
    with (
        zipfile.ZipFile(raw_path, "r") as src,
        zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as dst,
    ):
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(
                    rb'(<dimension\b[^>]*\bref=")[^"]+(")',
                    rb"\g<1>A1:A50001\2",
                    data,
                    count=1,
                )
            dst.writestr(info, data)
    return path


@pytest.fixture
def basic_xlsx(tmp_path: Path) -> Path:
    return _build_basic(tmp_path / "basic.xlsx")


@pytest.fixture
def styled_xlsx(tmp_path: Path) -> Path:
    return _build_styled(tmp_path / "styled.xlsx")


@pytest.fixture
def mixed_xlsx(tmp_path: Path) -> Path:
    return _build_mixed_types(tmp_path / "mixed.xlsx")


@pytest.fixture
def sparse_xlsx(tmp_path: Path) -> Path:
    return _build_sparse(tmp_path / "sparse.xlsx")


@pytest.fixture
def formula_no_date_xlsx(tmp_path: Path) -> Path:
    return _build_formula_no_date_styles(tmp_path / "formula-no-date.xlsx")


# ---------------------------------------------------------------------------
# 1. Basic — 1000 rows in order via values_only.
# ---------------------------------------------------------------------------


def test_streaming_values_only_yields_1000_rows(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1000
    assert rows[0] == (11, 12, 13, 14, 15)
    assert rows[-1] == (10001, 10002, 10003, 10004, 10005)


def test_streaming_values_only_prefers_compact_chunks(
    monkeypatch: pytest.MonkeyPatch,
    basic_xlsx: Path,
) -> None:
    import wolfxl._rust as rust

    calls = []

    class CompactReader:
        def __init__(self) -> None:
            self._done = False

        def read_next_values_row(self) -> None:
            raise AssertionError("compact chunk path should avoid row-at-a-time reads")

        def read_next_values_chunk_with_flags(self) -> None:
            raise AssertionError("compact chunk path should avoid row-indexed chunks")

        def read_next_values_compact_chunk_with_flags(
            self,
            chunk_size: int,
        ) -> tuple[list[tuple[int, ...]], bool, int, int, bool] | None:
            calls.append(chunk_size)
            if self._done:
                return None
            self._done = True
            return ([(11, 12, 13), (21, 22, 23)], False, 1, 2, True)

        def close(self) -> None:
            pass

    class CompactReaderFactory:
        @staticmethod
        def open(*args: object, **kwargs: object) -> CompactReader:
            return CompactReader()

    monkeypatch.setattr(rust, "StreamingSheetReader", CompactReaderFactory)
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    rows = list(
        wb["Sheet1"].iter_rows(
            min_row=1,
            max_row=2,
            min_col=1,
            max_col=3,
            values_only=True,
        )
    )

    assert rows == [(11, 12, 13), (21, 22, 23)]
    assert calls == [1024, 1024]


def test_streaming_iter_value_chunks_match_values_only_rows(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]

    chunks = list(ws.iter_value_chunks(chunk_size=128))
    rows = [row for chunk in chunks for row in chunk]

    assert len(chunks) == 8
    assert rows == list(ws.iter_rows(values_only=True))


def test_streaming_iter_value_chunks_match_sparse_rows(sparse_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(sparse_xlsx, read_only=True)
    ws = wb.active

    chunked_rows = [row for chunk in ws.iter_value_chunks(chunk_size=2) for row in chunk]

    assert chunked_rows == list(ws.iter_rows(values_only=True))


def test_data_only_iter_value_chunks_uses_direct_plain_value_reader(
    basic_xlsx: Path,
) -> None:
    class DirectChunkCounter:
        def __init__(self, inner: object) -> None:
            self._inner = inner
            self.calls: list[int] = []

        def __getattr__(self, name: str) -> object:
            return getattr(self._inner, name)

        def read_sheet_value_chunks_plain(
            self,
            *args: object,
            **kwargs: object,
        ) -> object:
            self.calls.append(int(args[1]))
            return self._inner.read_sheet_value_chunks_plain(*args, **kwargs)

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True, data_only=True)
    counter = DirectChunkCounter(wb._rust_reader)  # noqa: SLF001
    wb._rust_reader = counter  # noqa: SLF001
    ws = wb["Sheet1"]

    rows = [row for chunk in ws.iter_value_chunks(chunk_size=128) for row in chunk]

    assert rows == list(ws.iter_rows(values_only=True))
    assert counter.calls == [128]


def test_data_only_iter_rows_keeps_streaming_reader(
    basic_xlsx: Path,
) -> None:
    class DirectChunkGuard:
        def __init__(self, inner: object) -> None:
            self._inner = inner

        def __getattr__(self, name: str) -> object:
            return getattr(self._inner, name)

        def read_sheet_value_chunks_plain(
            self,
            *args: object,
            **kwargs: object,
        ) -> object:
            raise AssertionError("iter_rows must keep the streaming row reader")

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True, data_only=True)
    wb._rust_reader = DirectChunkGuard(wb._rust_reader)  # noqa: SLF001

    rows = list(wb["Sheet1"].iter_rows(values_only=True))

    assert rows[0] == (11, 12, 13, 14, 15)


def test_iter_value_chunks_rejects_invalid_chunk_size(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)

    with pytest.raises(ValueError, match="chunk_size"):
        list(wb["Sheet1"].iter_value_chunks(chunk_size=0))


def test_read_only_values_only_does_not_hydrate_cells_or_bulk_read(
    basic_xlsx: Path,
) -> None:
    class BulkGuard:
        def __init__(self, inner: object) -> None:
            self._inner = inner

        def __getattr__(self, name: str) -> object:
            return getattr(self._inner, name)

        def read_sheet_values_plain(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("streaming values_only must not use bulk eager read")

        def read_sheet_values(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("streaming values_only must not use bulk eager read")

        def read_sheet_bounds(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("read_only streaming must not use eager bounds read")

        def read_sheet_dimensions(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("read_only streaming must not use eager dimensions read")

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    wb._rust_reader = BulkGuard(wb._rust_reader)  # noqa: SLF001
    ws = wb["Sheet1"]
    assert ws._cells == {}  # noqa: SLF001

    rows = list(ws.iter_rows(values_only=True))

    assert rows[0] == (11, 12, 13, 14, 15)
    assert ws._cells == {}  # noqa: SLF001


def test_read_only_open_skips_eager_merged_border_hydration(
    monkeypatch: pytest.MonkeyPatch,
    basic_xlsx: Path,
) -> None:
    import wolfxl._workbook_state as workbook_state

    archive_probe_called = False

    def mark_archive_probe(*args: object, **kwargs: object) -> bool:
        nonlocal archive_probe_called
        archive_probe_called = True
        return False

    def fail_if_called(*args: object, **kwargs: object) -> None:
        raise AssertionError("read_only=True must not parse worksheet XML at open")

    monkeypatch.setattr(
        workbook_state, "_archive_has_merged_cells", mark_archive_probe
    )
    monkeypatch.setattr(
        workbook_state, "_hydrate_merged_cell_borders", fail_if_called
    )

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)

    assert wb.active.title == "Sheet1"
    assert archive_probe_called is False


def test_eager_open_skips_eager_merged_border_hydration(
    monkeypatch: pytest.MonkeyPatch,
    basic_xlsx: Path,
) -> None:
    import wolfxl._workbook_state as workbook_state

    def fail_if_called(*args: object, **kwargs: object) -> None:
        raise AssertionError("normal load must not parse worksheet XML at open")

    monkeypatch.setattr(
        workbook_state, "_hydrate_merged_cell_borders", fail_if_called
    )

    wb = wolfxl.load_workbook(basic_xlsx)

    assert wb.active.title == "Sheet1"


def test_read_only_values_property_does_not_hydrate_cells(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]

    first = next(ws.values)

    assert first == (11, 12, 13, 14, 15)
    assert ws._cells == {}  # noqa: SLF001


# ---------------------------------------------------------------------------
# 2. Style access — read-mode StreamingCell exposes font.bold etc.
# ---------------------------------------------------------------------------


def test_streaming_cells_expose_styles(styled_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(styled_xlsx, read_only=True)
    ws = wb["Styled"]
    rows = list(ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2))
    # All cells are StreamingCell instances.
    for r in rows:
        for c in r:
            assert isinstance(c, StreamingCell)
    # A1 = "bold" with bold font.
    assert rows[0][0].value == "bold"
    assert rows[0][0].font.bold is True
    # A2 = 3.14, number_format "0.00".
    assert rows[1][0].value == 3.14
    assert rows[1][0].number_format == "0.00"
    # B2 alignment horizontal=center.
    assert rows[1][1].alignment.horizontal == "center"


# ---------------------------------------------------------------------------
# 3. Bounded range — min_row=10, max_row=20 yields 11 rows.
# ---------------------------------------------------------------------------


def test_streaming_bounded_range(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(
        ws.iter_rows(min_row=10, max_row=20, min_col=1, max_col=3, values_only=True)
    )
    assert len(rows) == 11
    assert rows[0] == (101, 102, 103)
    assert rows[-1] == (201, 202, 203)


# ---------------------------------------------------------------------------
# 4. Mixed types — number, str, bool, date, formula all surface correctly.
# ---------------------------------------------------------------------------


def test_streaming_mixed_types(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=5, values_only=True)
    )
    # Row 1: int, float, str, True, False.
    r1 = rows[0]
    assert r1[0] == 42
    assert r1[1] == pytest.approx(3.14)
    assert r1[2] == "hello"
    assert r1[3] is True
    assert r1[4] is False
    # Row 2 col 3 — a formula cached as a string.
    r2 = rows[1]
    val_a = r2[2]
    # Streaming surfaces the formula text from `Cell.value` to match
    # openpyxl read_only=True semantics.
    assert val_a in ("=A1*2", "84")  # cached value may be either


def test_streaming_formula_values_only_without_date_styles(
    formula_no_date_xlsx: Path,
) -> None:
    wb = wolfxl.load_workbook(formula_no_date_xlsx, read_only=True)
    ws = wb["FormulaPlain"]
    row = next(ws.iter_rows(values_only=True))
    assert row == (2, "=A1*2")


# ---------------------------------------------------------------------------
# 5. Empty rows / sparse cells — yields tuples of Nones for missing cells.
# ---------------------------------------------------------------------------


def test_streaming_sparse_rows(sparse_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(sparse_xlsx, read_only=True)
    ws = wb["Sparse"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True)
    )
    # Explicit bounds match openpyxl: missing rows are padded with empty tuples.
    assert len(rows) == 10
    # Sparse rows should appear with the right shape:
    found_first_cells = [r[0] for r in rows]
    assert "r1" in found_first_cells
    assert "r10" in found_first_cells


# ---------------------------------------------------------------------------
# 6. SST reference — a `t="s"` cell resolves to its shared-strings entry.
# ---------------------------------------------------------------------------


def test_streaming_sst_resolves(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    # Repeat strings → shared in SST.
    for i in range(10):
        ws.cell(row=i + 1, column=1, value="repeated")
        ws.cell(row=i + 1, column=2, value=f"unique-{i}")
    path = tmp_path / "sst.xlsx"
    wb.save(path)

    wb2 = wolfxl.load_workbook(path, read_only=True)
    ws2 = wb2["Sheet"]
    rows = list(
        ws2.iter_rows(min_row=1, max_row=10, min_col=1, max_col=2, values_only=True)
    )
    assert len(rows) == 10
    for i, r in enumerate(rows):
        assert r == ("repeated", f"unique-{i}")


# ---------------------------------------------------------------------------
# 7. Style index — `<c s="N">` surfaces non-zero style_id when set.
# ---------------------------------------------------------------------------


def test_streaming_style_id_is_set(styled_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(styled_xlsx, read_only=True)
    ws = wb["Styled"]
    rows = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1))
    cell = rows[0][0]
    # Style id is internal — we expose it via the public font property,
    # which must have bold=True since the underlying `s=` index points
    # at a styled xf.
    assert cell.font.bold is True


# ---------------------------------------------------------------------------
# 8. Mutation rejection — every setter raises a clear RuntimeError.
# ---------------------------------------------------------------------------


def test_streaming_cell_value_setter_raises(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.value = "X"


def test_streaming_cell_font_setter_raises(basic_xlsx: Path) -> None:
    from wolfxl import Font

    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.font = Font(bold=True)


def test_streaming_cell_typo_attr_raises(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    cell = next(iter(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1)))[0]
    with pytest.raises(RuntimeError, match="read_only=True"):
        cell.bogus_attr = 123


# ---------------------------------------------------------------------------
# 9. Mode independence — read_only=False still works (no regression).
# ---------------------------------------------------------------------------


def test_eager_mode_unchanged(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx)
    ws = wb["Sheet1"]
    # Direct cell access still returns a regular Cell, not StreamingCell.
    cell = ws["A1"]
    assert not isinstance(cell, StreamingCell)
    assert cell.value == 11
    # iter_rows in eager mode also returns regular Cell-tuple rows.
    rows = list(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3))
    assert all(not isinstance(c, StreamingCell) for c in rows[0])


# ---------------------------------------------------------------------------
# 10. Auto-trigger — a > 50k-row workbook streams even without read_only=True.
# ---------------------------------------------------------------------------


def test_auto_trigger_uses_dimension_head_not_eager_reader(tmp_path: Path) -> None:
    class EagerGuard:
        def __init__(self, inner: object) -> None:
            self._inner = inner

        def __getattr__(self, name: str) -> object:
            return getattr(self._inner, name)

        def read_sheet_values_plain(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("auto-stream detection must not bulk-read values")

        def read_sheet_values(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("auto-stream detection must not bulk-read values")

        def read_sheet_bounds(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("auto-stream detection must not eagerly read bounds")

        def read_sheet_dimensions(self, *args: object, **kwargs: object) -> object:
            raise AssertionError("auto-stream detection must not eagerly read dimensions")

    path = _build_fake_high_dimension(tmp_path / "fake-high-dimension.xlsx")
    wb = wolfxl.load_workbook(path)
    wb._rust_reader = EagerGuard(wb._rust_reader)  # noqa: SLF001
    ws = wb["BigDim"]

    assert should_auto_stream(ws) is True
    assert next(ws.iter_rows(values_only=True)) == ("only-row",)
    assert ws._cells == {}  # noqa: SLF001


def test_auto_stream_open_skips_eager_merged_border_hydration(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    import wolfxl._workbook_state as workbook_state

    def fail_if_called(*args: object, **kwargs: object) -> None:
        raise AssertionError("auto-stream sheets must not parse worksheet XML at open")

    monkeypatch.setattr(
        workbook_state, "_hydrate_merged_cell_borders", fail_if_called
    )
    path = _build_fake_high_dimension(tmp_path / "fake-high-dimension.xlsx")

    wb = wolfxl.load_workbook(path)

    assert wb.active.title == "BigDim"
    assert should_auto_stream(wb.active) is True


@pytest.mark.slow
def test_auto_trigger_above_threshold(tmp_path: Path) -> None:
    # Skip locally if generation is too slow; pinned slow.
    path = _build_synthetic_60k(tmp_path / "huge.xlsx")
    wb = wolfxl.load_workbook(path)  # read_only=False
    ws = wb.active
    assert should_auto_stream(ws)
    # Reading via iter_rows doesn't OOM and yields ALL rows.
    n = sum(1 for _ in ws.iter_rows(values_only=True))
    assert n > AUTO_STREAM_ROW_THRESHOLD


# ---------------------------------------------------------------------------
# 11. close() / generator cleanup — file handle is released after iteration.
# ---------------------------------------------------------------------------


def test_streaming_iterator_cleanup(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    gen = ws.iter_rows(values_only=True)
    # Pull a single row, then drop the generator.
    next(gen)
    del gen
    # File is no longer locked; we can rename / delete it.
    new_path = basic_xlsx.parent / "renamed.xlsx"
    os.rename(basic_xlsx, new_path)
    assert new_path.exists()


# ---------------------------------------------------------------------------
# 12. Auto-trigger — should_auto_stream returns False for small sheets.
# ---------------------------------------------------------------------------


def test_no_auto_trigger_below_threshold(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx)  # read_only=False
    ws = wb["Sheet1"]
    assert should_auto_stream(ws) is False


# ---------------------------------------------------------------------------
# 13. read_only=True surfaces the explicit flag on the workbook.
# ---------------------------------------------------------------------------


def test_workbook_read_only_property(basic_xlsx: Path) -> None:
    wb_r = wolfxl.load_workbook(basic_xlsx, read_only=True)
    assert wb_r.read_only is True
    wb_e = wolfxl.load_workbook(basic_xlsx)
    assert wb_e.read_only is False


# ---------------------------------------------------------------------------
# 14. min_col / max_col bounds are honored.
# ---------------------------------------------------------------------------


def test_streaming_column_bounds(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(
        ws.iter_rows(min_row=1, max_row=2, min_col=2, max_col=4, values_only=True)
    )
    assert rows == [(12, 13, 14), (22, 23, 24)]


# ---------------------------------------------------------------------------
# 15. Coordinate / row / column on streaming cells match openpyxl.
# ---------------------------------------------------------------------------


def test_streaming_cell_coords(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, min_col=3, max_col=3))
    cell = rows[0][0]
    assert cell.row == 2
    assert cell.column == 3
    assert cell.column_letter == "C"
    assert cell.coordinate == "C2"


# ---------------------------------------------------------------------------
# 16. Iteration yields exactly one tuple per `<row>` element present.
# ---------------------------------------------------------------------------


def test_streaming_yields_per_row_element(basic_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(basic_xlsx, read_only=True)
    ws = wb["Sheet1"]
    count = 0
    for _ in ws.iter_rows(values_only=True):
        count += 1
    assert count == 1000


# ---------------------------------------------------------------------------
# 17. Sprint Λ Pod-γ — datetime cells convert via the styles table.
# A `<c s="N">` whose number format passes ``is_date_format`` must surface
# as a ``datetime`` (or ``date``/``time``) — not as the raw Excel serial.
# ---------------------------------------------------------------------------


def test_streaming_datetime_yields_datetime_values_only(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(
        ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2, values_only=True)
    )
    a2, b2 = rows[0]
    assert isinstance(a2, dt.datetime), f"A2 datetime divergence: {type(a2).__name__} {a2!r}"
    assert a2 == dt.datetime(2024, 1, 15, 12, 30)
    # B2 is a date — openpyxl read_only path returns a datetime at midnight.
    assert isinstance(b2, dt.datetime), f"B2 date divergence: {type(b2).__name__} {b2!r}"
    assert b2 == dt.datetime(2024, 6, 1, 0, 0)


def test_streaming_datetime_yields_datetime_via_streaming_cell(mixed_xlsx: Path) -> None:
    wb = wolfxl.load_workbook(mixed_xlsx, read_only=True)
    ws = wb["Mixed"]
    rows = list(ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2))
    a2_cell, b2_cell = rows[0]
    assert isinstance(a2_cell.value, dt.datetime), (
        f"StreamingCell.value (A2) divergence: {type(a2_cell.value).__name__} {a2_cell.value!r}"
    )
    assert a2_cell.value == dt.datetime(2024, 1, 15, 12, 30)
    assert isinstance(b2_cell.value, dt.datetime), (
        f"StreamingCell.value (B2) divergence: {type(b2_cell.value).__name__} {b2_cell.value!r}"
    )
    assert b2_cell.value == dt.datetime(2024, 6, 1, 0, 0)
