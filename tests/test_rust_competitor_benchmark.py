from __future__ import annotations

import importlib.util
import sys
from pathlib import Path
from types import ModuleType, SimpleNamespace


def _load_module() -> ModuleType:
    scripts_dir = Path(__file__).resolve().parents[1] / "scripts"
    sys.path.insert(0, str(scripts_dir))
    script = scripts_dir / "benchmark_rust_competitors.py"
    spec = importlib.util.spec_from_file_location("benchmark_rust_competitors", script)
    assert spec is not None
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


bench = _load_module()


def test_umya_temp_project_uses_current_claim_version(tmp_path: Path) -> None:
    project_dir = tmp_path / "umya"

    bench._write_umya_project(project_dir)

    assert 'umya-spreadsheet = "3.0.0"' in (
        project_dir / "Cargo.toml"
    ).read_text(encoding="utf-8")
    main_rs = (project_dir / "src" / "main.rs").read_text(encoding="utf-8")
    assert "sheet_by_name_mut" in main_rs
    assert ".ok_or(" not in main_rs


def test_rust_competitor_comparison_rows_use_competitor_seconds_over_wolfxl() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_cell",
                "median_seconds": 2.0,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 3.0,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "write_cell_by_cell_plain",
            "wolfxl_engine": "wolfxl_cell",
            "wolfxl_api_surface": "python_public_api",
            "competitor": "rust_xlsxwriter",
            "competitor_engine": "rust_xlsxwriter",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": False,
            "speedup_vs_competitor": 1.5,
        }
    ]


def test_rust_competitor_comparison_rows_carry_competitor_version() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_write_rows",
                "median_seconds": 1.0,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "competitor_version": "0.95.0",
                "median_seconds": 2.0,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["competitor_version"] == "0.95.0"


def test_rust_competitor_comparison_rows_ignore_cases_without_wolfxl() -> None:
    assert (
        bench.rust_competitor_comparison_rows(
            [
                {
                    "case": "read_values_plain",
                    "engine": "calamine",
                    "competitor": "calamine",
                    "median_seconds": 1.0,
                }
            ]
        )
        == []
    )


def test_rust_competitor_comparison_rows_accept_read_engine_name() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl",
                "median_seconds": 1.0,
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 0.5,
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl"
    assert rows[0]["competitor"] == "calamine"
    assert rows[0]["speedup_vs_competitor"] == 0.5


def test_rust_competitor_comparison_rows_prefers_read_only_baseline() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl",
                "median_seconds": 2.0,
            },
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only",
                "median_seconds": 1.0,
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 0.5,
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_read_only"
    assert rows[0]["speedup_vs_competitor"] == 0.5


def test_rust_competitor_comparison_rows_prefers_chunked_read_baseline() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only",
                "median_seconds": 1.0,
            },
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only_chunks",
                "median_seconds": 0.75,
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 0.5,
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_read_only_chunks"
    assert rows[0]["speedup_vs_competitor"] == 0.5 / 0.75


def test_rust_competitor_comparison_rows_prefers_public_value_summary() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl_value_summary",
                "median_seconds": 0.2,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only_chunks",
                "median_seconds": 0.75,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 0.5,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_value_summary"
    assert rows[0]["speedup_vs_competitor"] == 0.5 / 0.2


def test_rust_competitor_comparison_rows_prefers_public_formula_summary() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_formula_text",
                "engine": "wolfxl_formula_summary",
                "median_seconds": 0.2,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_formula_text",
                "engine": "wolfxl",
                "median_seconds": 0.75,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_formula_text",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 0.5,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_formula_summary"
    assert rows[0]["speedup_vs_competitor"] == 0.5 / 0.2


def test_direct_reader_diagnostic_does_not_replace_public_claim_baseline() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only",
                "median_seconds": 2.0,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_values_plain",
                "engine": "wolfxl_read_only_chunks",
                "median_seconds": 1.5,
                "api_surface": "python_public_api",
            },
            {
                "case": "read_values_plain",
                "engine": "wolfxl_reader_direct",
                "median_seconds": 0.25,
                "api_surface": "direct_rust_api",
                "diagnostic": True,
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 1.0,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_read_only_chunks"
    assert rows[0]["wolfxl_api_surface"] == "python_public_api"
    assert rows[0]["speedup_vs_competitor"] == 1.0 / 1.5


def test_modify_phase_diagnostic_does_not_replace_public_claim_baseline() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "modify_existing_workbook_plain",
                "engine": "wolfxl_modify",
                "median_seconds": 2.0,
                "api_surface": "python_public_api",
            },
            {
                "case": "modify_existing_workbook_plain",
                "engine": "wolfxl_modify_internal_phases",
                "median_seconds": 0.25,
                "api_surface": "python_public_api",
                "diagnostic": True,
            },
            {
                "case": "modify_existing_workbook_plain",
                "engine": "office_oxide",
                "competitor": "office_oxide",
                "median_seconds": 1.0,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_modify"
    assert rows[0]["speedup_vs_competitor"] == 0.5


def test_wolfxl_direct_core_comparison_rows_use_direct_rust_only() -> None:
    rows = bench.wolfxl_direct_core_comparison_rows(
        [
            {
                "case": "read_values_plain",
                "engine": "wolfxl_reader_direct",
                "median_seconds": 0.25,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 1.0,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "read_values_plain",
                "engine": "python_calamine",
                "competitor": "python-calamine",
                "median_seconds": 2.0,
                "api_surface": "python_binding_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "read_values_plain",
            "wolfxl_engine": "wolfxl_reader_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "calamine",
            "competitor_engine": "calamine",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "speedup_vs_competitor": 4.0,
        }
    ]


def test_wolfxl_direct_reader_harness_counts_rows_during_checksum_pass() -> None:
    main = bench.WOLFXL_READER_MAIN

    assert "read_first_sheet_value_summary_path(path)?" in main
    assert "read_first_sheet_formula_summary_path(path)?" in main
    assert 'case_name == "read_formula_text"' in main
    assert "summary.numeric_checksum" in main
    assert "summary.formula_cells" in main
    assert "read_first_sheet_value_cells_path(path)?" not in main
    assert "cells.iter().map(|cell| cell.row as usize).max()" not in main


def test_wolfxl_direct_writer_comparison_rows_use_direct_rust_only() -> None:
    rows = bench.wolfxl_direct_writer_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_writer_direct",
                "median_seconds": 0.5,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 1.0,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "read_values_plain",
                "engine": "calamine",
                "competitor": "calamine",
                "median_seconds": 2.0,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "write_cell_by_cell_plain",
            "wolfxl_engine": "wolfxl_writer_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "rust_xlsxwriter",
            "competitor_engine": "rust_xlsxwriter",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "speedup_vs_competitor": 2.0,
        }
    ]


def test_wolfxl_streaming_direct_writer_memory_rows_use_direct_rust_only() -> None:
    rows = bench.wolfxl_streaming_direct_writer_memory_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_writer_streaming_direct",
                "peak_rss_bytes": 8_000_000,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "peak_rss_bytes": 6_000_000,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "python_backed_writer",
                "competitor": "python_backed_writer",
                "peak_rss_bytes": 10_000_000,
                "api_surface": "python_binding_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "write_cell_by_cell_plain",
            "wolfxl_engine": "wolfxl_writer_streaming_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "rust_xlsxwriter",
            "competitor_engine": "rust_xlsxwriter",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "wolfxl_peak_rss_bytes": 8_000_000,
            "competitor_peak_rss_bytes": 6_000_000,
            "peak_rss_excess_bytes": 2_000_000,
            "peak_rss_ratio_vs_competitor": 8_000_000 / 6_000_000,
        }
    ]


def test_rust_competitor_comparison_rows_uses_fastest_public_write_api() -> None:
    rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_cell",
                "median_seconds": 3.0,
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_append",
                "median_seconds": 1.5,
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_write_rows",
                "median_seconds": 1.0,
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 2.0,
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_write_rows"
    assert rows[0]["speedup_vs_competitor"] == 2.0

    styled_rows = bench.rust_competitor_comparison_rows(
        [
            {
                "case": "write_styled_cells",
                "engine": "wolfxl_write_styled_rows",
                "median_seconds": 3.0,
            },
            {
                "case": "write_styled_cells",
                "engine": "wolfxl_write_styled_rows_grid",
                "median_seconds": 1.0,
            },
            {
                "case": "write_styled_cells",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 2.0,
            },
        ]
    )

    assert styled_rows[0]["wolfxl_engine"] == "wolfxl_write_styled_rows_grid"
    assert styled_rows[0]["speedup_vs_competitor"] == 2.0


def test_rust_competitor_memory_rows_use_claim_facing_wolfxl_engine() -> None:
    rows = bench.rust_competitor_memory_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_cell",
                "median_seconds": 3.0,
                "peak_rss_bytes": 90_000_000,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_write_rows",
                "median_seconds": 1.0,
                "peak_rss_bytes": 60_000_000,
                "delta_rss_bytes": 12_000_000,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 2.0,
                "peak_rss_bytes": 100_000_000,
                "delta_rss_bytes": 20_000_000,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "write_cell_by_cell_plain",
            "wolfxl_engine": "wolfxl_write_rows",
            "wolfxl_api_surface": "python_public_api",
            "competitor": "rust_xlsxwriter",
            "competitor_engine": "rust_xlsxwriter",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": False,
            "wolfxl_peak_rss_bytes": 60_000_000,
            "competitor_peak_rss_bytes": 100_000_000,
            "peak_rss_excess_bytes": -40_000_000,
            "peak_rss_ratio_vs_competitor": 0.6,
            "wolfxl_delta_rss_bytes": 12_000_000,
            "competitor_delta_rss_bytes": 20_000_000,
            "delta_rss_excess_bytes": -8_000_000,
            "delta_rss_ratio_vs_competitor": 0.6,
        }
    ]


def test_rust_competitor_memory_rows_skip_unmeasured_engines() -> None:
    assert (
        bench.rust_competitor_memory_comparison_rows(
            [
                {
                    "case": "read_values_plain",
                    "engine": "wolfxl_value_summary",
                    "median_seconds": 1.0,
                },
                {
                    "case": "read_values_plain",
                    "engine": "calamine",
                    "competitor": "calamine",
                    "median_seconds": 0.5,
                    "peak_rss_bytes": 100_000_000,
                },
            ]
        )
        == []
    )


def test_rust_competitor_memory_rows_fall_back_to_measured_claim_engine() -> None:
    rows = bench.rust_competitor_memory_comparison_rows(
        [
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_append",
                "median_seconds": 0.5,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "wolfxl_write_rows_fast_plain",
                "median_seconds": 1.0,
                "peak_rss_bytes": 50_000_000,
                "api_surface": "python_public_api",
            },
            {
                "case": "write_cell_by_cell_plain",
                "engine": "rust_xlsxwriter",
                "competitor": "rust_xlsxwriter",
                "median_seconds": 2.0,
                "peak_rss_bytes": 25_000_000,
                "api_surface": "direct_rust_api",
            },
        ]
    )

    assert rows[0]["wolfxl_engine"] == "wolfxl_write_rows_fast_plain"
    assert rows[0]["peak_rss_ratio_vs_competitor"] == 2.0


def test_time_output_parsers_read_child_peak_rss() -> None:
    assert (
        bench._parse_macos_time_peak_rss(
            "  12345678  maximum resident set size\n"
        )
        == 12_345_678
    )
    assert (
        bench._parse_gnu_time_peak_rss(
            "Maximum resident set size (kbytes): 12345\n"
        )
        == 12_345 * 1024
    )


def test_parse_args_defaults_to_all_required_competitors() -> None:
    args = bench.parse_args([])

    assert args.competitors == list(bench.DEFAULT_RUST_COMPETITORS)


def test_parse_args_accepts_rustypyxl_watchlist_competitor() -> None:
    args = bench.parse_args(["--competitor", "rustypyxl"])

    assert args.competitors == ["rustypyxl"]


def test_parse_args_accepts_office_oxide_watchlist_competitor() -> None:
    args = bench.parse_args(["--competitor", "office_oxide"])

    assert args.competitors == ["office_oxide"]


def test_parse_args_accepts_wolfxl_rust_source_root(tmp_path: Path) -> None:
    args = bench.parse_args(["--wolfxl-rust-source-root", str(tmp_path)])

    assert args.wolfxl_rust_source_root == tmp_path


def test_wolfxl_direct_projects_use_configured_source_root(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_root = tmp_path / "source"
    project_dir = tmp_path / "project"
    monkeypatch.setattr(bench, "WOLFXL_RUST_SOURCE_ROOT", source_root)

    bench._write_wolfxl_writer_project(project_dir)

    cargo_toml = (project_dir / "Cargo.toml").read_text(encoding="utf-8")
    assert f'path = "{(source_root / "crates" / "wolfxl-writer").as_posix()}"' in cargo_toml


def test_plain_read_checksum_covers_all_numeric_columns() -> None:
    # row 1: 1 + 1.25 + 1, row 2: 2 + 2.5 + 2
    assert bench.plain_numeric_checksum(rows=2, cols=5) == 9.75
    assert bench.plain_numeric_checksum(rows=2, cols=1) == 3.0


def test_cargo_lock_package_version_reads_resolved_dependency(tmp_path: Path) -> None:
    lock = tmp_path / "Cargo.lock"
    lock.write_text(
        """
[[package]]
name = "wolfxl-bench"
version = "0.1.0"

[[package]]
name = "rust_xlsxwriter"
version = "0.95.0"
""".strip()
        + "\n",
        encoding="utf-8",
    )

    assert bench._cargo_lock_package_version(tmp_path, "rust_xlsxwriter") == "0.95.0"


def test_run_benchmark_marks_unrequested_competitors_as_not_requested(monkeypatch) -> None:
    def fake_writer_result(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "rust_xlsxwriter",
            "competitor": "rust_xlsxwriter",
            "median_seconds": 1.0,
        }

    monkeypatch.setattr(
        bench,
        "bench_write_phased",
        lambda *_args, **_kwargs: {
            "case": "write_cell_by_cell_plain",
            "engine": "wolfxl_cell",
            "median_seconds": 2.0,
        },
    )
    monkeypatch.setattr(bench, "bench_rust_xlsxwriter", fake_writer_result)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["rust_xlsxwriter"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    skipped = {item["competitor"]: item["reason"] for item in payload["skipped_competitors"]}
    assert skipped["xlsxwriter"] == "not requested in this run"
    assert skipped["calamine"] == "not requested in this run"
    assert skipped["fastexcel"] == "not requested in this run"


def test_run_benchmark_adds_direct_reader_diagnostic_for_calamine(monkeypatch) -> None:
    def fake_public_reader(engine: str = "wolfxl_read_only") -> dict[str, object]:
        return {
            "case": "read_values_plain",
            "engine": engine,
            "median_seconds": 2.0,
        }

    def fake_direct_reader(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "wolfxl_reader_direct",
            "median_seconds": 0.25,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    def fake_calamine(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "calamine",
            "competitor": "calamine",
            "median_seconds": 1.0,
            "api_surface": "direct_rust_api",
        }

    def fake_formula_reader(**_kwargs):
        return {
            "case": "read_formula_text",
            "engine": "calamine",
            "competitor": "calamine",
            "median_seconds": 1.0,
            "api_surface": "direct_rust_api",
        }

    def fake_direct_formula_reader(**_kwargs):
        return {
            "case": "read_formula_text",
            "engine": "wolfxl_reader_direct",
            "median_seconds": 0.5,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    monkeypatch.setattr(bench, "build_openpyxl_plain", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(bench, "build_openpyxl_formula", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        bench,
        "bench_read_phased",
        lambda _case, engine, *_args, **_kwargs: fake_public_reader(engine),
    )
    monkeypatch.setattr(
        bench,
        "bench_read_checked",
        lambda _case, engine, *_args, **_kwargs: {
            "case": _case,
            "engine": engine,
            "median_seconds": 2.0,
        },
    )
    monkeypatch.setattr(
        bench,
        "bench_wolfxl_reader_direct_read_values",
        fake_direct_reader,
    )
    monkeypatch.setattr(
        bench,
        "bench_wolfxl_reader_direct_read_formulas",
        fake_direct_formula_reader,
    )
    monkeypatch.setattr(bench, "bench_calamine_read_values", fake_calamine)
    monkeypatch.setattr(bench, "bench_calamine_read_formulas", fake_formula_reader)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["calamine"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    read_values = [
        row for row in payload["comparisons"] if row["case"] == "read_values_plain"
    ]
    assert read_values[0]["wolfxl_engine"] == "wolfxl_value_summary"
    assert payload["wolfxl_direct_core_comparisons"] == [
        {
            "case": "read_formula_text",
            "wolfxl_engine": "wolfxl_reader_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "calamine",
            "competitor_engine": "calamine",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "speedup_vs_competitor": 2.0,
        },
        {
            "case": "read_values_plain",
            "wolfxl_engine": "wolfxl_reader_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "calamine",
            "competitor_engine": "calamine",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "speedup_vs_competitor": 4.0,
        }
    ]


def test_run_benchmark_can_measure_rustypyxl_watchlist_reader(monkeypatch) -> None:
    def fake_public_reader(engine: str = "wolfxl_read_only") -> dict[str, object]:
        return {
            "case": "read_values_plain",
            "engine": engine,
            "median_seconds": 1.0,
        }

    def fake_optional_reader(**kwargs):
        return {
            "case": kwargs["case"],
            "engine": kwargs["engine"],
            "competitor": kwargs["competitor"],
            "median_seconds": 2.0,
            "api_surface": bench.PYTHON_BINDING_API,
            "competitor_version": "0.3.1",
        }

    monkeypatch.setattr(bench, "_optional_module_available", lambda module: module == "rustypyxl")
    monkeypatch.setattr(bench, "build_openpyxl_plain", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        bench,
        "bench_read_phased",
        lambda _case, engine, *_args, **_kwargs: fake_public_reader(engine),
    )
    monkeypatch.setattr(
        bench,
        "bench_read_checked",
        lambda _case, engine, *_args, **_kwargs: fake_public_reader(engine),
    )
    monkeypatch.setattr(bench, "bench_optional_python_reader", fake_optional_reader)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["rustypyxl"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    comparison = payload["comparisons"][0]
    assert comparison["competitor"] == "rustypyxl"
    assert comparison["wolfxl_api_surface"] == bench.PYTHON_PUBLIC_API
    assert comparison["competitor_api_surface"] == bench.PYTHON_BINDING_API
    assert comparison["speedup_vs_competitor"] == 2.0


def test_run_benchmark_can_measure_office_oxide_watchlist_lanes(monkeypatch) -> None:
    def fake_public_reader(engine: str = "wolfxl_read_only") -> dict[str, object]:
        return {
            "case": "read_values_plain",
            "engine": engine,
            "median_seconds": 1.0,
            "api_surface": bench.PYTHON_PUBLIC_API,
        }

    def fake_direct_reader(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "wolfxl_reader_direct",
            "median_seconds": 0.5,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    def fake_public_writer(case: str = "write_cell_by_cell_plain"):
        return {
            "case": case,
            "engine": "wolfxl_write_rows_fast_plain",
            "median_seconds": 1.0,
            "api_surface": bench.PYTHON_PUBLIC_API,
        }

    def fake_direct_writer(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "wolfxl_writer_direct",
            "median_seconds": 0.25,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    def fake_public_modify(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify",
            "median_seconds": 1.0,
            "api_surface": bench.PYTHON_PUBLIC_API,
        }

    def fake_public_modify_phases(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_internal_phases",
            "median_seconds": 1.1,
            "api_surface": bench.PYTHON_PUBLIC_API,
            "diagnostic": True,
            "phase_medians_seconds": {
                "load_workbook_modify": 0.4,
                "assign_cells": 0.01,
                "save": 0.69,
            },
        }

    def fake_direct_modify(**_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_direct",
            "median_seconds": 0.5,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    def fake_office_write_or_modify(**kwargs):
        return {
            "case": kwargs["case"],
            "engine": "office_oxide",
            "competitor": "office_oxide",
            "competitor_version": "0.1.2",
            "median_seconds": 2.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    def fake_office_read(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "office_oxide",
            "competitor": "office_oxide",
            "competitor_version": "0.1.2",
            "median_seconds": 2.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    monkeypatch.setattr(bench, "build_openpyxl_plain", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        bench,
        "bench_read_phased",
        lambda _case, engine, *_args, **_kwargs: fake_public_reader(engine),
    )
    monkeypatch.setattr(
        bench,
        "bench_read_checked",
        lambda _case, engine, *_args, **_kwargs: fake_public_reader(engine),
    )
    monkeypatch.setattr(
        bench,
        "bench_write",
        lambda case, *_args, **_kwargs: fake_public_writer(case),
    )
    monkeypatch.setattr(
        bench,
        "bench_write_phased",
        lambda case, *_args, **_kwargs: fake_public_writer(case),
    )
    monkeypatch.setattr(bench, "bench_wolfxl_reader_direct_read_values", fake_direct_reader)
    monkeypatch.setattr(bench, "bench_wolfxl_writer_direct", fake_direct_writer)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_existing_workbook", fake_public_modify)
    monkeypatch.setattr(
        bench,
        "bench_wolfxl_modify_internal_phases",
        fake_public_modify_phases,
    )
    monkeypatch.setattr(bench, "bench_wolfxl_modify_direct", fake_direct_modify)
    monkeypatch.setattr(bench, "bench_office_oxide_read_values", fake_office_read)
    monkeypatch.setattr(
        bench,
        "bench_office_oxide_write_or_modify",
        fake_office_write_or_modify,
    )
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["office_oxide"],
            rows=10,
            cols=2,
            rounds=1,
            cases=[
                "read_values_plain",
                "write_cell_by_cell_plain",
                "modify_existing_workbook_plain",
            ],
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    office_cases = {
        result["case"]
        for result in payload["results"]
        if result.get("competitor") == "office_oxide"
    }
    assert office_cases == {
        "read_values_plain",
        "write_cell_by_cell_plain",
        "modify_existing_workbook_plain",
    }
    assert payload["missing_competitors"] == list(bench.DEFAULT_RUST_COMPETITORS)
    assert {
        item["competitor"]: item["reason"]
        for item in payload["skipped_competitors"]
        if item["competitor"] in bench.DEFAULT_RUST_COMPETITORS
    }["calamine"] == "not requested in this run"
    assert any(
        comparison["competitor"] == "office_oxide"
        and comparison["competitor_api_surface"] == bench.DIRECT_RUST_API
        for comparison in payload["comparisons"]
    )
    assert any(
        comparison["competitor"] == "office_oxide"
        for comparison in payload["wolfxl_direct_core_comparisons"]
    )
    assert any(
        comparison["competitor"] == "office_oxide"
        for comparison in payload["wolfxl_direct_writer_comparisons"]
    )
    assert any(
        result["engine"] == "wolfxl_modify_internal_phases"
        and result.get("diagnostic") is True
        for result in payload["results"]
    )
    assert any(
        comparison["competitor"] == "office_oxide"
        for comparison in payload["wolfxl_direct_modify_comparisons"]
    )


def test_run_benchmark_adds_direct_writer_diagnostic_for_rust_writer(monkeypatch) -> None:
    def fake_public_writer(case: str = "write_cell_by_cell_plain"):
        engine = {
            "write_formula_cells": "wolfxl_append",
            "write_styled_cells": "wolfxl_styled",
            "write_multi_sheet_plain": "wolfxl_multi_sheet",
        }.get(case, "wolfxl_write_rows")
        return {
            "case": case,
            "engine": engine,
            "median_seconds": 2.0,
        }

    def fake_direct_writer(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "wolfxl_writer_direct",
            "median_seconds": 0.5,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    def fake_streaming_direct_writer(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "wolfxl_writer_streaming_direct",
            "median_seconds": 0.4,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    def fake_rust_xlsxwriter(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "rust_xlsxwriter",
            "competitor": "rust_xlsxwriter",
            "median_seconds": 1.0,
            "api_surface": "direct_rust_api",
        }

    monkeypatch.setattr(
        bench,
        "bench_write_phased",
        lambda case, *_args, **_kwargs: fake_public_writer(case),
    )
    monkeypatch.setattr(
        bench,
        "bench_write",
        lambda case, *_args, **_kwargs: fake_public_writer(case),
    )
    monkeypatch.setattr(bench, "bench_wolfxl_writer_direct", fake_direct_writer)
    monkeypatch.setattr(
        bench,
        "bench_wolfxl_writer_streaming_direct",
        fake_streaming_direct_writer,
    )
    monkeypatch.setattr(bench, "bench_rust_xlsxwriter", fake_rust_xlsxwriter)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["rust_xlsxwriter"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    comparison_cases = {row["case"] for row in payload["comparisons"]}
    assert comparison_cases >= set(bench.RUST_WRITER_REQUIRED_CASES)
    direct_cases = {row["case"] for row in payload["wolfxl_direct_writer_comparisons"]}
    assert direct_cases >= set(bench.RUST_WRITER_REQUIRED_CASES)
    streaming_direct_cases = {
        row["case"] for row in payload["wolfxl_streaming_direct_writer_comparisons"]
    }
    assert streaming_direct_cases >= set(bench.RUST_WRITER_REQUIRED_CASES)
    assert {
        result["case"]
        for result in payload["results"]
        if result["engine"] == "wolfxl_writer_streaming_direct"
    } >= set(bench.RUST_WRITER_REQUIRED_CASES)


def test_run_benchmark_filters_rust_writer_cases(monkeypatch) -> None:
    def fake_public_writer(case: str = "write_cell_by_cell_plain", *_args, **_kwargs):
        return {
            "case": case,
            "engine": "wolfxl_styled",
            "median_seconds": 2.0,
        }

    def fake_direct_writer(**kwargs):
        return {
            "case": kwargs["case"],
            "engine": "wolfxl_writer_direct",
            "median_seconds": 0.5,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    def fake_streaming_direct_writer(**kwargs):
        return {
            "case": kwargs["case"],
            "engine": "wolfxl_writer_streaming_direct",
            "median_seconds": 0.4,
            "api_surface": "direct_rust_api",
            "diagnostic": True,
        }

    def fake_rust_xlsxwriter(**kwargs):
        return {
            "case": kwargs["case"],
            "engine": "rust_xlsxwriter",
            "competitor": "rust_xlsxwriter",
            "median_seconds": 1.0,
            "api_surface": "direct_rust_api",
        }

    monkeypatch.setattr(bench, "bench_write_phased", fake_public_writer)
    monkeypatch.setattr(bench, "bench_write", fake_public_writer)
    monkeypatch.setattr(bench, "bench_wolfxl_writer_direct", fake_direct_writer)
    monkeypatch.setattr(
        bench,
        "bench_wolfxl_writer_streaming_direct",
        fake_streaming_direct_writer,
    )
    monkeypatch.setattr(bench, "bench_rust_xlsxwriter", fake_rust_xlsxwriter)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["rust_xlsxwriter"],
            rows=10,
            cols=2,
            rounds=1,
            cases=["write_styled_cells"],
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    result_cases = {row["case"] for row in payload["results"]}
    comparison_cases = {row["case"] for row in payload["comparisons"]}
    direct_cases = {row["case"] for row in payload["wolfxl_direct_writer_comparisons"]}
    streaming_direct_cases = {
        row["case"] for row in payload["wolfxl_streaming_direct_writer_comparisons"]
    }
    assert result_cases == {"write_styled_cells"}
    assert comparison_cases == {"write_styled_cells"}
    assert direct_cases == {"write_styled_cells"}
    assert streaming_direct_cases == {"write_styled_cells"}
    assert payload["parameters"]["cases"] == ["write_styled_cells"]


def test_run_benchmark_adds_umya_modify_comparison(monkeypatch) -> None:
    def fake_public_writer(*_args, **_kwargs):
        return {
            "case": "write_cell_by_cell_plain",
            "engine": "wolfxl_write_rows",
            "median_seconds": 2.0,
        }

    def fake_modify_baseline(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify",
            "median_seconds": 2.0,
        }

    def fake_modify_phases(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_internal_phases",
            "median_seconds": 2.1,
            "diagnostic": True,
        }

    def fake_modify_direct(**_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_direct",
            "median_seconds": 1.0,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    def fake_umya(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "umya-spreadsheet",
            "competitor": "umya-spreadsheet",
            "median_seconds": 4.0,
            "api_surface": "direct_rust_api",
        }

    def fake_umya_read(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "umya-spreadsheet",
            "competitor": "umya-spreadsheet",
            "median_seconds": 4.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    def fake_read_baseline(results, **_kwargs):
        results.append(
            {
                "case": "read_values_plain",
                "engine": "wolfxl_value_summary",
                "median_seconds": 2.0,
                "api_surface": bench.PYTHON_PUBLIC_API,
            }
        )

    def fake_read_direct(*_args, **_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "wolfxl_reader_direct",
            "median_seconds": 1.0,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    monkeypatch.setattr(bench, "bench_write_phased", fake_public_writer)
    monkeypatch.setattr(bench, "bench_write", fake_public_writer)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_existing_workbook", fake_modify_baseline)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_internal_phases", fake_modify_phases)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_direct", fake_modify_direct)
    monkeypatch.setattr(bench, "build_openpyxl_plain", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(bench, "_append_wolfxl_read_only_baseline", fake_read_baseline)
    monkeypatch.setattr(bench, "bench_wolfxl_reader_direct_read_values", fake_read_direct)
    monkeypatch.setattr(bench, "bench_wolfxl_writer_direct", fake_public_writer)
    monkeypatch.setattr(bench, "bench_umya_write_values", fake_umya)
    monkeypatch.setattr(bench, "bench_umya_read_values", fake_umya_read)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["umya-spreadsheet"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    modify_rows = [
        row for row in payload["comparisons"] if row["case"] == "modify_existing_workbook_plain"
    ]
    read_rows = [row for row in payload["comparisons"] if row["case"] == "read_values_plain"]
    assert read_rows[0]["competitor"] == "umya-spreadsheet"
    assert read_rows[0]["competitor_api_surface"] == bench.DIRECT_RUST_API
    assert modify_rows[0]["wolfxl_engine"] == "wolfxl_modify"
    assert modify_rows[0]["competitor"] == "umya-spreadsheet"
    assert payload["wolfxl_direct_modify_comparisons"] == [
        {
            "case": "modify_existing_workbook_plain",
            "wolfxl_engine": "wolfxl_modify_direct",
            "wolfxl_api_surface": bench.DIRECT_RUST_API,
            "competitor": "umya-spreadsheet",
            "competitor_engine": "umya-spreadsheet",
            "competitor_api_surface": bench.DIRECT_RUST_API,
            "same_api_surface": True,
            "speedup_vs_competitor": 4.0,
        }
    ]


def test_run_benchmark_adds_modify_fullrewrite_comparison(monkeypatch) -> None:
    """run_benchmark wires the same-algorithm full-rewrite lane end to end.

    The unit tests above cover the pure comparison-row helper and the case set;
    this exercises run_benchmark itself so the new lane and its
    `wolfxl_direct_modify_fullrewrite_comparisons` payload key cannot silently
    stop populating. Mirrors test_run_benchmark_adds_umya_modify_comparison.
    """

    def fake_public_writer(*_args, **_kwargs):
        return {
            "case": "write_cell_by_cell_plain",
            "engine": "wolfxl_write_rows",
            "median_seconds": 2.0,
        }

    def fake_modify_baseline(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify",
            "median_seconds": 2.0,
        }

    def fake_modify_phases(*_args, **_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_internal_phases",
            "median_seconds": 2.1,
            "diagnostic": True,
        }

    def fake_modify_direct(**_kwargs):
        return {
            "case": "modify_existing_workbook_plain",
            "engine": "wolfxl_modify_direct",
            "median_seconds": 1.0,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    def fake_modify_fullrewrite(**kwargs):
        # Echo the case so plain/wide/large each produce a distinct result row.
        return {
            "case": kwargs.get("case", "modify_fullrewrite_plain"),
            "engine": "wolfxl_modify_fullrewrite_direct",
            "median_seconds": 1.0,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    office_oxide_calls: list[str] = []

    def fake_office_oxide_fullrewrite(**kwargs):
        # Fix #3: prove office_oxide's full-rewrite row comes from the runner
        # driving bench_office_oxide_modify_fullrewrite through the append path,
        # not from a result hand-injected into the pure comparison-rows helper.
        case = kwargs.get("case", "modify_fullrewrite_plain")
        office_oxide_calls.append(case)
        return {
            "case": case,
            "engine": "office_oxide",
            "competitor": "office_oxide",
            "median_seconds": 2.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    def fake_umya(**kwargs):
        return {
            "case": kwargs.get("case", "write_cell_by_cell_plain"),
            "engine": "umya-spreadsheet",
            "competitor": "umya-spreadsheet",
            "median_seconds": 4.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    def fake_umya_read(**_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "umya-spreadsheet",
            "competitor": "umya-spreadsheet",
            "median_seconds": 4.0,
            "api_surface": bench.DIRECT_RUST_API,
        }

    def fake_read_baseline(results, **_kwargs):
        results.append(
            {
                "case": "read_values_plain",
                "engine": "wolfxl_value_summary",
                "median_seconds": 2.0,
                "api_surface": bench.PYTHON_PUBLIC_API,
            }
        )

    def fake_read_direct(*_args, **_kwargs):
        return {
            "case": "read_values_plain",
            "engine": "wolfxl_reader_direct",
            "median_seconds": 1.0,
            "api_surface": bench.DIRECT_RUST_API,
            "diagnostic": True,
        }

    monkeypatch.setattr(bench, "bench_write_phased", fake_public_writer)
    monkeypatch.setattr(bench, "bench_write", fake_public_writer)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_existing_workbook", fake_modify_baseline)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_internal_phases", fake_modify_phases)
    monkeypatch.setattr(bench, "bench_wolfxl_modify_direct", fake_modify_direct)
    monkeypatch.setattr(
        bench, "bench_wolfxl_modify_fullrewrite_direct", fake_modify_fullrewrite
    )
    monkeypatch.setattr(
        bench, "bench_office_oxide_modify_fullrewrite", fake_office_oxide_fullrewrite
    )
    monkeypatch.setattr(bench, "build_openpyxl_plain", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(bench, "_append_wolfxl_read_only_baseline", fake_read_baseline)
    monkeypatch.setattr(bench, "bench_wolfxl_reader_direct_read_values", fake_read_direct)
    monkeypatch.setattr(bench, "bench_wolfxl_writer_direct", fake_public_writer)
    monkeypatch.setattr(bench, "bench_umya_write_values", fake_umya)
    monkeypatch.setattr(bench, "bench_umya_read_values", fake_umya_read)
    monkeypatch.setattr(bench, "metadata", lambda rounds: {"rounds": rounds})

    payload = bench.run_benchmark(
        SimpleNamespace(
            competitors=["umya-spreadsheet"],
            rows=10,
            cols=2,
            rounds=1,
            cache_dir="/tmp/wolfxl-test-cache",
        )
    )

    fullrewrite_rows = payload["wolfxl_direct_modify_fullrewrite_comparisons"]
    cases = {row["case"] for row in fullrewrite_rows}
    assert cases == {
        "modify_fullrewrite_plain",
        "modify_fullrewrite_wide",
        "modify_fullrewrite_large",
    }

    def _row(case: str, competitor: str) -> dict:
        return next(
            row
            for row in fullrewrite_rows
            if row["case"] == case and row["competitor"] == competitor
        )

    assert _row("modify_fullrewrite_plain", "umya-spreadsheet") == {
        "case": "modify_fullrewrite_plain",
        "wolfxl_engine": "wolfxl_modify_fullrewrite_direct",
        "wolfxl_api_surface": bench.DIRECT_RUST_API,
        "competitor": "umya-spreadsheet",
        "competitor_engine": "umya-spreadsheet",
        "competitor_api_surface": bench.DIRECT_RUST_API,
        "same_api_surface": True,
        "speedup_vs_competitor": 4.0,
    }

    # Fix #3: office_oxide's full-rewrite participation is proven at the runner
    # level - the runner called bench_office_oxide_modify_fullrewrite for every
    # case and the resulting rows flow through the real append path, not a result
    # hand-fed to the pure comparison-rows helper.
    assert set(office_oxide_calls) == {
        "modify_fullrewrite_plain",
        "modify_fullrewrite_wide",
        "modify_fullrewrite_large",
    }
    assert _row("modify_fullrewrite_plain", "office_oxide") == {
        "case": "modify_fullrewrite_plain",
        "wolfxl_engine": "wolfxl_modify_fullrewrite_direct",
        "wolfxl_api_surface": bench.DIRECT_RUST_API,
        "competitor": "office_oxide",
        "competitor_engine": "office_oxide",
        "competitor_api_surface": bench.DIRECT_RUST_API,
        "same_api_surface": True,
        "speedup_vs_competitor": 2.0,
    }


def test_wolfxl_modify_rust_comparison_times_modify_phase_only(tmp_path: Path) -> None:
    fixture = tmp_path / "fixture.xlsx"
    bench.build_openpyxl_plain(fixture, rows=5, cols=3)

    def modifier(path: Path) -> str:
        wb = bench.openpyxl.load_workbook(path)
        ws = wb.active
        ws["B2"] = "changed"
        ws["C3"] = 12345
        wb.save(path)
        wb.close()
        return "ok"

    result = bench.bench_wolfxl_modify_existing_workbook(
        "modify_existing_workbook_plain",
        "wolfxl_modify",
        rows=5,
        cols=3,
        rounds=2,
        fixture=fixture,
        modifier=modifier,
        warmup=False,
    )

    assert result["timed_phase"] == "modify_save"
    assert result["samples_seconds"] == result["phase_samples_seconds"]["modify_save"]
    assert set(result["phase_samples_seconds"]) == {
        "modify_save",
        "setup_copy",
        "verify_load",
    }


def test_formula_helpers_match_benchmark_fixture_shape() -> None:
    assert bench._formula_numeric_checksum(3) == 6.0
    assert bench._formula_cell_count(rows=3, cols=5) == 6
    assert bench._formula_cell_count(rows=3, cols=6) == 9


def test_modify_fullrewrite_cases_cover_plain_wide_large() -> None:
    # Distinct case names (not the patcher lane's modify_existing_workbook_*) so
    # one case maps to exactly one algorithm class: the in-place patcher and the
    # full-rewrite competitors never collide in the by_case engine map.
    assert bench.MODIFY_FULLREWRITE_CASES == (
        ("modify_fullrewrite_plain", 1, 1),
        ("modify_fullrewrite_wide", 1, 10),
        ("modify_fullrewrite_large", 10, 1),
    )


def test_modify_fullrewrite_dims_scale_off_base() -> None:
    # plain keeps the base, wide multiplies columns, large multiplies rows.
    assert bench._modify_fullrewrite_dims(2000, 5, 1, 1) == (2000, 5)
    assert bench._modify_fullrewrite_dims(2000, 5, 1, 10) == (2000, 50)
    assert bench._modify_fullrewrite_dims(2000, 5, 10, 1) == (20000, 5)


def test_modify_fullrewrite_wide_fixture_is_genuinely_wide(tmp_path: Path) -> None:
    # Regression: plain_row caps at 5 columns, so the column-scaled wide case
    # (col_factor=10 -> 50 cols) routed through build_openpyxl_plain silently
    # produced a 2000x5 fixture byte-identical to plain -- the wide lane then
    # measured the plain workload twice. The lane must pick build_openpyxl_wide
    # for column-scaled cases so the wide fixture actually has its 50 columns.
    import openpyxl

    assert bench._modify_fullrewrite_fixture_builder(1) is bench.build_openpyxl_plain
    assert bench._modify_fullrewrite_fixture_builder(10) is bench.build_openpyxl_wide

    plain = tmp_path / "plain.xlsx"
    wide = tmp_path / "wide.xlsx"
    bench._modify_fullrewrite_fixture_builder(1)(plain, 8, 5)
    bench._modify_fullrewrite_fixture_builder(10)(wide, 8, 50)
    plain_wb = openpyxl.load_workbook(plain)
    wide_wb = openpyxl.load_workbook(wide)
    try:
        assert plain_wb.active is not None and wide_wb.active is not None
        assert plain_wb.active.max_column == 5
        # The whole point of "wide": genuinely more columns than plain.
        assert wide_wb.active.max_column == 50
    finally:
        plain_wb.close()
        wide_wb.close()


def test_wolfxl_direct_modify_fullrewrite_comparison_rows_use_direct_rust_only() -> None:
    rows = bench.wolfxl_direct_modify_fullrewrite_comparison_rows(
        [
            {
                "case": "modify_fullrewrite_plain",
                "engine": "wolfxl_modify_fullrewrite_direct",
                "median_seconds": 0.25,
                "api_surface": "direct_rust_api",
            },
            {
                "case": "modify_fullrewrite_plain",
                "engine": "umya-spreadsheet",
                "competitor": "umya-spreadsheet",
                "median_seconds": 1.0,
                "api_surface": "direct_rust_api",
            },
            # Public-surface modify must be ignored - it is cross-surface.
            {
                "case": "modify_fullrewrite_plain",
                "engine": "wolfxl_modify",
                "median_seconds": 0.4,
                "api_surface": "python_public_api",
            },
        ]
    )

    assert rows == [
        {
            "case": "modify_fullrewrite_plain",
            "wolfxl_engine": "wolfxl_modify_fullrewrite_direct",
            "wolfxl_api_surface": "direct_rust_api",
            "competitor": "umya-spreadsheet",
            "competitor_engine": "umya-spreadsheet",
            "competitor_api_surface": "direct_rust_api",
            "same_api_surface": True,
            "speedup_vs_competitor": 4.0,
        }
    ]
