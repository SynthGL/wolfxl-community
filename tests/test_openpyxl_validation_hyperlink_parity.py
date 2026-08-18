from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from wolfxl import Workbook, load_workbook
from wolfxl.worksheet.datavalidation import DataValidation
from wolfxl.worksheet.hyperlink import Hyperlink


def test_data_validation_public_range_and_attr_parity() -> None:
    od = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"A,B"',
        allow_blank=True,
        sqref="A1:A2",
        showErrorMessage=True,
        showInputMessage=True,
        error="Bad choice",
        errorTitle="Invalid",
        prompt="Pick one",
        promptTitle="Choice",
    )
    wd = DataValidation(
        type="list",
        formula1='"A,B"',
        allow_blank=True,
        sqref="A1:A2",
        showErrorMessage=True,
        showInputMessage=True,
        error="Bad choice",
        errorTitle="Invalid",
        prompt="Pick one",
        promptTitle="Choice",
    )

    od.add("B2")
    wd.add("B2")
    assert ("A1" in wd, "A1:A2" in wd, "B2" in wd) == (
        "A1" in od,
        "A1:A2" in od,
        "B2" in od,
    )
    assert dict(wd) == dict(od)

    wd.ranges.remove("A1:A2")
    od.ranges.remove("A1:A2")
    assert ("A1" in wd, "B2" in wd, str(wd.sqref)) == (
        "A1" in od,
        "B2" in od,
        str(od.sqref),
    )


def test_ws_add_data_validation_and_reload_prompt_error_fields(tmp_path: Path) -> None:
    out = tmp_path / "dv.xlsx"

    wb = Workbook()
    ws = wb.active
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="10",
        sqref="C2:C5",
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True,
        errorStyle="warning",
        errorTitle="Invalid",
        error="Pick a number from 1 to 10",
        promptTitle="Number",
        prompt="Enter 1-10",
    )
    ws.add_data_validation(dv)
    wb.save(out)

    owb = openpyxl.load_workbook(out)
    [round_tripped] = owb.active.data_validations.dataValidation
    assert round_tripped.type == "whole"
    assert round_tripped.operator == "between"
    assert round_tripped.formula1 == "1"
    assert round_tripped.formula2 == "10"
    assert str(round_tripped.sqref) == "C2:C5"
    assert round_tripped.allowBlank is True
    assert round_tripped.showErrorMessage is True
    assert round_tripped.showInputMessage is True
    assert round_tripped.errorStyle == "warning"
    assert round_tripped.errorTitle == "Invalid"
    assert round_tripped.error == "Pick a number from 1 to 10"
    assert round_tripped.promptTitle == "Number"
    assert round_tripped.prompt == "Enter 1-10"


@pytest.mark.parametrize(
    ("coord", "initial", "assigned", "expected_value", "expected_display"),
    [
        (
            "A1",
            None,
            "https://example.com/string-empty",
            "https://example.com/string-empty",
            None,
        ),
        (
            "A2",
            "before",
            "https://example.com/string-existing",
            "before",
            None,
        ),
        (
            "B1",
            None,
            Hyperlink(ref="Z99", target="https://example.com/object", display="Shown"),
            "https://example.com/object",
            "Shown",
        ),
        (
            "C1",
            None,
            Hyperlink(ref="C1", location="Sheet!A1", display="Inside"),
            "Sheet!A1",
            "Inside",
        ),
    ],
)
def test_hyperlink_assignment_and_reload_openpyxl_parity(
    tmp_path: Path,
    coord: str,
    initial: str | None,
    assigned: str | Hyperlink,
    expected_value: str,
    expected_display: str | None,
) -> None:
    out = tmp_path / f"{coord}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    if initial is not None:
        ws[coord] = initial
    ws[coord].hyperlink = assigned

    assert ws[coord].value == expected_value
    assert ws[coord].hyperlink is not None
    assert ws[coord].hyperlink.ref == coord
    assert ws[coord].hyperlink.display == expected_display

    wb.save(out)
    reloaded = load_workbook(out)
    cell = reloaded.active[coord]
    assert cell.value == expected_value
    assert cell.hyperlink is not None
    assert cell.hyperlink.ref == coord
    assert cell.hyperlink.display == expected_display


def test_hyperlink_removal_keeps_cell_value() -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "before"
    ws["A1"].hyperlink = "https://example.com"
    ws["A1"].hyperlink = None

    assert ws["A1"].value == "before"
    assert ws["A1"].hyperlink is None
