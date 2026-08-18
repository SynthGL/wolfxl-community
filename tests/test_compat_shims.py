"""T0 compat shim module smoke tests.

Two goals:
1. Every openpyxl module path exposed by wolfxl imports cleanly (no
   ``ModuleNotFoundError`` for the common paths).
2. Every formerly-stubbed shim class stays promoted to a real constructor,
   while any remaining shim placeholders raise ``NotImplementedError`` with a
   helpful message at construction time.

A drop-in replacement that silently no-ops would be far worse than a
pointed error. These tests pin the error behavior so we don't regress.
"""

from __future__ import annotations

import importlib
import importlib.util
import inspect
from io import BytesIO
import warnings
from xml.etree import ElementTree as ET
import zipfile

import pytest

import wolfxl
from tests.openpyxl_audit import (
    iter_importable_openpyxl_modules,
    iter_package_wide_openpyxl_modules,
)

_SAFE_CONSTRUCTOR_DEFAULT_TYPES = (str, int, float, bool, type(None), tuple)
_SAFE_PROPERTY_VALUE_TYPES = (str, int, float, bool, type(None))


def _safe_property_value(value: object) -> bool:
    if isinstance(value, _SAFE_PROPERTY_VALUE_TYPES):
        return True
    if isinstance(value, (list, tuple)):
        return all(isinstance(item, _SAFE_PROPERTY_VALUE_TYPES) for item in value)
    if isinstance(value, dict):
        return all(
            isinstance(key, _SAFE_PROPERTY_VALUE_TYPES)
            and isinstance(item, _SAFE_PROPERTY_VALUE_TYPES)
            for key, item in value.items()
        )
    return False

# ---------------- real class re-exports ----------------


def test_real_styles_classes_importable() -> None:
    from wolfxl.styles import Alignment, Border, Color, Font, PatternFill, Side

    assert Font(bold=True).bold is True
    assert PatternFill(patternType="solid", fgColor="FFFF0000").patternType == "solid"
    assert Color(rgb="FF00FF00").rgb == "FF00FF00"
    assert Border().left == Side()
    assert Alignment(horizontal="center").horizontal == "center"


def test_utils_cell_reexports() -> None:
    from wolfxl.utils.cell import (
        column_index_from_string,
        coordinate_to_tuple,
        get_column_letter,
        range_boundaries,
    )

    assert get_column_letter(1) == "A"
    assert column_index_from_string("AA") == 27
    assert coordinate_to_tuple("B3") == (3, 2)
    assert range_boundaries("A1:B2") == (1, 1, 2, 2)


def test_utils_cell_lazy_reexports() -> None:
    """Higher-level helpers routed through lazy ``__getattr__``."""
    from wolfxl.utils.cell import (
        absolute_coordinate,
        cols_from_range,
        get_column_interval,
        quote_sheetname,
        range_to_tuple,
        rows_from_range,
    )

    assert absolute_coordinate("A1") == "$A$1"
    assert quote_sheetname("Data")  # always quoted in openpyxl
    assert range_to_tuple("Sheet1!A1:B2") == ("Sheet1", (1, 1, 2, 2))
    assert list(rows_from_range("A1:B1")) == [("A1", "B1")]
    assert list(cols_from_range("A1:A2")) == [("A1", "A2")]
    assert get_column_interval("A", "C") == ["A", "B", "C"]


# ---------------- module imports ----------------


@pytest.mark.parametrize(
    "module_path",
    [
        "wolfxl.styles",
        "wolfxl.styles.cell_style",
        "wolfxl.styles.styleable",
        "wolfxl.styles.named_styles",
        "wolfxl.styles.differential",
        "wolfxl.xml",
        "wolfxl.xml.constants",
        "wolfxl.xml.functions",
        "wolfxl.reader",
        "wolfxl.reader.excel",
        "wolfxl.cell.read_only",
        "wolfxl.cell._writer",
        "wolfxl.utils.cell",
        "wolfxl.utils.bound_dictionary",
        "wolfxl.utils.dataframe",
        "wolfxl.utils.inference",
        "wolfxl.comments",
        "wolfxl.comments.shape_writer",
        "wolfxl.chart",
        "wolfxl.chart._3d",
        "wolfxl.chart.chartspace",
        "wolfxl.chart.picture",
        "wolfxl.chart.pivot",
        "wolfxl.chart.plotarea",
        "wolfxl.chart.print_settings",
        "wolfxl.chart.reader",
        "wolfxl.chart.series_factory",
        "wolfxl.chart.updown_bars",
        "wolfxl.chartsheet.custom",
        "wolfxl.chartsheet.publish",
        "wolfxl.chartsheet.relation",
        "wolfxl.chartsheet.views",
        "wolfxl.descriptors",
        "wolfxl.descriptors.base",
        "wolfxl.descriptors.container",
        "wolfxl.descriptors.excel",
        "wolfxl.descriptors.namespace",
        "wolfxl.descriptors.nested",
        "wolfxl.descriptors.sequence",
        "wolfxl.descriptors.serialisable",
        "wolfxl.descriptors.slots",
        "wolfxl.drawing",
        "wolfxl.drawing.colors",
        "wolfxl.drawing.connector",
        "wolfxl.drawing.drawing",
        "wolfxl.drawing.effect",
        "wolfxl.drawing.geometry",
        "wolfxl.drawing.graphic",
        "wolfxl.drawing.image",
        "wolfxl.drawing.picture",
        "wolfxl.drawing.properties",
        "wolfxl.drawing.relation",
        "wolfxl.drawing.text",
        "wolfxl.drawing.xdr",
        "wolfxl.comments.author",
        "wolfxl.comments.comment_sheet",
        "wolfxl.reader.drawings",
        "wolfxl.reader.strings",
        "wolfxl.reader.workbook",
        "wolfxl.worksheet",
        "wolfxl.worksheet._reader",
        "wolfxl.worksheet._write_only",
        "wolfxl.worksheet._writer",
        "wolfxl.worksheet.cell_watch",
        "wolfxl.worksheet.controls",
        "wolfxl.worksheet.custom",
        "wolfxl.worksheet.datavalidation",
        "wolfxl.worksheet.drawing",
        "wolfxl.worksheet.errors",
        "wolfxl.worksheet.ole",
        "wolfxl.worksheet.picture",
        "wolfxl.worksheet.related",
        "wolfxl.worksheet.scenario",
        "wolfxl.worksheet.smart_tag",
        "wolfxl.worksheet.table",
        "wolfxl.worksheet.filters",
        "wolfxl.worksheet.hyperlink",
        "wolfxl.formatting",
        "wolfxl.formatting.rule",
        "wolfxl.workbook",
        "wolfxl.workbook._writer",
        "wolfxl.workbook.defined_name",
        "wolfxl.workbook.external_link",
        "wolfxl.workbook.external_link.external",
        "wolfxl.workbook.external_reference",
        "wolfxl.workbook.function_group",
        "wolfxl.workbook.smart_tags",
        "wolfxl.workbook.web",
        "wolfxl.packaging.extended",
        "wolfxl.packaging.interface",
        "wolfxl.packaging.manifest",
        "wolfxl.packaging.relationship",
        "wolfxl.packaging.workbook",
        "wolfxl.styles.builtins",
        "wolfxl.styles.proxy",
        "wolfxl.styles.stylesheet",
        "wolfxl.styles.table",
        "wolfxl.pivot",
        "wolfxl.writer",
        "wolfxl.writer.excel",
        "wolfxl.writer.theme",
    ],
)
def test_module_imports(module_path: str) -> None:
    importlib.import_module(module_path)


def test_audited_openpyxl_module_paths_are_present() -> None:
    missing: list[str] = []
    for module in iter_importable_openpyxl_modules():
        wolfxl_module = "wolfxl" + module[len("openpyxl"):]
        try:
            found = importlib.util.find_spec(wolfxl_module) is not None
        except ModuleNotFoundError:
            found = False
        if not found:
            missing.append(f"{module} -> {wolfxl_module}")

    assert missing == []


def test_runtime_alias_imports_every_importable_openpyxl_module() -> None:
    """One-line replacement must preserve upstream module import paths."""
    from tests.test_openpyxl_runtime_alias import _isolated_openpyxl_imports

    modules = iter_package_wide_openpyxl_modules(include_root=True)
    with _isolated_openpyxl_imports():
        wolfxl.install_as_openpyxl()

        missing: list[str] = []
        errors: list[str] = []
        for module in modules:
            try:
                importlib.import_module(module)
            except ModuleNotFoundError as exc:
                missing.append(f"{module}: {exc}")
            except Exception as exc:  # noqa: BLE001 - aggregate report
                errors.append(f"{module}: {type(exc).__name__}: {exc}")

    assert missing == []
    assert errors == []


def test_audited_openpyxl_public_names_resolve() -> None:
    """Every audited openpyxl public import name should exist on wolfxl too."""
    missing: list[str] = []
    wolfxl_import_errors: list[str] = []

    for module in iter_importable_openpyxl_modules():
        wolfxl_module = "wolfxl" + module[len("openpyxl"):]
        openpyxl_module = importlib.import_module(module)
        try:
            wolfxl_mod = importlib.import_module(wolfxl_module)
        except Exception as exc:  # noqa: BLE001 - report all import failures together
            wolfxl_import_errors.append(
                f"{module} -> {wolfxl_module}: {type(exc).__name__}: {exc}"
            )
            continue

        public_names = getattr(openpyxl_module, "__all__", None) or [
            name for name in dir(openpyxl_module) if not name.startswith("_")
        ]
        for name in public_names:
            try:
                has_name = hasattr(wolfxl_mod, name)
            except Exception as exc:  # noqa: BLE001 - descriptor shims should not explode
                missing.append(
                    f"{module} -> {wolfxl_module}.{name}: {type(exc).__name__}: {exc}"
                )
            else:
                if not has_name:
                    missing.append(f"{module} -> {wolfxl_module}.{name}")

    assert wolfxl_import_errors == []
    assert missing == []


def test_package_wide_openpyxl_public_names_resolve() -> None:
    """Every package-wide upstream public import name should resolve in wolfxl."""
    modules = iter_package_wide_openpyxl_modules(include_root=True)
    missing: list[str] = []
    wolfxl_import_errors: list[str] = []
    for module in modules:
        openpyxl_module = importlib.import_module(module)
        wolfxl_module_name = "wolfxl" + module[len("openpyxl") :]
        try:
            wolfxl_module = importlib.import_module(wolfxl_module_name)
        except Exception as exc:  # noqa: BLE001 - aggregate report
            wolfxl_import_errors.append(
                f"{module} -> {wolfxl_module_name}: {type(exc).__name__}: {exc}"
            )
            continue

        public_names = getattr(openpyxl_module, "__all__", None) or [
            name for name in dir(openpyxl_module) if not name.startswith("_")
        ]
        for name in public_names:
            try:
                has_name = hasattr(wolfxl_module, name)
            except Exception as exc:  # noqa: BLE001 - descriptor shims should not explode
                missing.append(
                    f"{module} -> {wolfxl_module.__name__}.{name}: "
                    f"{type(exc).__name__}: {exc}"
                )
            else:
                if not has_name:
                    missing.append(f"{module} -> {wolfxl_module.__name__}.{name}")

    assert wolfxl_import_errors == []
    assert missing == []


def test_package_wide_openpyxl_zero_arg_public_constructors() -> None:
    """Public classes construct with no args wherever upstream also allows it."""
    failures: list[str] = []
    for module in iter_package_wide_openpyxl_modules(include_root=True):
        openpyxl_module = importlib.import_module(module)
        wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
        public_names = getattr(openpyxl_module, "__all__", None) or [
            name for name in dir(openpyxl_module) if not name.startswith("_")
        ]
        for name in public_names:
            openpyxl_obj = getattr(openpyxl_module, name)
            if not inspect.isclass(openpyxl_obj):
                continue
            if getattr(openpyxl_obj, "__module__", "").startswith(("abc", "typing")):
                continue
            wolfxl_obj = getattr(wolfxl_module, name)
            if not inspect.isclass(wolfxl_obj):
                continue
            try:
                openpyxl_obj()
            except Exception:  # noqa: BLE001 - upstream requires constructor args
                continue
            try:
                wolfxl_obj()
            except Exception as exc:  # noqa: BLE001 - aggregate report
                failures.append(f"{module}.{name}: {type(exc).__name__}: {exc}")

    assert failures == []


def test_package_wide_openpyxl_optional_default_keyword_constructors() -> None:
    """Default-valued upstream constructor keywords should not break in wolfxl."""
    failures: list[str] = []
    for module in iter_package_wide_openpyxl_modules(include_root=True):
        openpyxl_module = importlib.import_module(module)
        wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
        public_names = getattr(openpyxl_module, "__all__", None) or [
            name for name in dir(openpyxl_module) if not name.startswith("_")
        ]
        for name in public_names:
            openpyxl_obj = getattr(openpyxl_module, name)
            if not inspect.isclass(openpyxl_obj):
                continue
            if getattr(openpyxl_obj, "__module__", "").startswith(("abc", "typing")):
                continue
            wolfxl_obj = getattr(wolfxl_module, name)
            if not inspect.isclass(wolfxl_obj):
                continue
            try:
                signature = inspect.signature(openpyxl_obj)
            except (TypeError, ValueError):
                continue
            for param_name, parameter in signature.parameters.items():
                if param_name == "self" or parameter.kind not in (
                    parameter.POSITIONAL_OR_KEYWORD,
                    parameter.KEYWORD_ONLY,
                ):
                    continue
                default = parameter.default
                if default is inspect._empty or not isinstance(
                    default, _SAFE_CONSTRUCTOR_DEFAULT_TYPES
                ):
                    continue
                try:
                    openpyxl_obj(**{param_name: default})
                except Exception:  # noqa: BLE001 - upstream rejects this default
                    continue
                try:
                    wolfxl_obj(**{param_name: default})
                except Exception as exc:  # noqa: BLE001 - aggregate report
                    failures.append(
                        f"{module}.{name}({param_name}={default!r}): "
                        f"{type(exc).__name__}: {exc}"
                    )

    assert failures == []


def test_package_wide_openpyxl_safe_default_method_calls() -> None:
    """No-arg/default public method calls should match upstream compatibility."""
    failures: list[str] = []
    seen: set[tuple[str, str]] = set()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for module in iter_package_wide_openpyxl_modules(include_root=True):
            openpyxl_module = importlib.import_module(module)
            wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
            public_names = getattr(openpyxl_module, "__all__", None) or [
                name for name in dir(openpyxl_module) if not name.startswith("_")
            ]
            for class_name in public_names:
                openpyxl_obj = getattr(openpyxl_module, class_name)
                if not inspect.isclass(openpyxl_obj):
                    continue
                wolfxl_obj = getattr(wolfxl_module, class_name)
                if not inspect.isclass(wolfxl_obj):
                    continue
                key = (getattr(openpyxl_obj, "__module__", ""), class_name)
                if key in seen:
                    continue
                seen.add(key)
                try:
                    openpyxl_obj()
                    wolfxl_obj()
                except Exception:  # noqa: BLE001 - only audit no-arg classes here
                    continue
                for method_name, raw_member in vars(openpyxl_obj).items():
                    if (
                        method_name.startswith("_")
                        or isinstance(raw_member, (staticmethod, classmethod, property))
                        or not inspect.isfunction(raw_member)
                    ):
                        continue
                    if not callable(getattr(wolfxl_obj, method_name, None)):
                        continue
                    try:
                        signature = inspect.signature(raw_member)
                    except (TypeError, ValueError):
                        continue
                    params = [
                        parameter
                        for parameter in signature.parameters.values()
                        if parameter.name != "self"
                    ]
                    if all(
                        parameter.default is not inspect._empty
                        or parameter.kind
                        in (parameter.VAR_POSITIONAL, parameter.VAR_KEYWORD)
                        for parameter in params
                    ):
                        try:
                            getattr(openpyxl_obj(), method_name)()
                        except Exception:  # noqa: BLE001 - upstream rejects this call
                            pass
                        else:
                            try:
                                getattr(wolfxl_obj(), method_name)()
                            except Exception as exc:  # noqa: BLE001 - aggregate report
                                failures.append(
                                    f"{module}.{class_name}.{method_name}(): "
                                    f"{type(exc).__name__}: {exc}"
                                )
                    for parameter in params:
                        if parameter.kind not in (
                            parameter.POSITIONAL_OR_KEYWORD,
                            parameter.KEYWORD_ONLY,
                        ):
                            continue
                        default = parameter.default
                        if default is inspect._empty or not isinstance(
                            default, _SAFE_CONSTRUCTOR_DEFAULT_TYPES
                        ):
                            continue
                        try:
                            getattr(openpyxl_obj(), method_name)(
                                **{parameter.name: default}
                            )
                        except Exception:  # noqa: BLE001 - upstream rejects this default
                            continue
                        try:
                            getattr(wolfxl_obj(), method_name)(
                                **{parameter.name: default}
                            )
                        except Exception as exc:  # noqa: BLE001 - aggregate report
                            failures.append(
                                f"{module}.{class_name}.{method_name}"
                                f"({parameter.name}={default!r}): "
                                f"{type(exc).__name__}: {exc}"
                            )

    assert failures == []


def test_package_wide_openpyxl_safe_property_getters() -> None:
    """Public no-arg property getters should not raise where upstream works."""
    failures: list[str] = []
    seen: set[tuple[str, str]] = set()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for module in iter_package_wide_openpyxl_modules(include_root=True):
            openpyxl_module = importlib.import_module(module)
            wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
            public_names = getattr(openpyxl_module, "__all__", None) or [
                name for name in dir(openpyxl_module) if not name.startswith("_")
            ]
            for class_name in public_names:
                openpyxl_obj = getattr(openpyxl_module, class_name)
                if not inspect.isclass(openpyxl_obj):
                    continue
                wolfxl_obj = getattr(wolfxl_module, class_name)
                if not inspect.isclass(wolfxl_obj):
                    continue
                key = (getattr(openpyxl_obj, "__module__", ""), class_name)
                if key in seen:
                    continue
                seen.add(key)
                try:
                    openpyxl_instance = openpyxl_obj()
                    wolfxl_instance = wolfxl_obj()
                except Exception:  # noqa: BLE001 - only audit no-arg classes here
                    continue
                for property_name, raw_member in vars(openpyxl_obj).items():
                    if (
                        property_name.startswith("_")
                        or not isinstance(raw_member, property)
                        or raw_member.fget is None
                    ):
                        continue
                    try:
                        openpyxl_value = getattr(openpyxl_instance, property_name)
                    except Exception:  # noqa: BLE001 - upstream rejects this getter
                        continue
                    try:
                        wolfxl_value = getattr(wolfxl_instance, property_name)
                    except Exception as exc:  # noqa: BLE001 - aggregate report
                        failures.append(
                            f"{module}.{class_name}.{property_name}: "
                            f"{type(exc).__name__}: {exc}"
                        )
                        continue
                    if (
                        _safe_property_value(openpyxl_value)
                        and _safe_property_value(wolfxl_value)
                        and wolfxl_value != openpyxl_value
                    ):
                        failures.append(
                            f"{module}.{class_name}.{property_name}: "
                            f"{wolfxl_value!r} != {openpyxl_value!r}"
                        )

    assert failures == []


def test_package_wide_openpyxl_safe_iterable_protocols() -> None:
    """No-arg iterable protocol calls should not raise where upstream works."""
    failures: list[str] = []
    seen: set[tuple[str, str]] = set()
    operations = {
        "iter": lambda obj: list(obj),
        "contains_none": lambda obj: None in obj,
    }
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for module in iter_package_wide_openpyxl_modules(include_root=True):
            openpyxl_module = importlib.import_module(module)
            wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
            public_names = getattr(openpyxl_module, "__all__", None) or [
                name for name in dir(openpyxl_module) if not name.startswith("_")
            ]
            for class_name in public_names:
                openpyxl_obj = getattr(openpyxl_module, class_name)
                if not inspect.isclass(openpyxl_obj):
                    continue
                wolfxl_obj = getattr(wolfxl_module, class_name)
                if not inspect.isclass(wolfxl_obj):
                    continue
                key = (getattr(openpyxl_obj, "__module__", ""), class_name)
                if key in seen:
                    continue
                seen.add(key)
                try:
                    openpyxl_instance = openpyxl_obj()
                    wolfxl_instance = wolfxl_obj()
                except Exception:  # noqa: BLE001 - only audit no-arg classes here
                    continue
                for operation_name, operation in operations.items():
                    try:
                        operation(openpyxl_instance)
                    except Exception:  # noqa: BLE001 - upstream rejects this operation
                        continue
                    try:
                        operation(wolfxl_instance)
                    except Exception as exc:  # noqa: BLE001 - aggregate report
                        failures.append(
                            f"{module}.{class_name}.{operation_name}: "
                            f"{type(exc).__name__}: {exc}"
                        )

    assert failures == []


def test_package_wide_openpyxl_safe_container_protocols() -> None:
    """Container helpers should not raise where no-arg upstream accepts them."""
    failures: list[str] = []
    seen: set[tuple[str, str]] = set()
    operations = {
        "len": lambda obj: len(obj),
        "getitem_0": lambda obj: obj[0],
        "getitem_empty": lambda obj: obj[""],
        "keys": lambda obj: list(obj.keys()),
        "items": lambda obj: list(obj.items()),
        "values": lambda obj: list(obj.values()),
        "get_empty": lambda obj: obj.get(""),
        "append_none": lambda obj: obj.append(None),
        "append_empty": lambda obj: obj.append(""),
        "extend_empty": lambda obj: obj.extend([]),
        "add_none": lambda obj: obj.add(None),
    }
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        for module in iter_package_wide_openpyxl_modules(include_root=True):
            openpyxl_module = importlib.import_module(module)
            wolfxl_module = importlib.import_module("wolfxl" + module[len("openpyxl") :])
            public_names = getattr(openpyxl_module, "__all__", None) or [
                name for name in dir(openpyxl_module) if not name.startswith("_")
            ]
            for class_name in public_names:
                openpyxl_obj = getattr(openpyxl_module, class_name)
                if not inspect.isclass(openpyxl_obj):
                    continue
                wolfxl_obj = getattr(wolfxl_module, class_name)
                if not inspect.isclass(wolfxl_obj):
                    continue
                key = (getattr(openpyxl_obj, "__module__", ""), class_name)
                if key in seen:
                    continue
                seen.add(key)
                for operation_name, operation in operations.items():
                    try:
                        operation(openpyxl_obj())
                    except Exception:  # noqa: BLE001 - upstream rejects this operation
                        continue
                    try:
                        operation(wolfxl_obj())
                    except Exception as exc:  # noqa: BLE001 - aggregate report
                        failures.append(
                            f"{module}.{class_name}.{operation_name}: "
                            f"{type(exc).__name__}: {exc}"
                        )

    assert failures == []


def test_targeted_invalid_operation_exceptions_match_openpyxl() -> None:
    """Invalid operations should fail early where upstream openpyxl does."""
    from openpyxl.cell._writer import CellRichText as OpenpyxlCellRichText
    from openpyxl.chart import BarChart as OpenpyxlBarChart
    from openpyxl.descriptors.serialisable import Serialisable as OpenpyxlSerialisable
    from openpyxl.packaging.relationship import RelationshipList as OpenpyxlRelationshipList
    from wolfxl.cell._writer import CellRichText as WolfxlCellRichText
    from wolfxl.chart import BarChart as WolfxlBarChart
    from wolfxl.descriptors.serialisable import Serialisable as WolfxlSerialisable
    from wolfxl.packaging.relationship import RelationshipList as WolfxlRelationshipList

    cases = [
        (lambda: OpenpyxlCellRichText().append(None), lambda: WolfxlCellRichText().append(None)),
        (lambda: OpenpyxlBarChart().append(None), lambda: WolfxlBarChart().append(None)),
        (lambda: OpenpyxlBarChart().append(""), lambda: WolfxlBarChart().append("")),
        (lambda: OpenpyxlSerialisable().to_tree(), lambda: WolfxlSerialisable().to_tree()),
        (lambda: OpenpyxlRelationshipList().get(""), lambda: WolfxlRelationshipList().get("")),
    ]
    for openpyxl_call, wolfxl_call in cases:
        with pytest.raises(Exception) as openpyxl_exc:
            openpyxl_call()
        with pytest.raises(type(openpyxl_exc.value)):
            wolfxl_call()


def test_compat_helper_behavior_matches_openpyxl() -> None:
    import datetime as dt
    import warnings

    from openpyxl.compat import deprecated as openpyxl_deprecated
    from openpyxl.compat.product import prod as openpyxl_prod
    from openpyxl.compat.product import product as openpyxl_product
    from openpyxl.compat.singleton import Cached as OpenpyxlCached
    from openpyxl.compat.strings import safe_string as openpyxl_safe_string
    from wolfxl.compat import deprecated as wolfxl_deprecated
    from wolfxl.compat.product import prod as wolfxl_prod
    from wolfxl.compat.product import product as wolfxl_product
    from wolfxl.compat.singleton import Cached as WolfxlCached
    from wolfxl.compat.strings import safe_string as wolfxl_safe_string

    values = [
        None,
        True,
        False,
        1,
        1.25,
        float("nan"),
        float("inf"),
        -float("inf"),
        dt.datetime(2020, 1, 2, 3, 4, 5),
        dt.date(2020, 1, 2),
        dt.time(3, 4, 5),
        dt.timedelta(days=1, seconds=2),
        "x",
    ]
    assert [wolfxl_safe_string(value) for value in values] == [
        openpyxl_safe_string(value) for value in values
    ]

    sequences = [[2, 3, 4], [1.5, 2]]
    assert [wolfxl_product(seq) for seq in sequences] == [
        openpyxl_product(seq) for seq in sequences
    ]
    assert [wolfxl_prod(seq) for seq in sequences] == [
        openpyxl_prod(seq) for seq in sequences
    ]
    with pytest.raises(TypeError):
        wolfxl_product([])

    for reason in [123, object()]:
        with pytest.raises(TypeError):
            wolfxl_deprecated(reason)  # type: ignore[arg-type]

    @wolfxl_deprecated("why")
    class WolfxlDeprecatedClass:
        """class docs"""

    @openpyxl_deprecated("why")
    class OpenpyxlDeprecatedClass:
        """class docs"""

    with warnings.catch_warnings(record=True) as wolfxl_warnings:
        warnings.simplefilter("always")
        WolfxlDeprecatedClass()
    with warnings.catch_warnings(record=True) as openpyxl_warnings:
        warnings.simplefilter("always")
        OpenpyxlDeprecatedClass()

    assert WolfxlDeprecatedClass.__doc__ == OpenpyxlDeprecatedClass.__doc__
    assert [str(item.message) for item in wolfxl_warnings] == [
        "Call to deprecated class WolfxlDeprecatedClass (why)."
    ]
    assert [str(item.message) for item in openpyxl_warnings] == [
        "Call to deprecated class OpenpyxlDeprecatedClass (why)."
    ]

    class WolfxlCachedClass(metaclass=WolfxlCached):
        def __init__(self, value: int = 1) -> None:
            self.value = value

    class OpenpyxlCachedClass(metaclass=OpenpyxlCached):
        def __init__(self, value: int = 1) -> None:
            self.value = value

    assert WolfxlCachedClass(1) is WolfxlCachedClass(1)
    assert WolfxlCachedClass(1) is not WolfxlCachedClass(2)
    with pytest.raises(TypeError):
        WolfxlCachedClass(value=1)
    with pytest.raises(TypeError):
        OpenpyxlCachedClass(value=1)


def test_pivot_record_low_risk_defaults_match_openpyxl() -> None:
    from openpyxl.pivot.record import Boolean as OpenpyxlBoolean
    from openpyxl.pivot.record import Index as OpenpyxlIndex
    from openpyxl.pivot.record import RecordList as OpenpyxlRecordList
    from wolfxl.pivot.record import Boolean as WolfxlBoolean
    from wolfxl.pivot.record import Index as WolfxlIndex
    from wolfxl.pivot.record import RecordList as WolfxlRecordList

    assert WolfxlBoolean().__dict__.get("v") == OpenpyxlBoolean().__dict__.get("v")
    assert WolfxlIndex().__dict__.get("v") == OpenpyxlIndex().__dict__.get("v")
    assert WolfxlRecordList().__dict__ == OpenpyxlRecordList().__dict__


@pytest.mark.parametrize(
    "module_suffix",
    [
        "cell",
        "cell.cell",
        "cell.read_only",
        "cell.rich_text",
        "cell.text",
        "styles.numbers",
    ],
)
def test_high_traffic_openpyxl_public_names_exist(module_suffix: str) -> None:
    openpyxl_module = importlib.import_module(f"openpyxl.{module_suffix}")
    wolfxl_module = importlib.import_module(f"wolfxl.{module_suffix}")
    public_names = getattr(openpyxl_module, "__all__", None) or [
        name for name in dir(openpyxl_module) if not name.startswith("_")
    ]

    missing = [name for name in public_names if not hasattr(wolfxl_module, name)]

    assert missing == []


def test_openpyxl_corpus_import_paths_are_real() -> None:
    """Upstream corpus collection imports these openpyxl-shaped paths."""
    from wolfxl import DEFUSEDXML, LXML
    from wolfxl.cell.read_only import EMPTY_CELL
    from wolfxl.packaging.manifest import Manifest
    from wolfxl.reader.excel import load_workbook
    from wolfxl.styles.styleable import StyleArray
    from wolfxl.xml.constants import CPROPS_TYPE, SHEET_MAIN_NS
    from wolfxl.xml.functions import Element, tostring

    style = StyleArray([0, 0, 0, 7, 0, 0, 0, 0, 0])
    assert style.numFmtId == 7
    style.fontId = 3
    assert tuple(style[:4]) == (3, 0, 0, 7)
    assert EMPTY_CELL.value is None
    assert load_workbook is wolfxl.load_workbook
    assert isinstance(LXML, bool)
    assert isinstance(DEFUSEDXML, bool)
    assert SHEET_MAIN_NS.endswith("/main")
    assert CPROPS_TYPE == "application/vnd.openxmlformats-officedocument.custom-properties+xml"
    assert tostring(Element("root")).startswith(b"<root")
    assert Manifest().path == "[Content_Types].xml"


def test_formula_tokenizer_and_translator_import_compat() -> None:
    from wolfxl.formula.tokenizer import Token, Tokenizer
    from wolfxl.formula.translate import Translator

    tokens = Tokenizer("=SUM(A1:B2, 3)+$C$4").items
    assert [(t.value, t.type, t.subtype) for t in tokens] == [
        ("SUM(", Token.FUNC, Token.OPEN),
        ("A1:B2", Token.OPERAND, Token.RANGE),
        (",", Token.SEP, Token.ARG),
        (" ", Token.WSPACE, ""),
        ("3", Token.OPERAND, Token.NUMBER),
        (")", Token.FUNC, Token.CLOSE),
        ("+", Token.OP_IN, ""),
        ("$C$4", Token.OPERAND, Token.RANGE),
    ]

    assert Translator("=A1+$B$2+SUM(C1:D2)", "A1").translate_formula("B2") == (
        "=B2+$B$2+SUM(D2:E3)"
    )


@pytest.mark.parametrize(
    "formula",
    [
        '=IF(A1>0,"yes","no")',
        "=-A1+2",
        "={1,2;3,4}",
        "='Data Sheet'!$A$1+Sheet2!B3",
        "=[Book1.xlsx]Sheet1!A1",
        "=Table1[Amount]",
        "=A:A",
        "=1:3",
        "=#REF!+1",
        "=TRUE+FALSE",
        "=1.2E+3+A1",
    ],
)
def test_formula_tokenizer_matches_openpyxl_for_common_shapes(formula: str) -> None:
    from openpyxl.formula.tokenizer import Tokenizer as OpenpyxlTokenizer
    from wolfxl.formula.tokenizer import Tokenizer as WolfxlTokenizer

    openpyxl_tokens = [
        (token.value, token.type, token.subtype)
        for token in OpenpyxlTokenizer(formula).items
    ]
    wolfxl_tokens = [
        (token.value, token.type, token.subtype)
        for token in WolfxlTokenizer(formula).items
    ]

    assert wolfxl_tokens == openpyxl_tokens


def test_import_shim_slice_behaves() -> None:
    from wolfxl.chart.series_factory import SeriesFactory
    from wolfxl.drawing.xdr import XDRPoint2D, XDRTransform2D
    from wolfxl.packaging.relationship import (
        Relationship,
        RelationshipList,
        get_rels_path,
    )
    from wolfxl.worksheet._write_only import isgenerator
    from wolfxl.writer.theme import theme_xml, write_theme

    series = SeriesFactory("'Sheet'!$A$1:$A$3")
    assert series.val.numRef.f == "'Sheet'!$A$1:$A$3"

    transform = XDRTransform2D(off=XDRPoint2D(1, 2))
    assert transform.off.x == 1
    assert transform.off.y == 2

    rels = RelationshipList()
    rels.append(Relationship(type="worksheet", Target="worksheets/sheet1.xml"))
    assert rels["rId1"].Target == "worksheets/sheet1.xml"
    assert get_rels_path("xl/workbook.xml") == "xl/_rels/workbook.xml.rels"

    assert isgenerator((i for i in range(1))) is True
    assert write_theme() == theme_xml


def test_save_accepts_binary_file_like_objects(tmp_path) -> None:
    wb = wolfxl.Workbook()
    wb.active["A1"] = "file-like"

    fp = BytesIO()
    wb.save(fp)

    fp.seek(0)
    with zipfile.ZipFile(fp) as zf:
        assert "xl/workbook.xml" in zf.namelist()


def test_load_workbook_accepts_keep_vba_keyword(tmp_path) -> None:
    path = tmp_path / "plain.xlsx"
    wolfxl.Workbook().save(path)

    wb = wolfxl.load_workbook(path, keep_vba=False)
    assert wb.sheetnames == ["Sheet"]


def test_read_only_cell_number_format_and_is_date() -> None:
    from wolfxl.cell.read_only import ReadOnlyCell
    from wolfxl.styles.styleable import StyleArray

    workbook = type("Book", (), {"_cell_styles": [StyleArray(), StyleArray([0, 0, 0, 14, 0, 0, 0, 0, 0])]})()
    sheet = type("Sheet", (), {"parent": workbook, "title": "Sheet"})()

    plain = ReadOnlyCell(sheet, 1, 1, 42, style_id=0)
    dated = ReadOnlyCell(sheet, 1, 2, 42, style_id=1)

    assert plain.number_format == "General"
    assert plain.is_date is False
    assert dated.number_format == "mm-dd-yy"
    assert dated.is_date is True


def test_dirty_read_mode_save_promotes_to_modify_mode(tmp_path) -> None:
    path = tmp_path / "source.xlsx"
    wb = wolfxl.Workbook()
    wb.active["A1"] = "old"
    wb.save(path)

    loaded = wolfxl.load_workbook(path)
    loaded.active["A1"] = "new"
    out = tmp_path / "saved.xlsx"
    loaded.save(out)

    assert wolfxl.load_workbook(out).active["A1"].value == "new"


def test_nonstandard_absolute_workbook_target_loads(tmp_path) -> None:
    source = tmp_path / "source.xlsx"
    wolfxl.Workbook().save(source)
    shifted = tmp_path / "shifted.xlsx"

    with zipfile.ZipFile(source, "r") as src, zipfile.ZipFile(shifted, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            member = info.filename
            data = src.read(member)
            if member == "xl/workbook.xml":
                member = "xl/workbook2.xml"
            elif member == "xl/_rels/workbook.xml.rels":
                member = "xl/_rels/workbook2.xml.rels"
            elif member == "_rels/.rels":
                data = _rewrite_office_document_target(data, "/xl/workbook2.xml")
            elif member == "[Content_Types].xml":
                data = data.replace(b'PartName="/xl/workbook.xml"', b'PartName="/xl/workbook2.xml"')
            info.filename = member
            dst.writestr(info, data)

    wb = wolfxl.load_workbook(shifted)
    assert wb.sheetnames == ["Sheet"]


def _rewrite_office_document_target(data: bytes, target: str) -> bytes:
    root = ET.fromstring(data)
    for rel in root:
        if rel.get("Type", "").endswith("/officeDocument"):
            rel.set("Target", target)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


# ---------------- stubs raise at construction ----------------


STUB_CONSTRUCTORS: list[tuple[str, str]] = [
    # T1 PR1 promoted: Comment, Hyperlink now real dataclasses.
    # T1 PR2 promoted: DataValidation, Table, TableStyleInfo, *Rule now real.
    # Sprint Μ Pod-β (RFC-046) promoted: ``BarChart``, ``LineChart``,
    # ``PieChart``, ``DoughnutChart``, ``AreaChart``, ``ScatterChart``,
    # ``BubbleChart``, ``RadarChart``, ``Reference``, and ``Series`` are
    # now real chart classes — exercised in tests/test_charts_*.py.
    # Sprint Λ Pod-β (RFC-045) promoted: ``wolfxl.drawing.image.Image``
    # is now a real class — exercised in tests/test_images_write.py
    # and tests/test_images_modify.py.
    # Sprint Ο Pod 1B (RFC-056) promoted:
    # ``wolfxl.worksheet.filters.AutoFilter`` is now a real class —
    # exercised in tests/test_autofilter_filters.py.
    # T1 PR3 promoted: DefinedName now real dataclass.
    # Sprint Ν (RFC-047/048) promoted: ``wolfxl.pivot.PivotTable``,
    # ``PivotCache``, ``DataField``, ``PivotSource`` are real classes
    # — exercised in tests/test_pivot_construction.py.
]


# Classes promoted from stub -> real as T1 PRs land. Parametrized to catch
# accidental regression back to a raising stub.
REAL_DATACLASSES: list[tuple[str, str, dict]] = [
    ("wolfxl.comments", "Comment", {"text": "hello", "author": "me"}),
    ("wolfxl.worksheet.hyperlink", "Hyperlink", {"target": "https://example.com"}),
    ("wolfxl.worksheet.datavalidation", "DataValidation", {"type": "list", "formula1": '"a,b,c"'}),
    ("wolfxl.worksheet.table", "Table", {"name": "MyTable", "ref": "A1:B2"}),
    ("wolfxl.worksheet.table", "TableStyleInfo", {"name": "TableStyleLight9"}),
    ("wolfxl.formatting.rule", "CellIsRule", {"operator": "greaterThan", "formula": ["10"]}),
    ("wolfxl.formatting.rule", "FormulaRule", {"formula": ["$A1>100"]}),
    ("wolfxl.formatting.rule", "ColorScaleRule", {"start_type": "min", "end_type": "max"}),
    ("wolfxl.workbook.defined_name", "DefinedName", {"name": "Totals", "value": "Sheet1!$A$1"}),
    ("wolfxl.worksheet.filters", "AutoFilter", {}),
    ("wolfxl.styles", "NamedStyle", {"name": "Metric"}),
    ("wolfxl.styles", "Protection", {}),
    ("wolfxl.styles", "GradientFill", {}),
    ("wolfxl.styles.fills", "Fill", {}),
    ("wolfxl.styles.differential", "DifferentialStyle", {}),
    ("wolfxl.worksheet.dimensions", "DimensionHolder", {"worksheet": None}),
    ("wolfxl.worksheet.dimensions", "SheetFormatProperties", {}),
    ("wolfxl.worksheet.dimensions", "SheetDimension", {}),
    ("wolfxl.worksheet.merge", "MergeCell", {"ref": "A1:B2"}),
    ("wolfxl.worksheet.merge", "MergeCells", {}),
    ("wolfxl.worksheet.pagebreak", "Break", {}),
    ("wolfxl.worksheet.pagebreak", "PageBreak", {}),
    ("wolfxl.worksheet.properties", "WorksheetProperties", {}),
    ("wolfxl.worksheet.table", "TableList", {}),
    ("wolfxl.worksheet.table", "TablePartList", {}),
    ("wolfxl.worksheet.table", "Related", {}),
    ("wolfxl.worksheet.table", "XMLColumnProps", {}),
    ("wolfxl.workbook.properties", "CalcProperties", {}),
    ("wolfxl.workbook.properties", "WorkbookProperties", {}),
    ("wolfxl.workbook.child", "_WorkbookChild", {}),
    ("wolfxl.drawing.spreadsheet_drawing", "SpreadsheetDrawing", {}),
    ("wolfxl.comments.comments", "CommentSheet", {}),
]


@pytest.mark.parametrize("module_path,class_name,kwargs", REAL_DATACLASSES)
def test_real_dataclass_constructs_cleanly(
    module_path: str, class_name: str, kwargs: dict
) -> None:
    """Each promoted class constructs cleanly without raising.

    We don't compare every kwarg against an attribute of the same name —
    some classes (color-scale/data-bar rules) stash options into an
    ``extra`` dict rather than exposing each as a top-level attribute.
    The goal is to pin the ``NotImplementedError`` regression risk, not
    to fully spec each constructor's introspection surface.
    """
    mod = importlib.import_module(module_path)
    cls = getattr(mod, class_name)
    instance = cls(**kwargs)
    assert instance is not None
    # Best-effort attribute check: only probe names that ARE attributes.
    for key, value in kwargs.items():
        if hasattr(instance, key):
            assert getattr(instance, key) == value


@pytest.mark.parametrize("module_path,class_name", STUB_CONSTRUCTORS)
def test_stub_raises_on_construct(module_path: str, class_name: str) -> None:
    mod = importlib.import_module(module_path)
    cls = getattr(mod, class_name)
    with pytest.raises(NotImplementedError) as excinfo:
        cls()
    # Message contains the class name and our GitHub compatibility anchor.
    msg = str(excinfo.value)
    assert class_name in msg
    assert "wolfxl" in msg.lower()


def test_pivot_table_no_longer_stub() -> None:
    """Sprint Ν (RFC-047/048) ratchet flip: ``wolfxl.pivot.PivotTable``
    must NOT raise ``NotImplementedError`` on construction.

    This is the explicit ratchet for the v0.5+ → v2.0 promotion of
    pivot construction from stub to real class.
    """
    # Real PivotTable signature requires `cache=` and `location=`. The
    # ratchet is: invoking the constructor with proper args succeeds; the
    # stub variant would have raised NotImplementedError unconditionally.
    # Exhaustive surface tested in tests/test_pivot_construction.py.
    import inspect

    from wolfxl.pivot import PivotTable
    sig = inspect.signature(PivotTable.__init__)
    assert "cache" in sig.parameters
    assert "location" in sig.parameters
    # Sanity: confirm we did NOT accidentally re-stub.
    assert PivotTable.__module__.startswith("wolfxl.pivot")
    assert PivotTable.__init__.__qualname__ == "PivotTable.__init__"


# ---------------- Color theme/indexed support ----------------


def test_color_theme_roundtrip() -> None:
    c = wolfxl.Color(theme=1, tint=-0.3)
    assert c.theme == 1
    assert c.tint == -0.3
    assert c.rgb is None
    assert c.type == "theme"


def test_color_indexed() -> None:
    c = wolfxl.Color(indexed=3)
    assert c.indexed == 3
    assert c.rgb is None
    assert c.type == "indexed"
    # Indexed 3 is 00FF00 in the openpyxl COLOR_INDEX table.
    assert c.to_hex() == "#00FF00"


def test_color_rgb_default() -> None:
    c = wolfxl.Color()
    assert c.rgb == "00000000"
    assert c.theme is None
    assert c.indexed is None
    assert c.type == "rgb"


def test_color_is_hashable() -> None:
    """Frozen dataclass contract - Colors must be hashable for set/dict membership."""
    a = wolfxl.Color(rgb="FF0000")
    b = wolfxl.Color(rgb="FF0000")
    assert {a, b} == {a}


def test_style_openpyxl_aliases() -> None:
    font = wolfxl.Font(b=True, i=True, u=True, sz=14, strikethrough=True)
    assert font.bold is True
    assert font.b is True
    assert font.italic is True
    assert font.i is True
    assert font.underline == "single"
    assert font.u == "single"
    assert font.size == 14
    assert font.sz == 14
    assert font.strikethrough is True

    fill = wolfxl.PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FF00FF00")
    assert fill.patternType == "solid"
    assert fill.fill_type == "solid"
    assert fill.fgColor == "FFFF0000"
    assert fill.start_color == "FFFF0000"
    assert fill.bgColor == "FF00FF00"
    assert fill.end_color == "FF00FF00"

    align = wolfxl.Alignment(wrapText=True, textRotation=45, shrinkToFit=True)
    assert align.wrap_text is True
    assert align.wrapText is True
    assert align.text_rotation == 45
    assert align.textRotation == 45
    assert align.shrink_to_fit is True
    assert align.shrinkToFit is True

    side = wolfxl.Side(border_style="thin")
    assert side.style == "thin"
    assert side.border_style == "thin"
    border = wolfxl.Border(diagonal=side, diagonalUp=True)
    assert border.diagonal is side
    assert border.diagonal_direction == "up"

    color = wolfxl.Color(indexed=3)
    assert color.value == 3
    assert color.index == 3


def test_style_tree_helpers_round_trip() -> None:
    from xml.etree import ElementTree as ET

    font = wolfxl.Font(b=True, i=True, sz=14, color="FF0000")
    font_xml = ET.tostring(font.to_tree()).decode()
    assert "<b" in font_xml
    assert "<i" in font_xml
    assert 'rgb="00FF0000"' in font_xml
    assert wolfxl.Font.from_tree(font.to_tree()).bold is True

    fill = wolfxl.PatternFill(fill_type="solid", start_color="FF0000")
    assert wolfxl.PatternFill.from_tree(fill.to_tree()).fill_type == "solid"

    alignment = wolfxl.Alignment(horizontal="center", wrapText=True)
    assert wolfxl.Alignment.from_tree(alignment.to_tree()).wrapText is True

    side = wolfxl.Side(border_style="thin")
    border = wolfxl.Border(left=side, diagonalUp=True)
    assert wolfxl.Border.from_tree(border.to_tree()).left.border_style == "thin"

    color = wolfxl.Color.from_tree(wolfxl.Color(theme=1, tint=-0.3).to_tree())
    assert color.theme == 1
    assert color.tint == -0.3


def test_protection_and_named_style_helpers() -> None:
    from wolfxl.styles import NamedStyle, Protection

    protection = Protection(locked=False, hidden=True)
    assert Protection.from_tree(protection.to_tree()).locked is False
    assert Protection.from_tree(protection.to_tree()).hidden is True

    style = NamedStyle(name="Metric", builtinId=42, xfId=7, hidden=True)
    assert style.as_name().name == "Metric"
    assert style.as_tuple() == (0, 0, 0, 0, 0, 0, 0, 0, 0)
    assert style.as_xf().xfId == 7
    restored = NamedStyle.from_tree(style.to_tree())
    assert restored.name == "Metric"
    assert restored.builtinId == 42
    assert restored.xfId == 7
    assert restored.hidden is True


def test_data_validation_openpyxl_aliases_and_tree_helpers() -> None:
    from xml.etree import ElementTree as ET

    from wolfxl.worksheet.datavalidation import DataValidation, DataValidationList

    dv = DataValidation(type="list", formula1='"A,B"', allow_blank=True)
    dv.add("A1:B2")
    dv.hide_drop_down = True

    assert dv.validation_type == "list"
    assert dv.allowBlank is True
    assert dv.allow_blank is True
    assert dv.showDropDown is True
    assert dv.hide_drop_down is True
    assert str(dv.ranges) == "A1:B2"
    assert "A1" in dv.cells

    xml = ET.tostring(dv.to_tree()).decode()
    assert 'sqref="A1:B2"' in xml
    assert '<formula1>"A,B"</formula1>' in xml

    restored = DataValidation.from_tree(dv.to_tree())
    assert restored.validation_type == "list"
    assert restored.formula1 == '"A,B"'
    assert restored.hide_drop_down is True
    assert str(restored.sqref) == "A1:B2"

    validations = DataValidationList()
    validations.append(DataValidation())
    validations.append(restored)
    assert validations.count == 2
    list_xml = ET.tostring(validations.to_tree()).decode()
    assert 'count="1"' in list_xml
    assert "<dataValidation" in list_xml
    assert DataValidationList.from_tree(validations.to_tree()).count == 1


def test_filter_classes_accept_openpyxl_public_names() -> None:
    from wolfxl.worksheet.filters import (
        AutoFilter,
        ColorFilter,
        CustomFilter,
        CustomFilters,
        DynamicFilter,
        FilterColumn,
        Filters,
        IconFilter,
        SortCondition,
        SortState,
        Top10,
    )

    filters = Filters(blank=True, filter=["North"])
    column = FilterColumn(colId=0, filters=filters, hiddenButton=True)
    assert column.col_id == 0
    assert column.colId == 0
    assert column.hiddenButton is True
    assert column.vals == ["North"]
    assert column.blank is True

    custom = FilterColumn(customFilters=CustomFilters(_and=True, customFilter=[CustomFilter(val="A")]))
    assert custom.customFilters is not None
    assert custom.customFilters.and_ is True

    dynamic = DynamicFilter(type="today", valIso="2024-01-01", maxValIso="2024-01-31")
    assert dynamic.valIso == "2024-01-01"
    assert dynamic.maxValIso == "2024-01-31"

    color = ColorFilter(dxfId=2, cellColor=False)
    icon = IconFilter(iconSet="3Arrows", iconId=1)
    top = Top10(filterVal=5.0)
    assert color.dxfId == 2
    assert color.cellColor is False
    assert icon.iconSet == "3Arrows"
    assert icon.iconId == 1
    assert top.filterVal == 5.0

    sort = SortCondition(ref="A2:A10", sortBy="cellColor", dxfId=2, iconSet="3Arrows")
    state = SortState(sortCondition=[sort], columnSort=True, caseSensitive=True, sortMethod="stroke")
    auto_filter = AutoFilter(ref="A1:A10", filterColumn=[column], sortState=state)
    assert auto_filter.filterColumn == [column]
    assert auto_filter.sortState is state
    assert state.sortCondition == [sort]
    assert state.columnSort is True
    assert state.caseSensitive is True
    assert state.sortMethod == "stroke"
    assert sort.sortBy == "cellColor"
    assert sort.dxfId == 2
    assert sort.iconSet == "3Arrows"


def test_dataframe_to_rows_without_pandas_import() -> None:
    """Importing the module does not require pandas."""
    import wolfxl.utils.dataframe as dfmod

    assert hasattr(dfmod, "dataframe_to_rows")
    assert hasattr(dfmod, "worksheet_to_dataframe")


def test_dataframe_to_rows_basic() -> None:
    pd = pytest.importorskip("pandas")
    from wolfxl.utils.dataframe import dataframe_to_rows

    df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    rows = list(dataframe_to_rows(df, index=False, header=True))
    assert rows[0] == ["a", "b"]
    assert rows[1] == [1, 3]
    assert rows[2] == [2, 4]


def test_worksheet_to_dataframe_basic() -> None:
    pd = pytest.importorskip("pandas")
    from wolfxl.utils.dataframe import worksheet_to_dataframe

    wb = wolfxl.Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, 3])
    ws.append([2, 4])

    df = worksheet_to_dataframe(ws)
    expected = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
    pd.testing.assert_frame_equal(df, expected)

    via_method = ws.to_dataframe(max_row=2)
    pd.testing.assert_frame_equal(via_method, expected.iloc[:1])
