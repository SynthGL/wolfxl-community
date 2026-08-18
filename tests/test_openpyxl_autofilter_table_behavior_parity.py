from __future__ import annotations

from xml.etree.ElementTree import tostring

import openpyxl
from openpyxl.worksheet.table import Table as OpenpyxlTable
import pytest

import wolfxl
from wolfxl.worksheet.table import Table as WolfXLTable


def _tree_xml(obj: object) -> str:
    return tostring(obj.to_tree()).decode("us-ascii")  # type: ignore[attr-defined]


def test_autofilter_builder_shape_matches_openpyxl() -> None:
    openpyxl_ws = openpyxl.Workbook().active
    wolfxl_ws = wolfxl.Workbook().active

    for ws in (openpyxl_ws, wolfxl_ws):
        ws.auto_filter.ref = "A1:D10"

    openpyxl_filter_return = openpyxl_ws.auto_filter.add_filter_column(
        0, ["Closed", "Open"], blank=True
    )
    wolfxl_filter_return = wolfxl_ws.auto_filter.add_filter_column(
        0, ["Closed", "Open"], blank=True
    )
    openpyxl_sort_return = openpyxl_ws.auto_filter.add_sort_condition(
        "A2:A10", descending=True
    )
    wolfxl_sort_return = wolfxl_ws.auto_filter.add_sort_condition(
        "A2:A10", descending=True
    )

    assert wolfxl_filter_return is openpyxl_filter_return is None
    assert wolfxl_sort_return is openpyxl_sort_return is None
    assert list(wolfxl_ws.auto_filter) == list(openpyxl_ws.auto_filter)
    assert list(wolfxl_ws.auto_filter.filterColumn[0]) == list(
        openpyxl_ws.auto_filter.filterColumn[0]
    )
    assert list(wolfxl_ws.auto_filter.sortState) == list(
        openpyxl_ws.auto_filter.sortState
    )
    assert list(wolfxl_ws.auto_filter.sortState.sortCondition[0]) == list(
        openpyxl_ws.auto_filter.sortState.sortCondition[0]
    )
    assert _tree_xml(wolfxl_ws.auto_filter) == _tree_xml(openpyxl_ws.auto_filter)


def test_add_table_rejects_duplicate_names_like_openpyxl() -> None:
    for workbook_cls, table_cls in (
        (openpyxl.Workbook, OpenpyxlTable),
        (wolfxl.Workbook, WolfXLTable),
    ):
        wb = workbook_cls()
        ws = wb.active
        ws.add_table(table_cls(displayName="SalesTable", ref="A1:B2"))

        with pytest.raises(ValueError, match="SalesTable"):
            ws.add_table(table_cls(displayName="SalesTable", ref="D1:E2"))


def test_add_table_rejects_cross_sheet_duplicate_names_like_openpyxl() -> None:
    for workbook_cls, table_cls in (
        (openpyxl.Workbook, OpenpyxlTable),
        (wolfxl.Workbook, WolfXLTable),
    ):
        wb = workbook_cls()
        wb.active.add_table(table_cls(displayName="SalesTable", ref="A1:B2"))
        ws2 = wb.create_sheet("Sheet2")

        with pytest.raises(ValueError, match="SalesTable"):
            ws2.add_table(table_cls(displayName="SalesTable", ref="A1:B2"))
