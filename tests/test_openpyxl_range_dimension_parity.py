from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest

import wolfxl


def _close(workbook: Any) -> None:
    close = getattr(workbook, "close", None)
    if close is not None:
        close()


def _raises(fn: Any) -> tuple[str, str]:
    with pytest.raises(Exception) as exc_info:
        fn()
    exc = exc_info.value
    return type(exc).__name__, str(exc)


def _cell_range_snapshot(CellRange: Any) -> dict[str, Any]:
    titled = CellRange("'My Sheet'!B2:C3")
    shifted = CellRange("B2:C3")
    shifted.shift(col_shift=2, row_shift=3)
    expanded = CellRange("B2:C3")
    expanded.expand(right=1, down=2, left=1, up=1)
    shrunk = CellRange("A1:D4")
    shrunk.shrink(right=1, bottom=1, left=1, top=1)
    return {
        "single": CellRange("B2").bounds,
        "explicit": CellRange(
            min_col=2,
            min_row=3,
            max_col=4,
            max_row=5,
        ).coord,
        "title": (titled.title, titled.coord, titled.bounds, str(titled), repr(titled)),
        "shift": shifted.coord,
        "expand": expanded.coord,
        "shrink": shrunk.coord,
        "empty_error": _raises(lambda: CellRange()),
        "bad_shift_error_type": _raises(
            lambda: CellRange("B2:C3").shift(col_shift=-2)
        )[0],
    }


def test_cell_range_construction_and_mutation_match_openpyxl() -> None:
    from openpyxl.worksheet.cell_range import CellRange as OpenpyxlCellRange
    from wolfxl.worksheet.cell_range import CellRange as WolfxlCellRange

    assert _cell_range_snapshot(WolfxlCellRange) == _cell_range_snapshot(
        OpenpyxlCellRange
    )


def _multi_cell_range_snapshot(MultiCellRange: Any, CellRange: Any) -> dict[str, Any]:
    multi = MultiCellRange("B2:C3 A1")
    duplicate = MultiCellRange()
    duplicate.add("A1:B2")
    duplicate.add(CellRange("A1:B2"))
    return {
        "ranges_type": type(multi.ranges).__name__,
        "str": str(multi),
        "repr": repr(multi),
        "iter": [str(rng) for rng in multi],
        "ranges": sorted(str(rng) for rng in multi.ranges),
        "duplicate": sorted(str(rng) for rng in duplicate.ranges),
        "contains_cell": "A1" in MultiCellRange("A1:B2"),
        "contains_range": "A1:B2" in MultiCellRange("A1:B2"),
        "contains_partial": "B2:C3" in MultiCellRange("A1:B2"),
        "contains_cell_range": CellRange("A1") in MultiCellRange("A1:B2"),
        "contains_bad_type": _raises(lambda: (1, 1) in MultiCellRange("A1:B2")),
        "remove_missing": _raises(lambda: MultiCellRange("A1").remove("C1")),
        "remove_bad_type": _raises(lambda: MultiCellRange().remove(None)),
        "add_bad_type": _raises(lambda: MultiCellRange().add(None)),
        "order_insensitive_eq": MultiCellRange("A1 B2") == MultiCellRange("B2 A1"),
    }


def test_multi_cell_range_add_remove_contains_match_openpyxl() -> None:
    from openpyxl.worksheet.cell_range import CellRange as OpenpyxlCellRange
    from openpyxl.worksheet.cell_range import MultiCellRange as OpenpyxlMultiCellRange
    from wolfxl.worksheet.cell_range import CellRange as WolfxlCellRange
    from wolfxl.worksheet.cell_range import MultiCellRange as WolfxlMultiCellRange

    assert _multi_cell_range_snapshot(
        WolfxlMultiCellRange,
        WolfxlCellRange,
    ) == _multi_cell_range_snapshot(OpenpyxlMultiCellRange, OpenpyxlCellRange)


def _merged_cells_snapshot(xl: Any, path: Path) -> dict[str, Any]:
    wb = xl.Workbook()
    ws = wb.active
    fresh = {
        "type": type(ws.merged_cells).__name__,
        "ranges_type": type(ws.merged_cells.ranges).__name__,
        "str": str(ws.merged_cells),
        "contains": "A1:B2" in ws.merged_cells,
    }
    ws.merged_cells.add("A1:B2")
    after_add = {
        "str": str(ws.merged_cells),
        "contains_range": "A1:B2" in ws.merged_cells,
        "contains_anchor": "A1" in ws.merged_cells,
        "contains_corner": "B2" in ws.merged_cells,
        "ranges": sorted(str(rng) for rng in ws.merged_cells.ranges),
    }
    ws.merged_cells.remove("A1:B2")
    after_remove = {
        "str": str(ws.merged_cells),
        "ranges": sorted(str(rng) for rng in ws.merged_cells.ranges),
    }
    ws.merge_cells("C1:D2")
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    reloaded_snapshot = {
        "str": str(ws2.merged_cells),
        "contains_range": "C1:D2" in ws2.merged_cells,
        "contains_anchor": "C1" in ws2.merged_cells,
        "contains_corner": "D2" in ws2.merged_cells,
        "ranges": sorted(str(rng) for rng in ws2.merged_cells.ranges),
        "corner_type": type(ws2["D2"]).__name__,
    }
    _close(reloaded)
    return {
        "fresh": fresh,
        "after_add": after_add,
        "after_remove": after_remove,
        "reloaded": reloaded_snapshot,
    }


def test_worksheet_merged_cells_collection_matches_openpyxl(tmp_path: Path) -> None:
    assert _merged_cells_snapshot(
        wolfxl,
        tmp_path / "wolfxl-merged.xlsx",
    ) == _merged_cells_snapshot(openpyxl, tmp_path / "openpyxl-merged.xlsx")


def _dimension_rows(ws: Any) -> list[tuple[int, Any, bool, bool, int]]:
    return [
        (key, dim.height, dim.customHeight, dim.hidden, dim.outlineLevel)
        for key, dim in ws.row_dimensions.items()
    ]


def _dimension_cols(ws: Any) -> list[tuple[str, Any, bool, bool, int]]:
    return [
        (key, dim.width, dim.customWidth, dim.hidden, dim.outlineLevel)
        for key, dim in ws.column_dimensions.items()
    ]


def _dimensions_snapshot(xl: Any, path: Path) -> dict[str, Any]:
    wb = xl.Workbook()
    ws = wb.active
    fresh = {
        "rows": (len(ws.row_dimensions), list(ws.row_dimensions.keys())),
        "row_get": ws.row_dimensions.get(1, "missing"),
        "cols": (len(ws.column_dimensions), list(ws.column_dimensions.keys())),
        "col_get": ws.column_dimensions.get("A", "missing"),
    }
    row_one = ws.row_dimensions[1]
    col_a = ws.column_dimensions["A"]
    after_getitem = {
        "rows": (len(ws.row_dimensions), list(ws.row_dimensions.keys())),
        "row_items": _dimension_rows(ws),
        "row_one": (row_one.height, row_one.customHeight, row_one.hidden),
        "row_get_type": type(ws.row_dimensions.get(1, "missing")).__name__,
        "cols": (len(ws.column_dimensions), list(ws.column_dimensions.keys())),
        "col_items": _dimension_cols(ws),
        "col_a": (col_a.width, col_a.customWidth, col_a.hidden),
        "col_get_type": type(ws.column_dimensions.get("A", "missing")).__name__,
    }
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].outlineLevel = 2
    ws.column_dimensions["D"].width = 18
    before_save = {
        "row_keys": list(ws.row_dimensions.keys()),
        "row_items": _dimension_rows(ws),
        "col_keys": list(ws.column_dimensions.keys()),
        "col_items": _dimension_cols(ws),
    }
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    reloaded_snapshot = {
        "row_keys": list(ws2.row_dimensions.keys()),
        "row_items": _dimension_rows(ws2),
        "col_keys": list(ws2.column_dimensions.keys()),
        "col_items": _dimension_cols(ws2),
    }
    _close(reloaded)
    return {
        "fresh": fresh,
        "after_getitem": after_getitem,
        "before_save": before_save,
        "reloaded": reloaded_snapshot,
    }


def test_row_and_column_dimension_mapping_matches_openpyxl(tmp_path: Path) -> None:
    assert _dimensions_snapshot(
        wolfxl,
        tmp_path / "wolfxl-dimensions.xlsx",
    ) == _dimensions_snapshot(openpyxl, tmp_path / "openpyxl-dimensions.xlsx")
