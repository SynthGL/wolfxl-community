"""Runtime alias tests for one-line openpyxl replacement."""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
import importlib
import sys

import pytest

import wolfxl
from tests.openpyxl_audit import (
    iter_importable_openpyxl_modules,
    iter_package_wide_openpyxl_modules,
)


@contextmanager
def _isolated_openpyxl_imports() -> Iterator[None]:
    saved_modules = {
        name: module
        for name, module in sys.modules.items()
        if name == "openpyxl" or name.startswith("openpyxl.")
    }
    saved_meta_path = list(sys.meta_path)
    for name in list(saved_modules):
        del sys.modules[name]
    try:
        yield
    finally:
        for name in list(sys.modules):
            if name == "openpyxl" or name.startswith("openpyxl."):
                del sys.modules[name]
        sys.modules.update(saved_modules)
        sys.meta_path[:] = saved_meta_path


def test_install_as_openpyxl_supports_unchanged_imports(tmp_path) -> None:
    with _isolated_openpyxl_imports():
        wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.reader.excel import load_workbook
        from openpyxl.styles import Font, PatternFill
        import openpyxl.styles as styles
        from openpyxl.worksheet.table import Table, TableStyleInfo

        assert openpyxl is wolfxl
        assert Workbook is wolfxl.Workbook
        assert styles is importlib.import_module("wolfxl.styles")
        assert Font is importlib.import_module("wolfxl.styles").Font
        assert PatternFill is importlib.import_module("wolfxl.styles").PatternFill
        assert Table is importlib.import_module("wolfxl.worksheet.table").Table
        assert TableStyleInfo is importlib.import_module(
            "wolfxl.worksheet.table"
        ).TableStyleInfo
        assert BarChart is importlib.import_module("wolfxl.chart").BarChart
        assert Reference is importlib.import_module("wolfxl.chart").Reference
        assert load_workbook is wolfxl.load_workbook

        path = tmp_path / "aliased.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "amount"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = 42
        ws["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
        wb.save(path)

        reloaded = openpyxl.load_workbook(path)
        assert reloaded.sheetnames == ["Data"]
        assert reloaded["Data"]["A1"].value == "amount"
        assert reloaded["Data"]["A1"].font.bold is True
        assert reloaded["Data"]["A2"].value == 42


def test_install_as_openpyxl_aliases_audited_module_paths() -> None:
    with _isolated_openpyxl_imports():
        modules = iter_importable_openpyxl_modules(include_root=True)
        wolfxl.install_as_openpyxl()

        for module in modules:
            aliased = importlib.import_module(module)
            mapped = importlib.import_module("wolfxl" + module[len("openpyxl") :])
            assert aliased is mapped, module


def test_install_as_openpyxl_exposes_package_wide_public_names() -> None:
    """Alias mode must preserve package-wide public imports, not just modules."""
    public_names_by_module: dict[str, list[str]] = {}
    for module in iter_package_wide_openpyxl_modules(include_root=True):
        upstream_module = importlib.import_module(module)
        public_names = getattr(upstream_module, "__all__", None) or [
            name for name in dir(upstream_module) if not name.startswith("_")
        ]
        public_names_by_module[module] = list(public_names)

    with _isolated_openpyxl_imports():
        wolfxl.install_as_openpyxl()

        missing: list[str] = []
        errors: list[str] = []
        for module, public_names in public_names_by_module.items():
            aliased_module = importlib.import_module(module)
            for name in public_names:
                try:
                    has_name = hasattr(aliased_module, name)
                except Exception as exc:  # noqa: BLE001 - aggregate descriptor errors
                    errors.append(f"{module}.{name}: {type(exc).__name__}: {exc}")
                else:
                    if not has_name:
                        missing.append(f"{module}.{name}")

    assert errors == []
    assert missing == []


def test_install_as_openpyxl_replaces_preloaded_upstream_module() -> None:
    pytest.importorskip("openpyxl")

    with _isolated_openpyxl_imports():
        upstream = importlib.import_module("openpyxl")
        assert upstream is not wolfxl

        wolfxl.install_as_openpyxl()

        assert sys.modules["openpyxl"] is wolfxl
        assert importlib.import_module("openpyxl") is wolfxl
        assert importlib.import_module("openpyxl.styles") is importlib.import_module(
            "wolfxl.styles"
        )


def test_install_as_openpyxl_force_false_reports_existing_import() -> None:
    pytest.importorskip("openpyxl")

    with _isolated_openpyxl_imports():
        importlib.import_module("openpyxl")

        with pytest.raises(RuntimeError, match="openpyxl is already imported"):
            wolfxl.install_as_openpyxl(force=False)
