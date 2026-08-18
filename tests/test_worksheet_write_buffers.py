from __future__ import annotations

from datetime import date

import openpyxl

from wolfxl import Workbook
from wolfxl._worksheet_write_buffers import extract_non_batchable
from wolfxl._worksheet_writer_flush import _flush_bulk_writes
from wolfxl.styles import Font


def test_extract_non_batchable_keeps_bool_formula_and_error_on_batch_path() -> None:
    grid = [[1, True, "=A1*2", "#N/A", "plain"]]

    individual = extract_non_batchable(grid, start_row=1, start_col=1)

    assert individual == []
    assert grid == [[1, True, "=A1*2", "#N/A", "plain"]]


def test_extract_non_batchable_preserves_date_per_cell_format_path() -> None:
    value = date(2026, 5, 20)
    grid = [[1, value, "plain"]]

    individual = extract_non_batchable(grid, start_row=10, start_col=3)

    assert individual == [(10, 4, value)]
    assert grid == [[1, None, "plain"]]


def test_append_mixed_batch_values_round_trip(tmp_path) -> None:
    path = tmp_path / "mixed-batch.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["amount", "flag", "formula", "error"])
    ws.append([5, True, "=A2*2", "#N/A"])
    wb.save(str(path))
    wb.close()

    loaded = openpyxl.load_workbook(path, data_only=False)
    sheet = loaded.active
    assert sheet["A2"].value == 5
    assert sheet["B2"].value is True
    assert sheet["C2"].value == "=A2*2"
    assert sheet["D2"].value == "#N/A"
    loaded.close()


def test_write_styled_rows_flushes_style_id_grid() -> None:
    class Writer:
        def __init__(self) -> None:
            self.values: list[tuple[str, str, list[list[object]]]] = []
            self.interned: list[dict[str, object]] = []
            self.style_ids: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_values(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.values.append((sheet, start, grid))

        def intern_format(self, payload: dict[str, object]) -> int:
            self.interned.append(payload)
            return len(self.interned)

        def write_sheet_style_ids(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.style_ids.append((sheet, start, grid))

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.write_styled_rows(
        [[1, 2], [3, 4]],
        {1: {"font": Font(bold=True)}, 2: "#,##0.00"},
        plain_values_only=True,
    )
    writer = Writer()

    _flush_bulk_writes(ws, writer, lambda value: {"value": value})

    assert writer.values == [("Sheet", "A1", [[1, 2], [3, 4]])]
    assert writer.interned == [{"bold": True}, {"number_format": "#,##0.00"}]
    assert writer.style_ids == [("Sheet", "A1", [[1, 2], [1, 2]])]


def test_write_styled_rows_uses_combined_value_style_batch_when_available() -> None:
    class Writer:
        def __init__(self) -> None:
            self.values: list[tuple[str, str, list[list[object]]]] = []
            self.combined: list[tuple[str, str, list[list[object]], list[list[object]]]] = []
            self.interned: list[dict[str, object]] = []
            self.style_ids: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_values(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.values.append((sheet, start, grid))

        def intern_format(self, payload: dict[str, object]) -> int:
            self.interned.append(payload)
            return len(self.interned)

        def write_sheet_style_ids(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.style_ids.append((sheet, start, grid))

        def write_sheet_values_with_style_ids(
            self,
            sheet: str,
            start: str,
            values: list[list[object]],
            style_ids: list[list[object]],
        ) -> None:
            self.combined.append((sheet, start, values, style_ids))

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.write_styled_rows(
        [[1, 2], [3, None]],
        {1: {"font": Font(bold=True)}, 2: "#,##0.00"},
        plain_values_only=True,
    )
    writer = Writer()

    _flush_bulk_writes(ws, writer, lambda value: {"value": value})

    assert writer.values == []
    assert writer.interned == [{"bold": True}, {"number_format": "#,##0.00"}]
    assert writer.combined == [
        ("Sheet", "A1", [[1, 2], [3, None]], [[1, 2], [1, 2]])
    ]
    assert writer.style_ids == []


def test_write_styled_rows_accepts_normalized_style_grid() -> None:
    class Writer:
        def __init__(self) -> None:
            self.interned: list[dict[str, object]] = []
            self.combined: list[tuple[str, str, list[list[object]], list[list[object]]]] = []

        def intern_format(self, payload: dict[str, object]) -> int:
            self.interned.append(payload)
            return len(self.interned)

        def write_sheet_style_ids(
            self,
            _sheet: str,
            _start: str,
            _grid: list[list[object]],
        ) -> None:
            raise AssertionError("combined path should handle normalized styles")

        def write_sheet_values_with_style_ids(
            self,
            sheet: str,
            start: str,
            values: list[list[object]],
            style_ids: list[list[object]],
        ) -> None:
            self.combined.append((sheet, start, values, style_ids))

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.write_styled_rows(
        [[1, 2]],
        [[{"bold": True}, {"number_format": "#,##0.00"}]],
        plain_values_only=True,
        normalized_styles=True,
    )
    writer = Writer()

    _flush_bulk_writes(ws, writer, lambda value: {"value": value})

    assert writer.interned == [{"bold": True}, {"number_format": "#,##0.00"}]
    assert writer.combined == [("Sheet", "A1", [[1, 2]], [[1, 2]])]
