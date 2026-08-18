from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import wolfxl
from wolfxl._worksheet_iteration import _reader_has_array_formulas


def test_plain_bulk_reader_returns_row_tuples(tmp_path: Path) -> None:
    path = tmp_path / "values.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "amount"])
    ws.append(["alpha", 10])
    wb.save(str(path))
    wb.close()

    loaded = wolfxl.load_workbook(path)
    reader = loaded._rust_reader  # noqa: SLF001
    rows = reader.read_sheet_values_plain(loaded.sheetnames[0], "A1:B2")

    assert rows == [("name", "amount"), ("alpha", 10)]
    assert all(isinstance(row, tuple) for row in rows)
    loaded.close()


def test_plain_bulk_reader_unbounded_range_returns_used_grid(tmp_path: Path) -> None:
    path = tmp_path / "values.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["B2"] = "start"
    ws["D4"] = 42
    wb.save(str(path))
    wb.close()

    loaded = wolfxl.load_workbook(path)
    reader = loaded._rust_reader  # noqa: SLF001
    rows = reader.read_sheet_values_plain(loaded.sheetnames[0], None)

    assert rows == [
        (None, None, None, None),
        (None, "start", None, None),
        (None, None, None, None),
        (None, None, None, 42),
    ]
    assert all(isinstance(row, tuple) for row in rows)
    loaded.close()


def test_iter_rows_values_only_sparse_unbounded_matches_openpyxl_anchor(
    tmp_path: Path,
) -> None:
    path = tmp_path / "values.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["B2"] = "start"
    ws["D4"] = 42
    wb.save(str(path))
    wb.close()

    loaded = wolfxl.load_workbook(path)
    rows = list(loaded.active.iter_rows(values_only=True))

    assert rows == [
        (None, None, None, None),
        (None, "start", None, None),
        (None, None, None, None),
        (None, None, None, 42),
    ]
    loaded.close()


def test_iter_rows_values_only_uses_tuple_rows_from_bulk_reader(tmp_path: Path) -> None:
    path = tmp_path / "values.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["name", "amount"])
    ws.append(["alpha", 10])
    wb.save(str(path))
    wb.close()

    loaded = wolfxl.load_workbook(path)
    rows = list(loaded.active.iter_rows(values_only=True))

    assert rows == [("name", "amount"), ("alpha", 10)]
    assert all(isinstance(row, tuple) for row in rows)
    loaded.close()


def test_dense_plain_bulk_reader_preserves_date_formats(tmp_path: Path) -> None:
    path = tmp_path / "dense-date.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = date(2026, 5, 31)
    ws["B1"] = "sentinel"
    wb.save(path)
    wb.close()

    loaded = wolfxl.load_workbook(path, data_only=True)
    rows = list(loaded.active.iter_rows(values_only=True))

    assert rows == [(datetime(2026, 5, 31), "sentinel")]
    loaded.close()


def test_array_formula_presence_check_prefers_boolean_reader_method() -> None:
    class Reader:
        def __init__(self) -> None:
            self.boolean_calls = 0

        def has_sheet_array_formulas(self, sheet: str) -> bool:
            self.boolean_calls += 1
            assert sheet == "Sheet"
            return False

        def read_sheet_array_formulas(self, sheet: str) -> object:
            raise AssertionError(f"unexpected full array formula read for {sheet}")

    class Workbook:
        def __init__(self) -> None:
            self._rust_reader = Reader()

    class Worksheet:
        _title = "Sheet"

        def __init__(self) -> None:
            self._workbook = Workbook()
            self._reader_has_array_formulas_cache = None

    ws = Worksheet()

    assert _reader_has_array_formulas(ws) is False  # type: ignore[arg-type]
    assert ws._workbook._rust_reader.boolean_calls == 1
    assert _reader_has_array_formulas(ws) is False  # type: ignore[arg-type]
    assert ws._workbook._rust_reader.boolean_calls == 1
