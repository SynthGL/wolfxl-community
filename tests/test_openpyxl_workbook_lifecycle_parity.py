from __future__ import annotations

from io import BytesIO
from tempfile import NamedTemporaryFile
import zipfile

import openpyxl
import pytest

import wolfxl


def _two_sheet_workbook(xl):
    wb = xl.Workbook()
    wb.active.title = "First"
    wb.create_sheet("Second")
    return wb


def _active_title(wb) -> str | None:
    active = wb.active
    return None if active is None else active.title


@pytest.mark.parametrize(
    ("index", "expected_title"),
    [
        (1, "Second"),
        (-1, "Second"),
        (2, None),
        (99, None),
    ],
)
def test_active_int_setter_matches_openpyxl(index: int, expected_title: str | None) -> None:
    openpyxl_wb = _two_sheet_workbook(openpyxl)
    wolfxl_wb = _two_sheet_workbook(wolfxl)

    openpyxl_wb.active = index
    wolfxl_wb.active = index

    assert _active_title(openpyxl_wb) == expected_title
    assert _active_title(wolfxl_wb) == expected_title


def test_active_worksheet_setter_matches_openpyxl() -> None:
    openpyxl_wb = _two_sheet_workbook(openpyxl)
    wolfxl_wb = _two_sheet_workbook(wolfxl)

    openpyxl_wb.active = openpyxl_wb["Second"]
    wolfxl_wb.active = wolfxl_wb["Second"]

    assert _active_title(openpyxl_wb) == "Second"
    assert _active_title(wolfxl_wb) == "Second"


@pytest.mark.parametrize("index", [1, -1])
def test_active_sheet_selection_round_trips_through_save(index: int) -> None:
    for xl in (openpyxl, wolfxl):
        wb = _two_sheet_workbook(xl)
        wb.active = index

        stream = BytesIO()
        wb.save(stream)

        stream.seek(0)
        reloaded = xl.load_workbook(stream)
        assert _active_title(reloaded) == "Second"


@pytest.mark.parametrize("bad_value", ["1", None, object()])
def test_active_rejects_invalid_value_types_like_openpyxl(bad_value: object) -> None:
    for xl in (openpyxl, wolfxl):
        wb = _two_sheet_workbook(xl)
        with pytest.raises(TypeError):
            wb.active = bad_value


def test_active_rejects_foreign_worksheet_even_with_matching_title() -> None:
    for xl in (openpyxl, wolfxl):
        wb = xl.Workbook()
        foreign = xl.Workbook().active

        with pytest.raises(ValueError):
            wb.active = foreign


def test_active_rejects_hidden_worksheet_like_openpyxl() -> None:
    for xl in (openpyxl, wolfxl):
        wb = _two_sheet_workbook(xl)
        hidden = wb["Second"]
        hidden.sheet_state = "hidden"

        with pytest.raises(ValueError):
            wb.active = hidden


def test_save_to_bytesio_leaves_valid_zip_at_stream_end() -> None:
    for xl in (openpyxl, wolfxl):
        wb = xl.Workbook()
        wb.active["A1"] = "file-like"

        stream = BytesIO()
        wb.save(stream)

        assert stream.tell() == len(stream.getvalue())
        assert zipfile.is_zipfile(BytesIO(stream.getvalue()))

        stream.seek(0)
        reloaded = xl.load_workbook(stream)
        assert reloaded.active["A1"].value == "file-like"


def test_repeated_save_to_same_bytesio_preserves_latest_workbook() -> None:
    for xl in (openpyxl, wolfxl):
        wb = xl.Workbook()
        wb.active["A1"] = "first"

        stream = BytesIO()
        wb.save(stream)

        wb.active["A1"] = "second"
        wb.save(stream)

        assert stream.tell() == len(stream.getvalue())
        assert zipfile.is_zipfile(BytesIO(stream.getvalue()))

        stream.seek(0)
        reloaded = xl.load_workbook(stream)
        assert reloaded.active["A1"].value == "second"


def test_named_temporary_file_handle_save_and_load_lifecycle() -> None:
    for xl in (openpyxl, wolfxl):
        wb = xl.Workbook()
        wb.active["A1"] = "temp-handle"

        with NamedTemporaryFile(suffix=".xlsx") as handle:
            wb.save(handle)

            assert handle.closed is False
            assert handle.tell() > 0
            handle.seek(0)
            with zipfile.ZipFile(handle) as archive:
                assert archive.testzip() is None

            handle.seek(0)
            reloaded = xl.load_workbook(handle)
            assert reloaded.active["A1"].value == "temp-handle"
            assert handle.closed is False
