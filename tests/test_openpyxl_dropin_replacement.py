"""Program-level drop-in checks for openpyxl-shaped user code.

These tests run the same user-style snippet twice: once with openpyxl imports
and once with the matching wolfxl imports.  The observable workbook result must
match.  This is the north-star gate for "change the import line and keep going".
"""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest

import wolfxl
from wolfxl.cell.cell import Cell as WolfxlCell
from wolfxl.comments import Comment as WolfxlComment
from wolfxl.formatting.rule import CellIsRule as WolfxlCellIsRule
from wolfxl.styles import PatternFill as WolfxlPatternFill
import wolfxl.styles as wolfxl_styles
from wolfxl.workbook import defined_name as wolfxl_defined_name
from wolfxl.worksheet.worksheet import Worksheet as WolfxlWorksheet
from wolfxl.worksheet import datavalidation as wolfxl_datavalidation
from wolfxl.worksheet.table import Table as WolfxlTable
from wolfxl.worksheet.table import TableStyleInfo as WolfxlTableStyleInfo

openpyxl = pytest.importorskip("openpyxl")
from openpyxl.cell.cell import Cell as OpenpyxlCell  # noqa: E402
from openpyxl.comments import Comment as OpenpyxlComment  # noqa: E402
from openpyxl.formatting.rule import CellIsRule as OpenpyxlCellIsRule  # noqa: E402
from openpyxl.styles import PatternFill as OpenpyxlPatternFill  # noqa: E402
import openpyxl.styles as openpyxl_styles  # noqa: E402
from openpyxl.workbook import defined_name as openpyxl_defined_name  # noqa: E402
from openpyxl.worksheet import datavalidation as openpyxl_datavalidation  # noqa: E402
from openpyxl.worksheet.table import Table as OpenpyxlTable  # noqa: E402
from openpyxl.worksheet.table import TableStyleInfo as OpenpyxlTableStyleInfo  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet  # noqa: E402


@dataclass(frozen=True)
class WorkbookModules:
    xl: Any
    styles: Any
    datavalidation: Any
    defined_name: Any
    Comment: Any
    Table: Any
    TableStyleInfo: Any
    CellIsRule: Any
    PatternFill: Any
    Worksheet: Any
    Cell: Any


@dataclass(frozen=True)
class DropInProgram:
    label: str
    run: Callable[[WorkbookModules, Path], Any]


OPENPYXL_MODULES = WorkbookModules(
    xl=openpyxl,
    styles=openpyxl_styles,
    datavalidation=openpyxl_datavalidation,
    defined_name=openpyxl_defined_name,
    Comment=OpenpyxlComment,
    Table=OpenpyxlTable,
    TableStyleInfo=OpenpyxlTableStyleInfo,
    CellIsRule=OpenpyxlCellIsRule,
    PatternFill=OpenpyxlPatternFill,
    Worksheet=OpenpyxlWorksheet,
    Cell=OpenpyxlCell,
)
WOLFXL_MODULES = WorkbookModules(
    xl=wolfxl,
    styles=wolfxl_styles,
    datavalidation=wolfxl_datavalidation,
    defined_name=wolfxl_defined_name,
    Comment=WolfxlComment,
    Table=WolfxlTable,
    TableStyleInfo=WolfxlTableStyleInfo,
    CellIsRule=WolfxlCellIsRule,
    PatternFill=WolfxlPatternFill,
    Worksheet=WolfxlWorksheet,
    Cell=WolfxlCell,
)


def _close(workbook: Any) -> None:
    close = getattr(workbook, "close", None)
    if close is not None:
        close()


def _visible_rgb(color: Any) -> Any:
    rgb = getattr(color, "rgb", color)
    if isinstance(rgb, str):
        value = rgb.removeprefix("#").upper()
        if len(value) == 8:
            return value[-6:]
        return value
    return rgb


def _basic_cells_and_sheets(modules: WorkbookModules, path: Path) -> dict[str, Any]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "name"
    ws["B1"] = "amount"
    ws.append(["alpha", 10])
    ws.cell(row=3, column=1, value="beta")
    ws.cell(row=3, column=2, value=20)
    wb.create_sheet("Summary")
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path, data_only=False)
    data = reloaded["Data"]
    result = {
        "sheetnames": tuple(reloaded.sheetnames),
        "active_title": reloaded.active.title,
        "dimensions": (data.max_row, data.max_column),
        "rows": tuple(
            tuple(row)
            for row in data.iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=2,
                values_only=True,
            )
        ),
    }
    _close(reloaded)
    return result


def _append_loaded_workbook(modules: WorkbookModules, path: Path) -> tuple[Any, ...]:
    xl = modules.xl
    seed = xl.Workbook()
    seed.active.title = "Events"
    seed.active["A1"] = "event"
    seed.save(path)
    _close(seed)

    wb = xl.load_workbook(path)
    wb["Events"].append(["launched"])
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = tuple(
        tuple(row)
        for row in reloaded["Events"].iter_rows(
            min_row=1,
            max_row=2,
            min_col=1,
            max_col=1,
            values_only=True,
        )
    )
    _close(reloaded)
    return result


def _styles_round_trip(modules: WorkbookModules, path: Path) -> tuple[Any, ...]:
    xl = modules.xl
    styles = modules.styles
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Styled"
    ws["A1"] = "Revenue"
    ws["B2"] = 1234.5
    ws["A1"].font = styles.Font(bold=True, italic=True, name="Arial", size=14)
    ws["B2"].fill = styles.PatternFill(fill_type="solid", fgColor="FFFF0000")
    ws["B2"].alignment = styles.Alignment(horizontal="center")
    ws["B2"].number_format = "$#,##0.00"
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    label = reloaded["Styled"]["A1"]
    amount = reloaded["Styled"]["B2"]
    result = (
        label.value,
        label.font.bold,
        label.font.italic,
        label.font.name,
        float(label.font.size),
        amount.value,
        amount.alignment.horizontal,
        amount.number_format,
    )
    _close(reloaded)
    return result


def _data_validation_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    datavalidation = modules.datavalidation
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Input"
    ws["A1"] = "Choice"
    dv = datavalidation.DataValidation(
        type="list",
        formula1='"A,B,C"',
        allow_blank=True,
    )
    dv.add("A2:A5")
    ws.add_data_validation(dv)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    validations = list(reloaded["Input"].data_validations.dataValidation)
    result = (
        len(validations),
        validations[0].type,
        validations[0].formula1,
        str(validations[0].sqref),
        validations[0].allowBlank,
    )
    _close(reloaded)
    return result


def _defined_name_xml_round_trip(
    modules: WorkbookModules,
    path: Path,  # noqa: ARG001 - same callable shape as workbook programs
) -> tuple[Any, ...]:
    defined_name = modules.defined_name
    dn = defined_name.DefinedName(
        "Totals",
        attr_text="Sheet1!$A$1:$A$2",
        comment="rollup",
        localSheetId=0,
        hidden=True,
        function=True,
        functionGroupId=2,
        workbookParameter=True,
    )
    node = dn.to_tree()
    restored = defined_name.DefinedName.from_tree(node)
    return (
        node.tag,
        dict(node.attrib),
        node.text,
        restored.name,
        restored.attr_text,
        restored.comment,
        restored.localSheetId,
        restored.hidden,
        restored.function,
        restored.functionGroupId,
        restored.workbookParameter,
    )


def _comments_hyperlinks_and_merges(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "note"
    ws["A1"].comment = modules.Comment("hello", "me")
    ws["B1"] = "site"
    ws["B1"].hyperlink = "https://example.com"
    ws["B1"].style = "Hyperlink"
    ws["A3"] = "merged"
    ws.merge_cells("A3:B4")
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded["Sheet1"]
    comment = ws2["A1"].comment
    hyperlink = ws2["B1"].hyperlink
    result = (
        comment.text if comment else None,
        comment.author if comment else None,
        ws2["B1"].value,
        hyperlink.target if hyperlink else None,
        ws2["B1"].style,
        ws2["A3"].value,
        ws2["B3"].value,
        type(ws2["B3"]).__name__,
        ws2["B3"].number_format,
        sorted(str(rng) for rng in ws2.merged_cells.ranges),
    )
    _close(reloaded)
    return result


def _merged_ranges_shape_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B2")
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ranges = reloaded.active.merged_cells.ranges
    range_objects = list(ranges)
    result = (
        type(ranges).__name__,
        tuple(sorted(type(rng).__name__ for rng in range_objects)),
        tuple(sorted(str(rng) for rng in range_objects)),
        "A1:B2" in ranges,
    )
    _close(reloaded)
    return result


def _merged_cells_collection_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    merged = ws.merged_cells
    fresh = (
        type(merged).__name__,
        isinstance(merged, modules.xl.worksheet.cell_range.MultiCellRange),
        type(merged.ranges).__name__,
        str(merged),
        "A1:B2" in merged,
    )
    merged.add("A1:B2")
    after_add = (
        str(merged),
        "A1:B2" in merged,
        "A1" in merged,
        "B2" in merged,
        "Z1" in merged,
        tuple(sorted(str(rng) for rng in merged.ranges)),
        type(ws["B2"]).__name__,
    )
    merged.remove("A1:B2")
    after_remove = (str(merged), tuple(sorted(str(rng) for rng in merged.ranges)))
    merged.add("C1:D2")
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    reloaded_merged = reloaded.active.merged_cells
    result = (
        fresh,
        after_add,
        after_remove,
        (
            str(reloaded_merged),
            "C1:D2" in reloaded_merged,
            "C1" in reloaded_merged,
            "D2" in reloaded_merged,
            tuple(sorted(str(rng) for rng in reloaded_merged.ranges)),
            type(reloaded.active["D2"]).__name__,
        ),
    )
    _close(reloaded)
    return result


def _merge_cells_coordinate_args_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    ws.merge_cells("C1:D2", start_row=9, start_column=9, end_row=10, end_column=10)
    ws.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
    in_memory = tuple(sorted(str(rng) for rng in ws.merged_cells.ranges))
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ranges = tuple(sorted(str(rng) for rng in reloaded.active.merged_cells.ranges))
    result = (in_memory, ranges, type(reloaded.active["D2"]).__name__)
    _close(reloaded)
    return result


def _table_display_name_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["Name", "Amount"])
    ws.append(["A", 1])
    ws.append(["B", 2])
    table = modules.Table(displayName="Sales", ref="A1:B3")
    table.tableStyleInfo = modules.TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
    )
    ws.add_table(table)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    table2 = reloaded.active.tables["Sales"]
    result = (
        table2.ref,
        table2.displayName,
        table2.tableStyleInfo.name,
        table2.tableStyleInfo.showRowStripes,
    )
    _close(reloaded)
    return result


def _table_list_shape_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["Name", "Amount"])
    ws.append(["A", 1])
    table = modules.Table(displayName="Sales", ref="A1:B2")
    ws.tables.add(table)

    fresh = (
        type(ws.tables).__name__,
        isinstance(ws.tables, dict),
        list(ws.tables),
        list(ws.tables.keys()),
        list(ws.tables.items()),
        "Sales" in ws.tables,
    )
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    tables = reloaded.active.tables
    result = (
        fresh,
        (
            type(tables).__name__,
            isinstance(tables, dict),
            list(tables),
            list(tables.keys()),
            list(tables.items()),
            tables["Sales"].ref,
            "Sales" in tables,
        ),
    )
    _close(reloaded)
    return result


def _freeze_and_dimensions_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.freeze_panes = "B2"
    ws.row_dimensions[2].height = 24
    ws.column_dimensions["C"].width = 18
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    result = (
        ws2.freeze_panes,
        ws2.row_dimensions[2].height,
        ws2.column_dimensions["C"].width,
    )
    _close(reloaded)
    return result


def _column_dimensions_mapping_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    fresh = (
        len(ws.column_dimensions),
        list(ws.column_dimensions),
        list(ws.column_dimensions.keys()),
        ws.column_dimensions.get("A", "missing"),
    )
    dim_a = ws.column_dimensions["A"]
    after_getitem = (
        len(ws.column_dimensions),
        list(ws.column_dimensions),
        list(ws.column_dimensions.keys()),
        [(key, dim.width, dim.customWidth) for key, dim in ws.column_dimensions.items()],
        dim_a.width,
        dim_a.customWidth,
        dim_a.hidden,
        dim_a.outlineLevel,
        ws.column_dimensions.get("D", "missing"),
    )
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].outlineLevel = 2
    ws.column_dimensions["D"].width = 18
    before_save = (
        list(ws.column_dimensions.keys()),
        [
            (key, dim.width, dim.customWidth, dim.hidden, dim.outlineLevel)
            for key, dim in ws.column_dimensions.items()
        ],
    )
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    result = (
        fresh,
        after_getitem,
        before_save,
        (
            list(ws2.column_dimensions.keys()),
            [
                (key, dim.width, dim.customWidth, dim.hidden, dim.outlineLevel)
                for key, dim in ws2.column_dimensions.items()
            ],
        ),
    )
    _close(reloaded)
    return result


def _conditional_formatting_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=row)
    fill = modules.PatternFill(fill_type="solid", fgColor="FFFF0000")
    rule = modules.CellIsRule(operator="greaterThan", formula=["1"], fill=fill)
    ws.conditional_formatting.add("A1:A3", rule)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    result = []
    for cf_range in ws2.conditional_formatting:
        for rule in ws2.conditional_formatting[cf_range]:
            dxf = getattr(rule, "dxf", None)
            fill = getattr(dxf, "fill", None)
            color = getattr(fill, "fgColor", None) if fill is not None else None
            rgb = getattr(color, "rgb", color)
            result.append(
                (
                    str(cf_range.sqref),
                    rule.type,
                    rule.operator,
                    tuple(str(part) for part in rule.formula),
                    dxf is not None,
                    rgb,
                )
            )
    _close(reloaded)
    return tuple(result)


def _openpyxl_keyword_aliases_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 5
    ws.add_data_validation(
        data_validation=modules.datavalidation.DataValidation(type="whole")
    )
    ws["A1"].comment = modules.Comment("note", "me")
    ws["A1"].comment.bind(cell=ws["A1"])
    ws.conditional_formatting.add(
        "A1:A2",
        cfRule=modules.CellIsRule(operator="greaterThan", formula=["1"]),
    )
    copied = wb.copy_worksheet(from_worksheet=ws)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        tuple(reloaded.sheetnames),
        copied.title,
        reloaded["Data"]["A1"].comment.text,
        len(list(reloaded["Data"].data_validations.dataValidation)),
        len(reloaded["Data"].conditional_formatting),
    )
    _close(reloaded)
    return result


def _openpyxl_direct_constructors(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook(False, True)
    default_ws = modules.Worksheet(wb)
    duplicate_title_ws = modules.Worksheet(wb, "Sheet")
    direct_ws = modules.Worksheet(wb, "Direct")
    direct_cell = modules.Cell(wb.active, 1, 2, "constructor value")
    keyword_cell = modules.Cell(
        wb.active,
        row=2,
        column=3,
        value="keyword value",
    )
    wb.active["B1"] = direct_cell.value
    wb.active["C2"] = keyword_cell.value
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        getattr(wb, "iso_dates", None),
        default_ws.title,
        duplicate_title_ws.title,
        direct_ws.title,
        direct_cell.coordinate,
        direct_cell.value,
        keyword_cell.coordinate,
        keyword_cell.value,
        tuple(reloaded.sheetnames),
        reloaded.active["B1"].value,
        reloaded.active["C2"].value,
    )
    _close(reloaded)
    return result


def _style_and_name_positional_constructors(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    font = modules.styles.Font(
        "Calibri",
        11,
        True,
        True,
        1,
        "single",
        True,
        "FF0000",
        "minor",
        2,
        12,
        False,
        False,
        False,
        "double",
        "superscript",
        True,
        True,
        True,
        True,
    )
    fill = modules.PatternFill(
        "solid",
        "FFFF0000",
        "FF00FF00",
        "darkGrid",
        "FF0000FF",
        "FFFFFF00",
    )
    alignment = modules.styles.Alignment(
        "center",
        "top",
        45,
        True,
        False,
        2,
        1,
        True,
        2,
        90,
        False,
        True,
        None,
    )
    named_style = modules.styles.NamedStyle(
        "PositionalStyle",
        None,
        None,
        None,
        None,
        "0.00",
        None,
        1,
        True,
    )
    defined_name = modules.defined_name.DefinedName(
        "PositionalName",
        "comment",
        "menu",
        "desc",
        "help",
        "status",
        0,
        True,
        False,
        False,
        False,
        1,
        "K",
        False,
        True,
        "Sheet!$A$1",
    )

    wb = xl.Workbook()
    ws = wb.active
    ws["A1"] = "styled"
    ws["A1"].font = font
    ws["A1"].fill = fill
    ws["A1"].alignment = alignment
    wb.add_named_style(named_style)
    wb.defined_names.add(defined_name)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        font.name,
        font.sz,
        font.b,
        font.i,
        font.charset,
        font.u,
        font.strike,
        _visible_rgb(font.color),
        font.scheme,
        font.family,
        font.size,
        font.bold,
        font.italic,
        font.strikethrough,
        font.underline,
        font.vertAlign,
        font.outline,
        font.shadow,
        font.condense,
        font.extend,
        fill.patternType,
        _visible_rgb(fill.fgColor),
        _visible_rgb(fill.bgColor),
        fill.fill_type,
        _visible_rgb(fill.start_color),
        _visible_rgb(fill.end_color),
        alignment.horizontal,
        alignment.vertical,
        alignment.textRotation,
        alignment.wrapText,
        alignment.shrinkToFit,
        alignment.indent,
        alignment.relativeIndent,
        alignment.justifyLastLine,
        alignment.readingOrder,
        alignment.text_rotation,
        alignment.wrap_text,
        alignment.shrink_to_fit,
        named_style.name,
        named_style.number_format,
        named_style.builtinId,
        named_style.hidden,
        defined_name.name,
        defined_name.comment,
        defined_name.customMenu,
        defined_name.description,
        defined_name.help,
        defined_name.statusBar,
        defined_name.localSheetId,
        defined_name.hidden,
        defined_name.function,
        defined_name.vbProcedure,
        defined_name.xlm,
        defined_name.functionGroupId,
        defined_name.shortcutKey,
        defined_name.publishToServer,
        defined_name.workbookParameter,
        defined_name.attr_text,
        reloaded.active["A1"].value,
    )
    _close(reloaded)
    return result


def _sheet_state_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.active.title = "Visible"
    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    very_hidden = wb.create_sheet("VeryHidden")
    very_hidden.sheet_state = "veryHidden"
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        tuple(reloaded.sheetnames),
        reloaded["Visible"].sheet_state,
        reloaded["Hidden"].sheet_state,
        reloaded["VeryHidden"].sheet_state,
    )
    _close(reloaded)
    return result


def _active_sheet_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.active.title = "First"
    wb.create_sheet("Second")
    wb.active = 1
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        tuple(reloaded.sheetnames),
        reloaded.active.title,
    )
    _close(reloaded)
    return result


def _sheet_view_tab_selected_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    first = wb.active
    first.title = "First"
    first.sheet_view.tabSelected = False
    second = wb.create_sheet("Second")
    second.sheet_view.tabSelected = True
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        reloaded.active.title,
        tuple(
            getattr(reloaded[name].sheet_view, "tabSelected", None)
            for name in reloaded.sheetnames
        ),
    )
    _close(reloaded)
    return result


def _sheet_view_defaults_and_flags_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    default = wb.active
    default.title = "Default"
    explicit = wb.create_sheet("Explicit")
    explicit.sheet_view.showGridLines = True
    explicit.sheet_view.showRowColHeaders = False
    explicit.sheet_view.showOutlineSymbols = True
    explicit.sheet_view.showZeros = False
    explicit.sheet_view.rightToLeft = True
    explicit.sheet_view.selection[0].activeCell = "B2"
    explicit.sheet_view.selection[0].sqref = "B2:C3"

    fresh = (
        default.sheet_view.showGridLines,
        default.sheet_view.showRowColHeaders,
        default.sheet_view.showOutlineSymbols,
        default.sheet_view.showZeros,
        default.sheet_view.rightToLeft,
        default.sheet_view.selection[0].activeCell,
        default.sheet_view.selection[0].sqref,
    )
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    default_view = reloaded["Default"].sheet_view
    explicit_view = reloaded["Explicit"].sheet_view
    result = (
        fresh,
        (
            default_view.showGridLines,
            default_view.showRowColHeaders,
            default_view.showOutlineSymbols,
            default_view.showZeros,
            default_view.rightToLeft,
            default_view.selection[0].activeCell,
            default_view.selection[0].sqref,
        ),
        (
            explicit_view.showGridLines,
            explicit_view.showRowColHeaders,
            explicit_view.showOutlineSymbols,
            explicit_view.showZeros,
            explicit_view.rightToLeft,
            explicit_view.selection[0].activeCell,
            explicit_view.selection[0].sqref,
        ),
    )
    _close(reloaded)
    return result


def _sheet_outline_properties_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    default = wb.active
    default.title = "Default"

    explicit_false = wb.create_sheet("ExplicitFalse")
    explicit_false.sheet_properties.outlinePr.applyStyles = False
    explicit_false.sheet_properties.outlinePr.showOutlineSymbols = False

    explicit_true = wb.create_sheet("ExplicitTrue")
    explicit_true.sheet_properties.outlinePr.applyStyles = True
    explicit_true.sheet_properties.outlinePr.showOutlineSymbols = True
    explicit_true.sheet_properties.outlinePr.summaryBelow = False
    explicit_true.sheet_properties.outlinePr.summaryRight = False

    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = tuple(
        (
            reloaded[name].sheet_properties.outlinePr.summaryBelow,
            reloaded[name].sheet_properties.outlinePr.summaryRight,
            reloaded[name].sheet_properties.outlinePr.applyStyles,
            reloaded[name].sheet_properties.outlinePr.showOutlineSymbols,
        )
        for name in ("Default", "ExplicitFalse", "ExplicitTrue")
    )
    _close(reloaded)
    return result


def _sheet_page_setup_properties_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    default = wb.active
    default.title = "Default"

    explicit_false = wb.create_sheet("ExplicitFalse")
    explicit_false.sheet_properties.pageSetUpPr.autoPageBreaks = False
    explicit_false.sheet_properties.pageSetUpPr.fitToPage = False

    explicit_true = wb.create_sheet("ExplicitTrue")
    explicit_true.sheet_properties.pageSetUpPr.autoPageBreaks = True
    explicit_true.sheet_properties.pageSetUpPr.fitToPage = True

    fresh = (
        default.sheet_properties.pageSetUpPr.autoPageBreaks,
        default.sheet_properties.pageSetUpPr.fitToPage,
    )
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        fresh,
        (
            reloaded["Default"].sheet_properties.pageSetUpPr.autoPageBreaks,
            reloaded["Default"].sheet_properties.pageSetUpPr.fitToPage,
        ),
        (
            reloaded["ExplicitFalse"].sheet_properties.pageSetUpPr.autoPageBreaks,
            reloaded["ExplicitFalse"].sheet_properties.pageSetUpPr.fitToPage,
        ),
        (
            reloaded["ExplicitTrue"].sheet_properties.pageSetUpPr.autoPageBreaks,
            reloaded["ExplicitTrue"].sheet_properties.pageSetUpPr.fitToPage,
        ),
    )
    _close(reloaded)
    return result


def _sheet_format_defaults_and_flags_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    default = wb.active
    default.title = "Default"

    explicit_false = wb.create_sheet("ExplicitFalse")
    explicit_false.sheet_format.customHeight = False
    explicit_false.sheet_format.zeroHeight = False
    explicit_false.sheet_format.thickTop = False
    explicit_false.sheet_format.thickBottom = False

    explicit_true = wb.create_sheet("ExplicitTrue")
    explicit_true.sheet_format.customHeight = True
    explicit_true.sheet_format.zeroHeight = True
    explicit_true.sheet_format.thickTop = True
    explicit_true.sheet_format.thickBottom = True

    fresh = (
        default.sheet_format.customHeight,
        default.sheet_format.zeroHeight,
        default.sheet_format.thickTop,
        default.sheet_format.thickBottom,
    )
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = (
        fresh,
        (
            reloaded["Default"].sheet_format.customHeight,
            reloaded["Default"].sheet_format.zeroHeight,
            reloaded["Default"].sheet_format.thickTop,
            reloaded["Default"].sheet_format.thickBottom,
        ),
        (
            reloaded["ExplicitFalse"].sheet_format.customHeight,
            reloaded["ExplicitFalse"].sheet_format.zeroHeight,
            reloaded["ExplicitFalse"].sheet_format.thickTop,
            reloaded["ExplicitFalse"].sheet_format.thickBottom,
        ),
        (
            reloaded["ExplicitTrue"].sheet_format.customHeight,
            reloaded["ExplicitTrue"].sheet_format.zeroHeight,
            reloaded["ExplicitTrue"].sheet_format.thickTop,
            reloaded["ExplicitTrue"].sheet_format.thickBottom,
        ),
    )
    _close(reloaded)
    return result


def _workbook_view_defaults_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    view = reloaded.views[0]
    result = (
        view.visibility,
        view.minimized,
        view.showHorizontalScroll,
        view.showVerticalScroll,
        view.showSheetTabs,
        view.xWindow,
        view.yWindow,
        view.windowWidth,
        view.windowHeight,
        view.tabRatio,
        view.firstSheet,
        view.activeTab,
        view.autoFilterDateGrouping,
    )
    _close(reloaded)
    return result


def _workbook_view_settings_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.create_sheet("Second")
    view = wb.views[0]
    view.visibility = "hidden"
    view.minimized = True
    view.showHorizontalScroll = False
    view.showVerticalScroll = False
    view.showSheetTabs = False
    view.xWindow = 10
    view.yWindow = 20
    view.windowWidth = 12000
    view.windowHeight = 8000
    view.tabRatio = 750
    view.firstSheet = 1
    view.activeTab = 1
    view.autoFilterDateGrouping = False
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    reloaded_view = reloaded.views[0]
    result = (
        reloaded_view.visibility,
        reloaded_view.minimized,
        reloaded_view.showHorizontalScroll,
        reloaded_view.showVerticalScroll,
        reloaded_view.showSheetTabs,
        reloaded_view.xWindow,
        reloaded_view.yWindow,
        reloaded_view.windowWidth,
        reloaded_view.windowHeight,
        reloaded_view.tabRatio,
        reloaded_view.firstSheet,
        reloaded_view.activeTab,
        reloaded_view.autoFilterDateGrouping,
        reloaded.active.title,
    )
    _close(reloaded)
    return result


def _named_style_number_format_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    styles = modules.styles
    wb = xl.Workbook()
    style = styles.NamedStyle(name="Money")
    style.number_format = "$#,##0.00"
    wb.add_named_style(style)
    ws = wb.active
    ws["A1"] = 12.5
    ws["A1"].style = "Money"
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    cell = reloaded.active["A1"]
    result = (
        cell.value,
        cell.style,
        cell.number_format,
    )
    _close(reloaded)
    return result


def _sheet_tab_color_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> Any:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.sheet_properties.tabColor = "FF0000"
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    color = reloaded.active.sheet_properties.tabColor
    result = getattr(color, "rgb", color)
    _close(reloaded)
    return result


def _write_mode_structural_edits(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    for row in range(1, 4):
        for col in range(1, 4):
            ws.cell(row=row, column=col, value=row * 10 + col)
    ws.insert_rows(2)
    ws.delete_cols(3)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    result = tuple(
        tuple(
            ws2.cell(row=row, column=col).value
            for col in range(1, 4)
        )
        for row in range(1, 5)
    )
    _close(reloaded)
    return result


def _write_mode_move_range_translate(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    ws["B1"] = "=A1+1"
    ws["C1"] = "f"
    ws.move_range("A1:C1", rows=1, cols=1, translate=True)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path, data_only=False)
    ws2 = reloaded.active
    result = (
        ws2["A1"].value,
        ws2["B2"].value,
        ws2["C2"].value,
        ws2["D2"].value,
    )
    _close(reloaded)
    return result


def _auto_filter_value_filter_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    ws.auto_filter.ref = "A1:B3"
    ws.auto_filter.add_filter_column(0, ["A"])
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    auto_filter = reloaded.active.auto_filter
    columns = list(auto_filter.filterColumn)
    result = (
        auto_filter.ref,
        len(columns),
        columns[0].colId,
        tuple(columns[0].filters.filter),
    )
    _close(reloaded)
    return result


def _private_style_copy_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    from copy import copy

    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws["A1"] = "source"
    ws["A1"].font = modules.styles.Font(bold=True)
    ws["A1"].comment = modules.Comment("note", "me")
    ws["B1"] = "target"
    ws["B1"]._style = copy(ws["A1"]._style)
    ws["B1"].comment = copy(ws["A1"].comment)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    target = reloaded.active["B1"]
    result = (
        target.value,
        target.font.bold,
        target.comment.text if target.comment else None,
        target.comment.author if target.comment else None,
    )
    _close(reloaded)
    return result


def _remove_sheet_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[str, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.active.title = "A"
    removed = wb.create_sheet("B")
    wb.remove(removed)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = tuple(reloaded.sheetnames)
    _close(reloaded)
    return result


def _append_dict_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.append({"A": "name", "C": "amount"})
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded.active
    result = (ws2["A1"].value, ws2["B1"].value, ws2["C1"].value)
    _close(reloaded)
    return result


def _named_style_full_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    styles = modules.styles
    wb = xl.Workbook()
    style = styles.NamedStyle(name="Metric")
    style.font = styles.Font(bold=True)
    style.fill = styles.PatternFill(fill_type="solid", fgColor="FFFF0000")
    style.number_format = "0.0"
    wb.add_named_style(style)
    ws = wb.active
    ws["A1"] = 1.2
    ws["A1"].style = "Metric"
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    cell = reloaded.active["A1"]
    result = (
        cell.style,
        cell.font.bold,
        _visible_rgb(cell.fill.fgColor),
        cell.number_format,
    )
    _close(reloaded)
    return result


def _create_sheet_index_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[str, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    wb.active.title = "A"
    wb.create_sheet("B")
    wb.create_sheet("Start", 0)
    wb.create_sheet("Neg", -1)
    wb.create_sheet("Far", 99)
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    result = tuple(reloaded.sheetnames)
    _close(reloaded)
    return result


def _print_settings_round_trip(
    modules: WorkbookModules,
    path: Path,
) -> tuple[Any, ...]:
    xl = modules.xl
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Print"
    ws.print_title_rows = "1:2"
    ws.print_title_cols = "A:B"
    ws.print_area = "A1:C10"
    ws.page_margins.left = 0.25
    ws.page_setup.orientation = "landscape"
    ws.print_options.gridLines = True
    wb.save(path)
    _close(wb)

    reloaded = xl.load_workbook(path)
    ws2 = reloaded["Print"]
    result = (
        ws2.print_title_rows,
        ws2.print_title_cols,
        str(ws2.print_area),
        ws2.page_margins.left,
        ws2.page_setup.orientation,
        ws2.print_options.gridLines,
    )
    _close(reloaded)
    return result


PROGRAMS = [
    DropInProgram("basic_cells_and_sheets", _basic_cells_and_sheets),
    DropInProgram("append_loaded_workbook", _append_loaded_workbook),
    DropInProgram("styles_round_trip", _styles_round_trip),
    DropInProgram("data_validation_round_trip", _data_validation_round_trip),
    DropInProgram("defined_name_xml_round_trip", _defined_name_xml_round_trip),
    DropInProgram("comments_hyperlinks_and_merges", _comments_hyperlinks_and_merges),
    DropInProgram("merged_ranges_shape_round_trip", _merged_ranges_shape_round_trip),
    DropInProgram("merged_cells_collection_round_trip", _merged_cells_collection_round_trip),
    DropInProgram(
        "merge_cells_coordinate_args_round_trip",
        _merge_cells_coordinate_args_round_trip,
    ),
    DropInProgram("table_display_name_round_trip", _table_display_name_round_trip),
    DropInProgram("table_list_shape_round_trip", _table_list_shape_round_trip),
    DropInProgram("freeze_and_dimensions_round_trip", _freeze_and_dimensions_round_trip),
    DropInProgram(
        "column_dimensions_mapping_round_trip",
        _column_dimensions_mapping_round_trip,
    ),
    DropInProgram("conditional_formatting_round_trip", _conditional_formatting_round_trip),
    DropInProgram("openpyxl_keyword_aliases_round_trip", _openpyxl_keyword_aliases_round_trip),
    DropInProgram("openpyxl_direct_constructors", _openpyxl_direct_constructors),
    DropInProgram(
        "style_and_name_positional_constructors",
        _style_and_name_positional_constructors,
    ),
    DropInProgram("sheet_state_round_trip", _sheet_state_round_trip),
    DropInProgram("active_sheet_round_trip", _active_sheet_round_trip),
    DropInProgram(
        "named_style_number_format_round_trip",
        _named_style_number_format_round_trip,
    ),
    DropInProgram("sheet_view_tab_selected_round_trip", _sheet_view_tab_selected_round_trip),
    DropInProgram(
        "sheet_view_defaults_and_flags_round_trip",
        _sheet_view_defaults_and_flags_round_trip,
    ),
    DropInProgram(
        "sheet_outline_properties_round_trip",
        _sheet_outline_properties_round_trip,
    ),
    DropInProgram(
        "sheet_page_setup_properties_round_trip",
        _sheet_page_setup_properties_round_trip,
    ),
    DropInProgram(
        "sheet_format_defaults_and_flags_round_trip",
        _sheet_format_defaults_and_flags_round_trip,
    ),
    DropInProgram("workbook_view_defaults_round_trip", _workbook_view_defaults_round_trip),
    DropInProgram("workbook_view_settings_round_trip", _workbook_view_settings_round_trip),
    DropInProgram("sheet_tab_color_round_trip", _sheet_tab_color_round_trip),
    DropInProgram("write_mode_structural_edits", _write_mode_structural_edits),
    DropInProgram("write_mode_move_range_translate", _write_mode_move_range_translate),
    DropInProgram(
        "auto_filter_value_filter_round_trip",
        _auto_filter_value_filter_round_trip,
    ),
    DropInProgram("private_style_copy_round_trip", _private_style_copy_round_trip),
    DropInProgram("remove_sheet_round_trip", _remove_sheet_round_trip),
    DropInProgram("append_dict_round_trip", _append_dict_round_trip),
    DropInProgram("named_style_full_round_trip", _named_style_full_round_trip),
    DropInProgram("create_sheet_index_round_trip", _create_sheet_index_round_trip),
    DropInProgram("print_settings_round_trip", _print_settings_round_trip),
]


@pytest.mark.parametrize("program", PROGRAMS, ids=lambda program: program.label)
def test_openpyxl_user_program_matches_wolfxl(
    program: DropInProgram,
    tmp_path: Path,
) -> None:
    expected = program.run(OPENPYXL_MODULES, tmp_path / f"openpyxl-{program.label}.xlsx")
    actual = program.run(WOLFXL_MODULES, tmp_path / f"wolfxl-{program.label}.xlsx")
    assert actual == expected
