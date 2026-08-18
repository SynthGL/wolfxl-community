"""Read-only openpyxl parity for sparse and malformed worksheet dimensions."""

from __future__ import annotations

from pathlib import Path
import re
import zipfile

import openpyxl
import pytest

import wolfxl


def _make_sparse_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["B2"] = "start"
    ws["E5"] = "end"
    ws["C10"] = "tail"
    wb.save(path)
    wb.close()
    return path


def _rewrite_sheet_dimension(src: Path, dst: Path, ref: str | None) -> Path:
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(
        dst, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                if ref is None:
                    data = re.sub(rb"<dimension[^>]*/>", b"", data, count=1)
                else:
                    data = re.sub(
                        rb'(<dimension[^>]*ref=")[^"]+("[^>]*/>)',
                        rf"\g<1>{ref}\g<2>".encode("ascii"),
                        data,
                        count=1,
                    )
            zout.writestr(info, data)
    return dst


def _read_only_snapshot(path: Path, module: object) -> dict[str, object]:
    wb = module.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sparse"]
    try:
        snapshot: dict[str, object] = {
            "min_row": ws.min_row,
            "min_column": ws.min_column,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "dimension": ws.calculate_dimension(),
            "values": list(ws.iter_rows(values_only=True)),
            "bounded_values": list(
                ws.iter_rows(
                    min_row=1,
                    max_row=10,
                    min_col=1,
                    max_col=5,
                    values_only=True,
                )
            ),
        }
    finally:
        wb.close()
    return snapshot


def test_read_only_sparse_used_range_matches_openpyxl(tmp_path: Path) -> None:
    path = _make_sparse_workbook(tmp_path / "sparse.xlsx")

    assert _read_only_snapshot(path, wolfxl) == _read_only_snapshot(path, openpyxl)


def test_read_only_absent_dimension_matches_openpyxl_and_can_force(
    tmp_path: Path,
) -> None:
    base = _make_sparse_workbook(tmp_path / "sparse.xlsx")
    path = _rewrite_sheet_dimension(base, tmp_path / "absent-dimension.xlsx", None)

    expected = openpyxl.load_workbook(path, read_only=True, data_only=True)
    actual = wolfxl.load_workbook(path, read_only=True, data_only=True)
    try:
        expected_ws = expected["Sparse"]
        actual_ws = actual["Sparse"]

        assert actual_ws.max_row == expected_ws.max_row is None
        assert actual_ws.max_column == expected_ws.max_column is None
        with pytest.raises(ValueError, match="Worksheet is unsized"):
            actual_ws.calculate_dimension()
        with pytest.raises(ValueError, match="Worksheet is unsized"):
            expected_ws.calculate_dimension()

        assert list(actual_ws.iter_rows(values_only=True)) == list(
            expected_ws.iter_rows(values_only=True)
        )
        assert actual_ws.calculate_dimension(force=True) == expected_ws.calculate_dimension(
            force=True
        )
        assert actual_ws.max_row == expected_ws.max_row == 10
        assert actual_ws.max_column == expected_ws.max_column == 5
        assert list(actual_ws.iter_rows(values_only=True)) == list(
            expected_ws.iter_rows(values_only=True)
        )
    finally:
        expected.close()
        actual.close()


def test_read_only_stale_dimension_matches_openpyxl_until_reset_force(
    tmp_path: Path,
) -> None:
    base = _make_sparse_workbook(tmp_path / "sparse.xlsx")
    path = _rewrite_sheet_dimension(base, tmp_path / "stale-dimension.xlsx", "A1:A1")

    expected = openpyxl.load_workbook(path, read_only=True, data_only=True)
    actual = wolfxl.load_workbook(path, read_only=True, data_only=True)
    try:
        expected_ws = expected["Sparse"]
        actual_ws = actual["Sparse"]

        assert actual_ws.max_row == expected_ws.max_row == 1
        assert actual_ws.max_column == expected_ws.max_column == 1
        assert actual_ws.calculate_dimension() == expected_ws.calculate_dimension() == "A1:A1"
        assert list(actual_ws.iter_rows(values_only=True)) == list(
            expected_ws.iter_rows(values_only=True)
        )

        expected_ws.reset_dimensions()
        actual_ws.reset_dimensions()
        assert actual_ws.max_row == expected_ws.max_row is None
        assert actual_ws.max_column == expected_ws.max_column is None
        with pytest.raises(ValueError, match="Worksheet is unsized"):
            actual_ws.calculate_dimension()
        assert actual_ws.calculate_dimension(force=True) == expected_ws.calculate_dimension(
            force=True
        )
        assert actual_ws.max_row == expected_ws.max_row == 10
        assert actual_ws.max_column == expected_ws.max_column == 5
        assert list(actual_ws.iter_rows(values_only=True)) == list(
            expected_ws.iter_rows(values_only=True)
        )
    finally:
        expected.close()
        actual.close()
