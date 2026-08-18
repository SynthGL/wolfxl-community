"""Public openpyxl reader behavior parity for generated workbooks."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

import wolfxl


def _make_reader_sample(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["text", "integer", "float", "date", "datetime", "bool", "formula"])
    ws.append(
        [
            "alpha",
            42,
            3.5,
            date(2026, 5, 13),
            datetime(2026, 5, 13, 14, 30),
            True,
            "=B2+C2",
        ]
    )
    ws.append(["beta", -7, 0, None, None, False, "=B3+C3"])
    wb.save(path)
    wb.close()
    return path


def _signature(value: Any) -> tuple[type[Any], Any]:
    return type(value), value


def _value_signatures(values: list[tuple[Any, ...]]) -> list[tuple[tuple[type[Any], Any], ...]]:
    return [tuple(_signature(value) for value in row) for row in values]


def _cell_signatures(cells: Any) -> Any:
    if cells and isinstance(cells[0], tuple):
        return tuple(tuple(_cell_signature(cell) for cell in row) for row in cells)
    return tuple(_cell_signature(cell) for cell in cells)


def _cell_value_signatures(cells: Any) -> Any:
    if cells and isinstance(cells[0], tuple):
        return tuple(tuple(_signature(cell.value) for cell in row) for row in cells)
    return tuple(_signature(cell.value) for cell in cells)


def _cell_signature(cell: Any) -> tuple[str, str, tuple[type[Any], Any]]:
    return (
        cell.coordinate,
        cell.data_type,
        _signature(cell.value),
    )


def _assert_reader_behavior_matches(path: Path, *, read_only: bool) -> None:
    expected = openpyxl.load_workbook(path, read_only=read_only, data_only=False)
    actual = wolfxl.load_workbook(path, read_only=read_only, data_only=False)
    try:
        expected_ws = expected["Data"]
        actual_ws = actual["Data"]

        assert actual.sheetnames == expected.sheetnames
        assert actual_ws.max_row == expected_ws.max_row == 3
        assert actual_ws.max_column == expected_ws.max_column == 7

        for key in ("A1", "B2", "C2", "D2", "E2", "F2", "G2", "G3"):
            assert _cell_signature(actual_ws[key]) == _cell_signature(expected_ws[key])

        assert _cell_signatures(actual_ws["A1:C2"]) == _cell_signatures(expected_ws["A1:C2"])
        assert _cell_value_signatures(actual_ws["2:3"]) == _cell_value_signatures(
            expected_ws["2:3"]
        )

        expected_rows = list(
            expected_ws.iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=7,
                values_only=True,
            )
        )
        actual_rows = list(
            actual_ws.iter_rows(
                min_row=1,
                max_row=3,
                min_col=1,
                max_col=7,
                values_only=True,
            )
        )
        assert _value_signatures(actual_rows) == _value_signatures(expected_rows)
        assert _value_signatures(list(actual_ws.values)) == _value_signatures(
            list(expected_ws.values)
        )
    finally:
        expected.close()
        actual.close()


def test_eager_reader_matches_openpyxl_public_behavior(tmp_path: Path) -> None:
    path = _make_reader_sample(tmp_path / "reader-sample.xlsx")

    _assert_reader_behavior_matches(path, read_only=False)


def test_read_only_reader_matches_openpyxl_public_behavior(tmp_path: Path) -> None:
    path = _make_reader_sample(tmp_path / "reader-sample.xlsx")

    _assert_reader_behavior_matches(path, read_only=True)


def test_data_only_reader_blanks_uncached_formulas_like_openpyxl(tmp_path: Path) -> None:
    path = _make_reader_sample(tmp_path / "reader-sample.xlsx")

    for read_only in (False, True):
        expected = openpyxl.load_workbook(path, read_only=read_only, data_only=True)
        actual = wolfxl.load_workbook(path, read_only=read_only, data_only=True)
        try:
            expected_ws = expected["Data"]
            actual_ws = actual["Data"]

            assert _signature(actual_ws["G2"].value) == _signature(expected_ws["G2"].value)
            assert _signature(actual_ws["G3"].value) == _signature(expected_ws["G3"].value)
            assert _value_signatures(list(actual_ws.values)) == _value_signatures(
                list(expected_ws.values)
            )
        finally:
            expected.close()
            actual.close()
