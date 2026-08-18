"""Focused openpyxl-oracle tests for workbook/sheet protection behavior."""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils.protection import hash_password as openpyxl_hash_password

import wolfxl
from wolfxl.utils.protection import check_password, hash_password
from wolfxl.workbook.protection import FileSharing, WorkbookProtection
from wolfxl.worksheet.protection import SheetProtection


@pytest.fixture(autouse=True)
def _force_test_epoch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("WOLFXL_TEST_EPOCH", "0")


def _workbook_xml(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        return zf.read("xl/workbook.xml").decode("utf-8")


def test_sheet_protection_set_password_disable_and_ctor_match_openpyxl() -> None:
    ref = openpyxl.worksheet.protection.SheetProtection()
    got = SheetProtection()

    ref.set_password("hunter2")
    got.set_password("hunter2")
    assert got.sheet == ref.sheet is True
    assert got.password == ref.password == "C258"
    assert list(got) == list(ref)
    assert got.check_password("hunter2") is True

    ref.disable()
    got.disable()
    assert got.sheet == ref.sheet is False
    assert got.password == ref.password == "C258"
    assert list(got) == list(ref)

    ref_ctor = openpyxl.worksheet.protection.SheetProtection(password="secret")
    got_ctor = SheetProtection(password="secret")
    assert got_ctor.sheet == ref_ctor.sheet is True
    assert got_ctor.password == ref_ctor.password == openpyxl_hash_password("secret")
    assert list(got_ctor) == list(ref_ctor)


def test_hash_password_and_check_password_match_openpyxl_hashes() -> None:
    for plaintext in ("", "secret", "hunter2"):
        assert hash_password(plaintext) == openpyxl_hash_password(plaintext)

    expected = openpyxl_hash_password("secret")
    assert check_password("secret", expected) is True
    assert check_password("wrong", expected) is False


def test_workbook_protection_aliases_and_attr_iteration_match_openpyxl() -> None:
    ref = openpyxl.workbook.protection.WorkbookProtection(
        workbookPassword="secret",
        lockStructure=True,
        lockWindows=True,
    )
    got = WorkbookProtection(
        workbookPassword="secret",
        lockStructure=True,
        lockWindows=True,
    )

    assert got.lockStructure == ref.lockStructure is True
    assert got.lockWindows == ref.lockWindows is True
    assert got.lockRevision == ref.lockRevision
    assert got.workbookPassword == ref.workbookPassword == openpyxl_hash_password("secret")
    assert got.check_workbook_password("secret") is True
    assert list(got) == list(ref)


def test_file_sharing_aliases_and_attr_iteration_match_openpyxl() -> None:
    password_hash = openpyxl_hash_password("reserve")
    ref = openpyxl.workbook.protection.FileSharing(
        readOnlyRecommended=True,
        userName="alice",
        reservationPassword=password_hash,
    )
    got = FileSharing(
        readOnlyRecommended=True,
        userName="alice",
        reservationPassword=password_hash,
    )

    assert got.readOnlyRecommended == ref.readOnlyRecommended is True
    assert got.userName == ref.userName == "alice"
    assert got.reservationPassword == ref.reservationPassword == password_hash
    assert got.check_reservation_password("reserve") is True
    assert list(got) == list(ref)


def test_sheet_workbook_and_file_sharing_save_reload_match_openpyxl(tmp_path: Path) -> None:
    path = tmp_path / "protection.xlsx"
    workbook_password = openpyxl_hash_password("structure")
    reservation_password = openpyxl_hash_password("reserve")

    wb = wolfxl.Workbook()
    ws = wb.active
    ws["A1"] = "protected"
    ws.protection.set_password("sheet")
    wb.security = WorkbookProtection(lockStructure=True, workbookPassword="structure")
    wb.fileSharing = FileSharing(
        readOnlyRecommended=True,
        userName="alice",
        reservationPassword=reservation_password,
    )
    wb.save(path)

    xml = _workbook_xml(path)
    assert f'workbookPassword="{workbook_password}"' in xml
    assert f'reservationPassword="{reservation_password}"' in xml

    ref = openpyxl.load_workbook(path)
    try:
        assert ref.active.protection.sheet is True
        assert ref.active.protection.password == openpyxl_hash_password("sheet")
        assert ref.security.lockStructure is True
        assert ref.security.workbookPassword == workbook_password
    finally:
        ref.close()

    got = wolfxl.load_workbook(path)
    try:
        assert got.active.protection.sheet is True
        assert got.active.protection.check_password("sheet") is True
        assert got.security.lockStructure is True
        assert got.security.workbookPassword == workbook_password
        assert got.fileSharing.readOnlyRecommended is True
        assert got.fileSharing.reservationPassword == reservation_password
    finally:
        got.close()
