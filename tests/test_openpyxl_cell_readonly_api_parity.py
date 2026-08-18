"""Focused openpyxl parity for cell bounds and read-only public shims."""

from __future__ import annotations

from collections.abc import Callable
from typing import Any

import openpyxl
import pytest

import wolfxl


def _exception_signature(call: Callable[[], Any]) -> tuple[type[BaseException], str]:
    try:
        call()
    except Exception as exc:  # noqa: BLE001 - openpyxl is the compatibility oracle
        return type(exc), str(exc)
    pytest.fail("openpyxl accepted an invalid coordinate")


@pytest.mark.parametrize(
    ("row", "column"),
    [
        (0, 1),
        (1, 0),
        (0, 0),
        (-1, 1),
        (1, -1),
    ],
)
def test_worksheet_cell_invalid_coordinates_match_openpyxl(row: int, column: int) -> None:
    expected_ws = openpyxl.Workbook().active
    actual_ws = wolfxl.Workbook().active

    expected_type, expected_message = _exception_signature(
        lambda: expected_ws.cell(row=row, column=column)
    )

    with pytest.raises(expected_type, match=expected_message):
        actual_ws.cell(row=row, column=column)


@pytest.mark.parametrize(
    ("row_offset", "column_offset"),
    [
        (-1, 0),
        (0, -1),
        (-2, 0),
        (0, -2),
        (-1, -1),
    ],
)
def test_cell_offset_invalid_target_matches_openpyxl(
    row_offset: int,
    column_offset: int,
) -> None:
    expected_cell = openpyxl.Workbook().active.cell(row=1, column=1)
    actual_cell = wolfxl.Workbook().active.cell(row=1, column=1)

    expected_type, expected_message = _exception_signature(
        lambda: expected_cell.offset(row=row_offset, column=column_offset)
    )

    with pytest.raises(expected_type, match=expected_message):
        actual_cell.offset(row=row_offset, column=column_offset)


def test_read_only_public_classes_and_empty_cell_match_openpyxl_surface() -> None:
    from openpyxl.cell.read_only import EMPTY_CELL as OpenpyxlEmptyCell
    from openpyxl.cell.read_only import ReadOnlyCell as OpenpyxlReadOnlyCell
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet as OpenpyxlReadOnlyWorksheet
    from wolfxl.cell.read_only import EMPTY_CELL as WolfxlEmptyCell
    from wolfxl.cell.read_only import ReadOnlyCell as WolfxlReadOnlyCell
    from wolfxl.worksheet._read_only import ReadOnlyWorksheet as WolfxlReadOnlyWorksheet

    assert WolfxlReadOnlyWorksheet.__name__ == OpenpyxlReadOnlyWorksheet.__name__
    assert WolfxlReadOnlyCell.__name__ == OpenpyxlReadOnlyCell.__name__
    assert repr(WolfxlEmptyCell) == repr(OpenpyxlEmptyCell) == "<EmptyCell>"
    assert WolfxlEmptyCell.value == OpenpyxlEmptyCell.value is None
    assert WolfxlEmptyCell.data_type == OpenpyxlEmptyCell.data_type == "n"
    assert not hasattr(WolfxlEmptyCell, "coordinate")
    assert not hasattr(OpenpyxlEmptyCell, "coordinate")
