"""Focused openpyxl oracle tests for worksheet layout/view compatibility."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

import wolfxl


def _close(workbook: Any) -> None:
    close = getattr(workbook, "close", None)
    if close is not None:
        close()


def _freeze_snapshot(ws: Any) -> tuple[Any, ...]:
    pane = ws.sheet_view.pane
    return (
        ws.freeze_panes,
        None
        if pane is None
        else (
            pane.xSplit,
            pane.ySplit,
            pane.topLeftCell,
            pane.activePane,
            pane.state,
        ),
        tuple(
            (selection.pane, selection.activeCell, selection.sqref)
            for selection in ws.sheet_view.selection
        ),
    )


def test_freeze_panes_setter_matches_openpyxl_in_memory() -> None:
    for value in (None, "A1", "A2", "B1", "B2"):
        op_wb = openpyxl.Workbook()
        wx_wb = wolfxl.Workbook()
        try:
            op_wb.active.freeze_panes = value
            wx_wb.active.freeze_panes = value

            assert _freeze_snapshot(wx_wb.active) == _freeze_snapshot(op_wb.active)
        finally:
            _close(op_wb)
            _close(wx_wb)


def test_freeze_panes_accepts_cell_objects_like_openpyxl() -> None:
    op_wb = openpyxl.Workbook()
    wx_wb = wolfxl.Workbook()
    try:
        op_wb.active.freeze_panes = op_wb.active["C4"]
        wx_wb.active.freeze_panes = wx_wb.active["C4"]

        assert _freeze_snapshot(wx_wb.active) == _freeze_snapshot(op_wb.active)
    finally:
        _close(op_wb)
        _close(wx_wb)


def test_freeze_panes_round_trip_matches_openpyxl(tmp_path: Path) -> None:
    for value in ("A1", "A2", "B1", "B2"):
        op_path = tmp_path / f"openpyxl-{value}.xlsx"
        wx_path = tmp_path / f"wolfxl-{value}.xlsx"

        op_wb = openpyxl.Workbook()
        wx_wb = wolfxl.Workbook()
        try:
            op_wb.active.freeze_panes = value
            wx_wb.active.freeze_panes = value
            op_wb.save(op_path)
            wx_wb.save(wx_path)
        finally:
            _close(op_wb)
            _close(wx_wb)

        op_reloaded = openpyxl.load_workbook(op_path)
        wx_reloaded = wolfxl.load_workbook(wx_path)
        try:
            assert _freeze_snapshot(wx_reloaded.active) == _freeze_snapshot(
                op_reloaded.active
            )
        finally:
            _close(op_reloaded)
            _close(wx_reloaded)


def test_print_area_setter_matches_openpyxl_for_common_inputs() -> None:
    for value in (None, "A1:D10", "$A$1:$D$10", ["A1:D10"], ["A1:B2", "D1:E2"]):
        op_wb = openpyxl.Workbook()
        wx_wb = wolfxl.Workbook()
        try:
            op_wb.active.title = "My Sheet"
            wx_wb.active.title = "My Sheet"
            op_wb.active.print_area = value
            wx_wb.active.print_area = value

            assert wx_wb.active.print_area == op_wb.active.print_area
        finally:
            _close(op_wb)
            _close(wx_wb)


def test_page_setup_and_margin_defaults_match_openpyxl() -> None:
    op_wb = openpyxl.Workbook()
    wx_wb = wolfxl.Workbook()
    try:
        assert wx_wb.active.page_setup.orientation == op_wb.active.page_setup.orientation
        assert list(wx_wb.active.page_setup) == list(op_wb.active.page_setup)
        assert (
            wx_wb.active.page_margins.left,
            wx_wb.active.page_margins.right,
            wx_wb.active.page_margins.top,
            wx_wb.active.page_margins.bottom,
            wx_wb.active.page_margins.header,
            wx_wb.active.page_margins.footer,
            list(wx_wb.active.page_margins),
        ) == (
            op_wb.active.page_margins.left,
            op_wb.active.page_margins.right,
            op_wb.active.page_margins.top,
            op_wb.active.page_margins.bottom,
            op_wb.active.page_margins.header,
            op_wb.active.page_margins.footer,
            list(op_wb.active.page_margins),
        )
    finally:
        _close(op_wb)
        _close(wx_wb)
