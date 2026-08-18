from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import pytest

from wolfxl import Workbook, load_workbook
from wolfxl._cell import _UNSET
from wolfxl._worksheet_writer_flush import (
    _flush_format_cells,
    _partition_dirty_cells,
    _plan_batch_value_grids,
    _plan_dirty_value_grids,
    _write_dirty_values,
    _write_batch_values,
)
from wolfxl._worksheet_flush import flush_compat_properties
from wolfxl.comments import Comment
from wolfxl.styles import Font


def test_plain_cell_assignments_remain_batchable() -> None:
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value=1)
    ws.cell(row=1, column=2, value="plain")
    ws.cell(row=1, column=3, value=True)
    ws.cell(row=1, column=4, value="=A1*2")
    ws.cell(row=1, column=5, value="#N/A")

    batch_values, individual_values, format_cells = _partition_dirty_cells(ws)

    assert sorted((row, col) for row, col, _value in batch_values) == [
        (1, 1),
        (1, 2),
        (1, 3),
        (1, 4),
        (1, 5),
    ]
    assert individual_values == []
    assert format_cells == []
    assert all(cell._explicit_data_type is _UNSET for cell in ws._cells.values())  # noqa: SLF001


def test_plain_cell_assignments_still_infer_openpyxl_data_types() -> None:
    wb = Workbook()
    ws = wb.active

    assert ws.cell(row=1, column=1, value=1).data_type == "n"
    assert ws.cell(row=1, column=2, value="plain").data_type == "s"
    assert ws.cell(row=1, column=3, value=True).data_type == "b"
    assert ws.cell(row=1, column=4, value="=A1*2").data_type == "f"
    assert ws.cell(row=1, column=5, value="#N/A").data_type == "e"
    assert ws.cell(row=1, column=6, value=None).data_type == "n"
    assert ws.cell(row=1, column=7, value="=").data_type == "s"
    assert ws.cell(row=1, column=8, value=Decimal("3.14")).data_type == "n"
    assert ws.cell(row=1, column=9, value=timedelta(days=1, seconds=2)).data_type == "d"


def test_scalar_fast_path_updates_worksheet_bounds() -> None:
    wb = Workbook()
    ws = wb.active

    ws.cell(row=4, column=7, value="x")

    assert ws.max_row == 4
    assert ws.max_column == 7
    assert ws._next_append_row == 5  # noqa: SLF001


def test_new_plain_scalar_cell_fast_path_initializes_full_cell_state() -> None:
    wb = Workbook()
    ws = wb.active

    cell = ws.cell(row=2, column=3, value="plain")

    assert cell.value == "plain"
    assert cell.coordinate == "C2"
    assert cell._value_dirty is True  # noqa: SLF001
    assert cell._format_dirty is False  # noqa: SLF001
    assert cell._explicit_data_type is _UNSET  # noqa: SLF001
    assert (2, 3) not in ws._dirty  # noqa: SLF001
    assert ws._dirty_values[(2, 3)] == "plain"  # noqa: SLF001


def test_new_plain_scalar_cell_updates_pending_overlay_and_bounds() -> None:
    wb = Workbook()
    ws = wb.active

    ws.cell(row=10, column=5, value="added")

    assert ws._collect_pending_overlay()[(10, 5)] == "added"  # noqa: SLF001
    assert ws._pending_writes_bounds() == (10, 5, 10, 5)  # noqa: SLF001


def test_modify_plain_scalar_setitem_uses_compact_dirty_values(tmp_path) -> None:
    path = tmp_path / "modify-fast-path.xlsx"
    wb = Workbook()
    wb.save(path)
    wb.close()

    loaded = load_workbook(path, modify=True)
    ws = loaded.active
    assert ws._merged_ranges_loaded is True  # noqa: SLF001

    ws["C2"] = "plain"

    assert ws._plain_cell_fast_path is True  # noqa: SLF001
    assert (2, 3) not in ws._dirty  # noqa: SLF001
    assert ws._dirty_values[(2, 3)] == "plain"  # noqa: SLF001

    loaded.save(path)
    loaded.close()

    verified = openpyxl.load_workbook(path)
    assert verified.active["C2"].value == "plain"
    verified.close()


def test_modify_plain_scalar_setitem_caches_merged_range_reads() -> None:
    class Reader:
        def __init__(self) -> None:
            self.calls = 0

        def read_merged_ranges(self, _title: str) -> list[str]:
            self.calls += 1
            return []

    wb = Workbook()
    ws = wb.active
    reader = Reader()
    wb._rust_reader = reader  # noqa: SLF001
    wb._rust_patcher = object()  # noqa: SLF001
    ws._plain_cell_fast_path = True  # noqa: SLF001

    ws["B2"] = "changed"
    ws["C3"] = 12345

    assert reader.calls == 1
    assert ws._dirty_values[(2, 2)] == "changed"  # noqa: SLF001
    assert ws._dirty_values[(3, 3)] == 12345  # noqa: SLF001


def test_append_disables_plain_cell_fast_path() -> None:
    wb = Workbook()
    ws = wb.active

    assert ws._plain_cell_fast_path is True  # noqa: SLF001
    ws.append(["buffered"])

    assert ws._plain_cell_fast_path is False  # noqa: SLF001


def test_merge_disables_plain_cell_fast_path() -> None:
    wb = Workbook()
    ws = wb.active

    assert ws._plain_cell_fast_path is True  # noqa: SLF001
    ws.merge_cells("A1:B2")

    assert ws._plain_cell_fast_path is False  # noqa: SLF001


def test_user_explicit_data_type_still_uses_type_preserving_path() -> None:
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="=literal")
    cell.data_type = "s"

    batch_values, individual_values, format_cells = _partition_dirty_cells(ws)

    assert batch_values == []
    assert [(row, col) for row, col, _cell in individual_values] == [(1, 1)]
    assert format_cells == []


def test_date_assignment_stays_on_format_preserving_path() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=date(2026, 5, 20))

    batch_values, individual_values, format_cells = _partition_dirty_cells(ws)

    assert batch_values == []
    assert [(row, col) for row, col, _cell in individual_values] == [(1, 1)]
    assert [(row, col) for row, col, _cell in format_cells] == [(1, 1)]


def test_plain_values_with_styles_stay_batchable() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="A").font = Font(bold=True)
    ws.cell(row=1, column=2, value="B").font = Font(bold=True)

    batch_values, individual_values, format_cells = _partition_dirty_cells(ws)

    assert sorted((row, col, value) for row, col, value in batch_values) == [
        (1, 1, "A"),
        (1, 2, "B"),
    ]
    assert individual_values == []
    assert ws._dirty == set()  # noqa: SLF001
    assert sorted((row, col) for row, col, _cell in format_cells) == [
        (1, 1),
        (1, 2),
    ]


def test_style_only_cells_still_mark_dirty_for_bounds() -> None:
    wb = Workbook()
    ws = wb.active

    ws.cell(row=4, column=2).font = Font(bold=True)

    assert ws._dirty == {(4, 2)}  # noqa: SLF001
    assert ws._pending_writes_bounds() == (4, 2, 4, 2)  # noqa: SLF001


def test_mixed_plain_and_styled_values_all_stay_batchable() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="styled").font = Font(bold=True)
    ws.cell(row=1, column=2, value="plain")

    batch_values, individual_values, format_cells = _partition_dirty_cells(ws)

    assert sorted((row, col, value) for row, col, value in batch_values) == [
        (1, 1, "styled"),
        (1, 2, "plain"),
    ]
    assert individual_values == []
    assert sorted((row, col) for row, col, _cell in format_cells) == [(1, 1)]


def test_comments_and_styles_are_not_lost_when_values_batch(tmp_path) -> None:
    path = tmp_path / "styled-commented.xlsx"
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="flag")
    cell.font = Font(bold=True)
    cell.comment = Comment("review this", "WolfXL")
    wb.save(str(path))
    wb.close()

    loaded = openpyxl.load_workbook(path)
    sheet = loaded.active
    assert sheet["A1"].value == "flag"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].comment is not None
    assert sheet["A1"].comment.text == "review this"
    loaded.close()


def test_mixed_styled_and_plain_values_survive_save(tmp_path) -> None:
    path = tmp_path / "mixed-styled-plain.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="styled").font = Font(bold=True)
    ws.cell(row=1, column=2, value="plain")
    wb.save(str(path))
    wb.close()

    loaded = openpyxl.load_workbook(path)
    sheet = loaded.active
    assert sheet["A1"].value == "styled"
    assert sheet["A1"].font.bold is True
    assert sheet["B1"].value == "plain"
    loaded.close()


def test_plan_batch_value_grids_keeps_dense_rectangle_together() -> None:
    grids = _plan_batch_value_grids(
        [
            (2, 2, "B2"),
            (1, 1, "A1"),
            (1, 2, "B1"),
            (2, 1, "A2"),
        ]
    )

    assert grids == [(1, 1, [["A1", "B1"], ["A2", "B2"]])]


def test_plan_batch_value_grids_splits_sparse_cells() -> None:
    grids = _plan_batch_value_grids(
        [
            (1, 1, "A1"),
            (1, 100, "CV1"),
            (2, 1, "A2"),
            (2, 100, "CV2"),
        ]
    )

    assert grids == [
        (1, 1, [["A1"], ["A2"]]),
        (1, 100, [["CV1"], ["CV2"]]),
    ]


def test_write_batch_values_uses_compact_sparse_calls() -> None:
    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_values(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.calls.append((sheet, start, grid))

    wb = Workbook()
    ws = wb.active
    writer = Writer()

    _write_batch_values(
        ws,
        writer,
        [
            (1, 1, "A1"),
            (1, 50, "AX1"),
            (2, 1, "A2"),
            (2, 50, "AX2"),
        ],
    )

    assert writer.calls == [
        ("Sheet", "A1", [["A1"], ["A2"]]),
        ("Sheet", "AX1", [["AX1"], ["AX2"]]),
    ]


def test_plan_dirty_value_grids_keeps_dense_map_together() -> None:
    grids = _plan_dirty_value_grids(
        {
            (1, 1): "A1",
            (1, 2): "B1",
            (2, 1): "A2",
            (2, 2): "B2",
        }
    )

    assert grids == [(1, 1, [["A1", "B1"], ["A2", "B2"]])]


def test_write_dirty_values_uses_compact_sparse_calls() -> None:
    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_values(self, sheet: str, start: str, grid: list[list[object]]) -> None:
            self.calls.append((sheet, start, grid))

    wb = Workbook()
    ws = wb.active
    writer = Writer()

    _write_dirty_values(
        ws,
        writer,
        {
            (1, 1): "A1",
            (1, 50): "AX1",
            (2, 1): "A2",
            (2, 50): "AX2",
        },
    )

    assert writer.calls == [
        ("Sheet", "A1", [["A1"], ["A2"]]),
        ("Sheet", "AX1", [["AX1"], ["AX2"]]),
    ]


def test_plain_sheet_skips_compat_metadata_flush() -> None:
    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, tuple[object, ...]]] = []

        def __getattr__(self, name: str):
            def record(*args: object) -> None:
                self.calls.append((name, args))

            return record

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "plain"
    writer = Writer()

    flush_compat_properties(ws, writer)

    assert writer.calls == []


def test_sheet_metadata_still_flushes_when_present() -> None:
    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, tuple[object, ...]]] = []

        def set_freeze_panes(self, *args: object) -> None:
            self.calls.append(("set_freeze_panes", args))

        def set_row_dimension(self, *args: object) -> None:
            self.calls.append(("set_row_dimension", args))

    wb = Workbook()
    ws = wb.active
    ws.freeze_panes = "B2"
    ws.row_dimensions[3].height = 24
    writer = Writer()

    flush_compat_properties(ws, writer)

    assert writer.calls == [
        ("set_freeze_panes", ("Sheet", {"mode": "freeze", "top_left_cell": "B2"})),
        (
            "set_row_dimension",
            (
                "Sheet",
                3,
                {"height": 24.0, "hidden": False, "outline_level": 0},
            ),
        ),
    ]


def test_format_flush_reuses_repeated_style_payloads() -> None:
    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_formats(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.calls.append((sheet, start, grid))

    wb = Workbook()
    ws = wb.active
    cell_a = ws.cell(row=1, column=1, value="A")
    cell_b = ws.cell(row=1, column=2, value="B")
    cell_a.font = Font(bold=True)
    cell_b.font = Font(bold=True)
    writer = Writer()
    font_payload_calls: list[Font] = []

    def font_payload(font: Font) -> dict[str, object]:
        font_payload_calls.append(font)
        return {"bold": bool(font.bold)}

    def unused_converter(_value: object) -> dict[str, object]:
        raise AssertionError("converter should not be called")

    _flush_format_cells(
        ws,
        writer,
        [(1, 1, cell_a), (1, 2, cell_b)],
        font_payload,
        unused_converter,
        unused_converter,
        unused_converter,
        unused_converter,
    )

    assert len(font_payload_calls) == 1
    assert writer.calls == [
        ("Sheet", "A1", [[{"bold": True}, {"bold": True}]])
    ]


def test_format_flush_cache_avoids_full_font_hash(monkeypatch: pytest.MonkeyPatch) -> None:
    """Repeated font styles should cache by cheap fields, not Font.__hash__."""

    class Writer:
        def __init__(self) -> None:
            self.calls: list[tuple[str, str, list[list[object]]]] = []

        def write_sheet_formats(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.calls.append((sheet, start, grid))

    def fail_hash(self: Font) -> int:
        raise AssertionError("style flush should not hash full Font objects")

    monkeypatch.setattr(Font, "__hash__", fail_hash)

    wb = Workbook()
    ws = wb.active
    cell_a = ws.cell(row=1, column=1, value="A")
    cell_b = ws.cell(row=1, column=2, value="B")
    cell_a.font = Font(bold=True)
    cell_b.font = Font(bold=True)
    writer = Writer()
    font_payload_calls: list[Font] = []

    def font_payload(font: Font) -> dict[str, object]:
        font_payload_calls.append(font)
        return {"bold": bool(font.bold)}

    def unused_converter(_value: object) -> dict[str, object]:
        raise AssertionError("converter should not be called")

    _flush_format_cells(
        ws,
        writer,
        [(1, 1, cell_a), (1, 2, cell_b)],
        font_payload,
        unused_converter,
        unused_converter,
        unused_converter,
        unused_converter,
    )

    assert len(font_payload_calls) == 1
    assert writer.calls == [
        ("Sheet", "A1", [[{"bold": True}, {"bold": True}]])
    ]


def test_format_flush_uses_style_id_grid_for_repeated_single_styles() -> None:
    class Writer:
        def __init__(self) -> None:
            self.format_calls: list[tuple[str, str, list[list[object]]]] = []
            self.style_id_calls: list[tuple[str, str, list[list[object]]]] = []
            self.interned: list[dict[str, object]] = []

        def intern_format(self, payload: dict[str, object]) -> int:
            self.interned.append(payload)
            return len(self.interned)

        def write_sheet_formats(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.format_calls.append((sheet, start, grid))

        def write_sheet_style_ids(
            self,
            sheet: str,
            start: str,
            grid: list[list[object]],
        ) -> None:
            self.style_id_calls.append((sheet, start, grid))

    wb = Workbook()
    ws = wb.active
    cell_a = ws.cell(row=1, column=1, value="A")
    cell_b = ws.cell(row=1, column=2, value="B")
    cell_a.font = Font(bold=True)
    cell_b.font = Font(bold=True)
    writer = Writer()
    font_payload_calls: list[Font] = []

    def font_payload(font: Font) -> dict[str, object]:
        font_payload_calls.append(font)
        return {"bold": bool(font.bold)}

    def unused_converter(_value: object) -> dict[str, object]:
        raise AssertionError("converter should not be called")

    _flush_format_cells(
        ws,
        writer,
        [(1, 1, cell_a), (1, 2, cell_b)],
        font_payload,
        unused_converter,
        unused_converter,
        unused_converter,
        unused_converter,
    )

    assert len(font_payload_calls) == 1
    assert writer.interned == [{"bold": True}]
    assert writer.style_id_calls == [("Sheet", "A1", [[1, 1]])]
    assert writer.format_calls == []
