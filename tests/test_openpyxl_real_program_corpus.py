"""Local real-program-style openpyxl snippets run through WolfXL's alias.

These snippets are distilled from openpyxl usage found in neighboring local
projects such as ExcelBench, SynthGL, and finance demo generators. They are not
API micro-smokes: each one keeps ordinary openpyxl imports after the one-line
runtime alias and exercises a small end-to-end workflow.
"""

from __future__ import annotations

from collections.abc import Iterator
from contextlib import contextmanager
import sys
import textwrap


@contextmanager
def _isolated_openpyxl_imports() -> Iterator[None]:
    saved_modules = {
        name: module
        for name, module in sys.modules.items()
        if name == "openpyxl" or name.startswith("openpyxl.")
    }
    saved_meta_path = list(sys.meta_path)
    for name in list(saved_modules):
        del sys.modules[name]
    try:
        yield
    finally:
        for name in list(sys.modules):
            if name == "openpyxl" or name.startswith("openpyxl."):
                del sys.modules[name]
        sys.modules.update(saved_modules)
        sys.meta_path[:] = saved_meta_path


def _run_snippet(source: str, tmp_path) -> dict[str, object]:
    namespace = {"tmp_path": tmp_path, "result": None}
    with _isolated_openpyxl_imports():
        exec(textwrap.dedent(source), namespace)  # noqa: S102 - intentional corpus exec
    result = namespace["result"]
    assert isinstance(result, dict)
    return result


def test_excelbench_snapshot_style_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.table import Table, TableStyleInfo
        from openpyxl.workbook.defined_name import DefinedName

        path = tmp_path / "snapshot-style.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Metric", "Value", "Status"])
        ws.append(["Revenue", 4200000, "Open"])
        ws["A1"].font = Font(bold=True)
        ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
        ws["C2"].comment = Comment("Reviewed", "ExcelBench")
        ws["C2"].hyperlink = "https://github.com/SynthGL/wolfxl"
        ws.merge_cells("A4:C4")
        ws["A4"] = "Summary"
        ws.freeze_panes = "B2"
        dv = DataValidation(type="list", formula1='"Open,Closed"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add("C2:C10")
        table = Table(displayName="CompatTable", ref="A1:C2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
        wb.defined_names.add(DefinedName("MetricCell", attr_text="'Sheet1'!$A$2"))
        wb.save(path)

        loaded = openpyxl.load_workbook(path, data_only=False)
        sheet = loaded["Sheet1"]
        result = {
            "sheets": loaded.sheetnames,
            "value": sheet["B2"].value,
            "comment": sheet["C2"].comment.text,
            "hyperlink": sheet["C2"].hyperlink.target,
            "merge": [str(rng) for rng in sheet.merged_cells.ranges],
            "freeze": str(sheet.freeze_panes),
            "tables": sorted(sheet.tables.keys()),
            "validations": [str(dv.sqref) for dv in sheet.data_validations.dataValidation],
            "defined_names": sorted(loaded.defined_names.keys()),
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "sheets": ["Sheet1"],
        "value": 4200000,
        "comment": "Reviewed",
        "hyperlink": "https://github.com/SynthGL/wolfxl",
        "merge": ["A4:C4"],
        "freeze": "B2",
        "tables": ["CompatTable"],
        "validations": ["C2:C10"],
        "defined_names": ["MetricCell"],
    }


def test_finance_generator_style_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers

        path = tmp_path / "finance-generator-style.xlsx"
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        money_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        thin_border = Border(bottom=Side(style="thin"))
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income Statement"
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18
        ws.cell(row=1, column=1, value="Acme SaaS Corp").font = title_font
        ws.cell(row=2, column=1, value="Income Statement").font = header_font
        for col in range(1, 3):
            cell = ws.cell(row=5, column=col)
            cell.fill = gray_fill
            cell.alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=1, value="Account")
        ws.cell(row=5, column=2, value="FY2025")
        ws.cell(row=6, column=1, value="Revenue")
        amount = ws.cell(row=6, column=2, value=4_200_000)
        amount.number_format = money_format
        amount.border = thin_border
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        sheet = loaded["Income Statement"]
        result = {
            "title": sheet["A1"].value,
            "amount": sheet["B6"].value,
            "number_format": sheet["B6"].number_format,
            "width": sheet.column_dimensions["A"].width,
            "header_center": sheet["A5"].alignment.horizontal,
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "title": "Acme SaaS Corp",
        "amount": 4_200_000,
        "number_format": "#,##0.00",
        "width": 40.0,
        "header_center": "center",
    }


def test_synthgl_metadata_reader_style_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.utils import column_index_from_string

        path = tmp_path / "metadata-reader-style.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].hidden = True
        ws.row_dimensions[3].height = 30
        ws.row_dimensions[4].hidden = True
        ws.merge_cells("A1:C1")
        ws["A1"] = "Header"
        ws["B3"] = 123
        wb.save(path)

        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
        worksheet = workbook.worksheets[0]
        hidden_cols = []
        column_widths = {}
        for key, dimension in worksheet.column_dimensions.items():
            col_index = column_index_from_string(str(key)) - 1
            if dimension.hidden:
                hidden_cols.append(col_index)
            if dimension.width is not None:
                column_widths[col_index] = float(dimension.width)
        hidden_rows = []
        row_heights = {}
        for row_key, dimension in worksheet.row_dimensions.items():
            row_index = int(row_key) - 1
            if dimension.hidden:
                hidden_rows.append(row_index)
            if dimension.height is not None:
                row_heights[row_index] = float(dimension.height)
        result = {
            "sheets": list(workbook.sheetnames),
            "hidden_cols": hidden_cols,
            "column_widths": column_widths,
            "hidden_rows": hidden_rows,
            "row_heights": row_heights,
            "merged": [(r.min_row - 1, r.min_col - 1, r.max_row, r.max_col) for r in worksheet.merged_cells.ranges],
            "value": worksheet["B3"].value,
        }
        workbook.close()
        """,
        tmp_path,
    )

    assert result == {
        "sheets": ["Data"],
        "hidden_cols": [2],
        "column_widths": {1: 22.0, 2: 13.0},
        "hidden_rows": [3],
        "row_heights": {2: 30.0},
        "merged": [(0, 0, 1, 3)],
        "value": 123,
    }


def test_file_like_financial_parser_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        from io import BytesIO
        import openpyxl

        source = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Input"
        ws.append(["Account", "FY2025"])
        ws.append(["Revenue", 4200000])
        ws.append([None, None])
        ws.append(["Gross Profit", 2700000])
        wb.save(source)
        position_after_save = source.tell()
        file_bytes = source.getvalue()

        parsed = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = parsed["Input"]
        rows_data = []
        headers = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            cleaned = []
            for cell in row:
                if cell is None:
                    cleaned.append("")
                elif isinstance(cell, (int, float)):
                    cleaned.append(cell)
                else:
                    cleaned.append(str(cell).strip())
            if all(c == "" or c == 0 for c in cleaned):
                continue
            if row_idx == 0:
                headers = [str(c) for c in cleaned]
            rows_data.append(cleaned)
        parsed.close()

        result = {
            "position_after_save": position_after_save,
            "bytes_len": len(file_bytes),
            "headers": headers,
            "rows": rows_data,
        }
        """,
        tmp_path,
    )

    assert result["position_after_save"] == result["bytes_len"]
    assert result["headers"] == ["Account", "FY2025"]
    assert result["rows"] == [
        ["Account", "FY2025"],
        ["Revenue", 4_200_000],
        ["Gross Profit", 2_700_000],
    ]


def test_read_only_formula_preflight_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        from datetime import date
        import openpyxl

        path = tmp_path / "read-only-formula-preflight.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Preflight"
        ws.append(["Label", "Amount", "Date", "Formula", "Error"])
        ws.append(["Revenue", 4200000, date(2026, 5, 12), "=B2*2", "#N/A"])
        ws["C2"].number_format = "yyyy-mm-dd"
        wb.save(path)

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        worksheet = workbook["Preflight"]
        cells = list(worksheet.iter_rows(min_row=2, max_row=2, values_only=False))[0]
        values = list(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        result = {
            "values": [
                value.isoformat() if hasattr(value, "isoformat") else value
                for value in values
            ],
            "data_types": [cell.data_type for cell in cells],
            "coordinates": [cell.coordinate for cell in cells],
        }
        workbook.close()
        """,
        tmp_path,
    )

    assert result == {
        "values": ["Revenue", 4_200_000, "2026-05-12T00:00:00", "=B2*2", "#N/A"],
        "data_types": ["s", "n", "d", "f", "e"],
        "coordinates": ["A2", "B2", "C2", "D2", "E2"],
    }


def test_append_multi_sheet_generator_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        from datetime import date
        from openpyxl import Workbook

        path = tmp_path / "append-generator-style.xlsx"
        wb = Workbook()
        wb.remove(wb.active)

        pnl = wb.create_sheet("P&L")
        pnl.append(["Account", "Jan 2024", "Q1 Total"])
        pnl.append(["Revenue", 692000, 2253300])
        pnl.append(["Operating Margin %", "28.8%", "29.9%"])
        pnl.column_dimensions[pnl.cell(row=1, column=1).column_letter].width = 28

        bs = wb.create_sheet("Balance Sheet")
        bs.append(["Account", date(2024, 3, 31), date(2023, 12, 31)])
        bs.append(["Cash", 1842500, 1120000])
        bs.cell(row=1, column=2).number_format = "yyyy-mm-dd"
        bs.cell(row=1, column=3).number_format = "yyyy-mm-dd"

        wb.save(path)

        loaded = __import__("openpyxl").load_workbook(path, data_only=True)
        result = {
            "sheets": loaded.sheetnames,
            "pnl_rows": list(loaded["P&L"].iter_rows(values_only=True)),
            "pnl_width": loaded["P&L"].column_dimensions["A"].width,
            "bs_header": loaded["Balance Sheet"]["B1"].value.isoformat(),
            "bs_format": loaded["Balance Sheet"]["B1"].number_format,
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "sheets": ["P&L", "Balance Sheet"],
        "pnl_rows": [
            ("Account", "Jan 2024", "Q1 Total"),
            ("Revenue", 692_000, 2_253_300),
            ("Operating Margin %", "28.8%", "29.9%"),
        ],
        "pnl_width": 28.0,
        "bs_header": "2024-03-31T00:00:00",
        "bs_format": "yyyy-mm-dd",
    }


def test_keyword_merge_archetype_generator_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        path = tmp_path / "keyword-merge-archetype.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unadj_to_Adj_PL"
        ws.cell(row=1, column=1, value="")
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
        header = ws.cell(row=1, column=2, value="Jan-24")
        header.font = Font(bold=True)
        header.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header.alignment = Alignment(horizontal="center")
        for offset, label in enumerate(["Unadjusted", "Adj_A", "Adj_B", "Adj_C", "Adj_D", "Adjusted"], start=2):
            ws.cell(row=2, column=offset, value=label)
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        sheet = loaded["Unadj_to_Adj_PL"]
        result = {
            "merged": [str(rng) for rng in sheet.merged_cells.ranges],
            "header": sheet["B1"].value,
            "alignment": sheet["B1"].alignment.horizontal,
            "labels": [sheet.cell(row=2, column=col).value for col in range(2, 8)],
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "merged": ["B1:G1"],
        "header": "Jan-24",
        "alignment": "center",
        "labels": ["Unadjusted", "Adj_A", "Adj_B", "Adj_C", "Adj_D", "Adjusted"],
    }


def test_excelbench_feature_import_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.filters import AutoFilter
        from openpyxl.worksheet.hyperlink import Hyperlink
        from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
        from openpyxl.worksheet.views import Pane

        path = tmp_path / "excelbench-feature-import-style.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Features"
        ws.append(["Name", "Value"])
        for row in range(2, 8):
            ws.cell(row=row, column=1, value=f"Row {row}")
            ws.cell(row=row, column=2, value=row)
        ws["A2"].hyperlink = Hyperlink(ref="A2", target="https://example.com")
        ws.conditional_formatting.add(
            "B2:B7",
            CellIsRule(
                operator="greaterThan",
                formula=["4"],
                fill=PatternFill(fill_type="solid", fgColor="FFFF0000"),
            ),
        )
        ws.conditional_formatting.add("B2:B7", FormulaRule(formula=["B2=2"]))
        ws.conditional_formatting.add(
            "B2:B7",
            ColorScaleRule(
                start_type="min",
                start_color="FFFFFF",
                end_type="max",
                end_color="00FF00",
            ),
        )
        table = Table(displayName="FeatureTable", ref="A1:B7")
        table.tableColumns = [TableColumn(id=1, name="Name"), TableColumn(id=2, name="Value")]
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.autoFilter = AutoFilter(ref="A1:B7")
        ws.add_table(table)
        ws.sheet_view.pane = Pane(
            xSplit=1,
            ySplit=1,
            topLeftCell="B2",
            activePane="bottomRight",
            state="frozen",
        )
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        sheet = loaded["Features"]
        result = {
            "hyperlink": sheet["A2"].hyperlink.target,
            "tables": sorted(sheet.tables.keys()),
            "conditional_formatting_count": len(list(sheet.conditional_formatting)),
            "freeze": str(sheet.freeze_panes),
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "hyperlink": "https://example.com",
        "tables": ["FeatureTable"],
        "conditional_formatting_count": 1,
        "freeze": "B2",
    }


def test_excelbench_conditional_formatting_dxf_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
        from openpyxl.styles import Font, PatternFill
        from openpyxl.styles.differential import DifferentialStyle

        path = tmp_path / "excelbench-conditional-formatting-dxf.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Features"
        ws.append(["Item", "Amount", "Flag"])
        for row, amount in enumerate([10, 20, 30, 40], start=2):
            ws.cell(row=row, column=1, value=f"Item {row}")
            ws.cell(row=row, column=2, value=amount)
            ws.cell(row=row, column=3, value="Y" if amount >= 30 else "N")

        dxf = DifferentialStyle(
            font=Font(color="FFFF0000"),
            fill=PatternFill(fill_type="solid", bgColor="FFFF0000"),
        )
        cell_rule = CellIsRule(operator="greaterThan", formula=["25"], stopIfTrue=True)
        cell_rule.dxf = dxf
        ws.conditional_formatting.add("B2:B5", cell_rule)
        ws.conditional_formatting.add(
            "C2:C5",
            FormulaRule(formula=['C2="Y"'], fill=PatternFill(fill_type="solid", fgColor="FFCCFFCC")),
        )
        ws.conditional_formatting.add(
            "B2:B5",
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=40,
                color="FF638EC6",
                showValue=True,
            ),
        )
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        sheet = loaded["Features"]
        rule_types = []
        for cf in sheet.conditional_formatting:
            for rule in cf.rules:
                rule_types.append(rule.type)
        result = {
            "conditional_formatting_count": len(rule_types),
            "rule_types": sorted(rule_types),
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "conditional_formatting_count": 3,
        "rule_types": ["cellIs", "dataBar", "expression"],
    }


def test_excelbench_style_color_object_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.styles import Border, Color, Font, PatternFill, Side
        from openpyxl.styles.colors import Color as ColorsColor

        path = tmp_path / "excelbench-style-color-object.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Styled"
        ws["A1"] = "styled"
        ws["A1"].font = Font(color=Color(rgb="FFFF0000"), bold=True)
        ws["A1"].fill = PatternFill(
            start_color=ColorsColor(rgb="FF00FF00"),
            end_color=ColorsColor(rgb="FF00FF00"),
            fill_type="solid",
        )
        ws["A1"].border = Border(left=Side(style="thin", color=Color(rgb="FF0000FF")))
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        cell = loaded["Styled"]["A1"]
        result = {
            "font_rgb": cell.font.color.rgb,
            "fill_rgb": cell.fill.fgColor.rgb,
            "border_rgb": cell.border.left.color.rgb,
        }
        loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "font_rgb": "FFFF0000",
        "fill_rgb": "FF00FF00",
        "border_rgb": "FF0000FF",
    }


def test_synthgl_workbook_preview_formula_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.cell.cell import Cell
        from openpyxl.utils.cell import get_column_letter
        from openpyxl.workbook.defined_name import DefinedName
        from openpyxl.worksheet.worksheet import Worksheet

        path = tmp_path / "synthgl-preview-style.xlsx"
        wb = openpyxl.Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary["A1"] = "Revenue"
        summary["B1"] = "=Data!A1"
        data = wb.create_sheet("Data")
        data["A1"] = 123
        wb.defined_names.add(DefinedName("RevenueCell", attr_text="'Data'!$A$1"))
        wb.save(path)

        raw_workbook = openpyxl.load_workbook(path, data_only=False)
        cached_workbook = openpyxl.load_workbook(path, data_only=True)
        try:
            raw_sheet = raw_workbook["Summary"]
            cached_sheet = cached_workbook["Summary"]
            formulas = []
            formula_values = {}
            grid = []
            for row_idx in range(1, raw_sheet.max_row + 1):
                row_values = []
                for col_idx in range(1, raw_sheet.max_column + 1):
                    raw_value = raw_sheet.cell(row=row_idx, column=col_idx).value
                    cached_value = cached_sheet.cell(row=row_idx, column=col_idx).value
                    display_value = raw_value
                    if isinstance(raw_value, str) and raw_value.startswith("="):
                        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                        formulas.append({"sheet": "Summary", "cell_ref": cell_ref, "formula": raw_value})
                        if cached_value is not None:
                            formula_values[f"Summary!{cell_ref}"] = cached_value
                            display_value = cached_value
                    row_values.append(display_value)
                grid.append(row_values)
            result = {
                "cell_type_import": isinstance(raw_sheet["A1"], Cell),
                "worksheet_type_import": isinstance(raw_sheet, Worksheet),
                "grid": grid,
                "formulas": formulas,
                "formula_values": formula_values,
                "defined_names": {
                    key: getattr(value, "attr_text", str(value))
                    for key, value in raw_workbook.defined_names.items()
                },
            }
        finally:
            raw_workbook.close()
            cached_workbook.close()
        """,
        tmp_path,
    )

    assert result == {
        "cell_type_import": True,
        "worksheet_type_import": True,
        "grid": [["Revenue", "=Data!A1"]],
        "formulas": [{"sheet": "Summary", "cell_ref": "B1", "formula": "=Data!A1"}],
        "formula_values": {},
        "defined_names": {"RevenueCell": "'Data'!$A$1"},
    }


def test_third_party_workbook_iteration_program_runs_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visible"
        hidden = wb.create_sheet("Hidden")
        hidden.sheet_state = "hidden"

        result = {
            "iterated": [
                {"title": sheet.title, "state": sheet.sheet_state}
                for sheet_name, sheet in zip(wb.sheetnames, wb)
            ]
        }
        """,
        tmp_path,
    )

    assert result == {
        "iterated": [
            {"title": "Visible", "state": "visible"},
            {"title": "Hidden", "state": "hidden"},
        ]
    }


def test_explicit_string_data_type_preserves_error_and_formula_text(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl

        path = tmp_path / "explicit-string-type.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "#N/A"
        ws["A1"].data_type = "s"
        ws["A2"] = "=1+2"
        ws["A2"].data_type = "s"
        wb.save(path)

        loaded = openpyxl.load_workbook(path, data_only=False)
        try:
            result = {
                "values": [loaded.active["A1"].value, loaded.active["A2"].value],
                "types": [loaded.active["A1"].data_type, loaded.active["A2"].data_type],
            }
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "values": ["#N/A", "=1+2"],
        "types": ["s", "s"],
    }


def test_data_validation_messages_round_trip_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation

        path = tmp_path / "data-validation-messages.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        dv = DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="10",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            errorTitle="Invalid",
            error="Use 1-10",
            promptTitle="Hint",
            prompt="Enter a whole number",
        )
        ws.add_data_validation(dv)
        dv.add("A1:A3")
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        try:
            loaded_dv = loaded.active.data_validations.dataValidation[0]
            result = {
                "formula1": loaded_dv.formula1,
                "formula2": loaded_dv.formula2,
                "allowBlank": loaded_dv.allowBlank,
                "showErrorMessage": loaded_dv.showErrorMessage,
                "showInputMessage": loaded_dv.showInputMessage,
                "errorTitle": loaded_dv.errorTitle,
                "error": loaded_dv.error,
                "promptTitle": loaded_dv.promptTitle,
                "prompt": loaded_dv.prompt,
            }
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "formula1": "1",
        "formula2": "10",
        "allowBlank": True,
        "showErrorMessage": True,
        "showInputMessage": True,
        "errorTitle": "Invalid",
        "error": "Use 1-10",
        "promptTitle": "Hint",
        "prompt": "Enter a whole number",
    }


def test_external_hyperlink_location_round_trips_through_alias(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.worksheet.hyperlink import Hyperlink

        path = tmp_path / "hyperlink-location.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Docs"
        ws["A1"].hyperlink = Hyperlink(
            ref="A1",
            target="https://example.com/page",
            location="section-2",
            tooltip="Jump",
        )
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        try:
            link = loaded.active["A1"].hyperlink
            result = {
                "target": link.target,
                "location": link.location,
                "tooltip": link.tooltip,
                "combined": link.target + "#" + link.location,
            }
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "target": "https://example.com/page",
        "location": "section-2",
        "tooltip": "Jump",
        "combined": "https://example.com/page#section-2",
    }


def test_conditional_formatting_private_cf_rules_matches_openpyxl_shape(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill

        path = tmp_path / "private-cf-rules.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = 12
        ws.conditional_formatting.add(
            "A1:A5",
            FormulaRule(
                formula=["A1>10"],
                fill=PatternFill(fill_type="solid", fgColor="FFFF00"),
                stopIfTrue=True,
            ),
        )
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        try:
            cf_rules = loaded.active.conditional_formatting._cf_rules
            result = {
                "rules": [
                    {
                        "range": str(getattr(sqref, "sqref", sqref)),
                        "type": rules[0].type,
                        "formula": rules[0].formula[0],
                        "stop": rules[0].stopIfTrue,
                    }
                    for sqref, rules in cf_rules.items()
                ]
            }
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "rules": [
            {
                "range": "A1:A5",
                "type": "expression",
                "formula": "A1>10",
                "stop": True,
            }
        ]
    }


def test_write_only_sheet_title_rename_updates_streaming_backend(tmp_path) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl

        path = tmp_path / "write-only-renamed-sheet.xlsx"
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.title = "S1"
        ws.append([True, "#REF!", ""])
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        try:
            result = {
                "sheets": loaded.sheetnames,
                "values": list(loaded["S1"].iter_rows(values_only=True)),
            }
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {
        "sheets": ["S1"],
        "values": [(True, "#REF!", "")],
    }


def test_column_dimension_dict_mutation_is_visible_to_adapter_style_reads(
    tmp_path,
) -> None:
    result = _run_snippet(
        """
        import wolfxl; wolfxl.install_as_openpyxl()

        import openpyxl

        path = tmp_path / "dimension-dict-mutation.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S1"
        ws["A1"] = "value"
        wb.save(path)

        loaded = openpyxl.load_workbook(path)
        try:
            dim = loaded["S1"].column_dimensions["A"]
            dim.__dict__["width"] = "invalid"
            same_dim = loaded["S1"].column_dimensions["A"]
            try:
                float(same_dim.width)
            except (TypeError, ValueError):
                result = {"same_object": dim is same_dim, "width": same_dim.width}
            else:
                result = {"same_object": dim is same_dim, "width": "numeric"}
        finally:
            loaded.close()
        """,
        tmp_path,
    )

    assert result == {"same_object": True, "width": "invalid"}
