from __future__ import annotations

import datetime as dt
from pathlib import Path
import zipfile

import openpyxl
from openpyxl.utils.datetime import CALENDAR_MAC_1904

import wolfxl
from wolfxl.utils.datetime import CALENDAR_MAC_1904 as WOLFXL_CALENDAR_MAC_1904


def test_openpyxl_builtin_number_format_catalog_parity() -> None:
    import openpyxl.styles.numbers as openpyxl_numbers
    import wolfxl.styles.numbers as wolfxl_numbers

    assert wolfxl_numbers.BUILTIN_FORMATS == openpyxl_numbers.BUILTIN_FORMATS
    assert wolfxl_numbers.BUILTIN_FORMATS_REVERSE == openpyxl_numbers.BUILTIN_FORMATS_REVERSE
    assert wolfxl_numbers.BUILTIN_FORMATS_MAX_SIZE == openpyxl_numbers.BUILTIN_FORMATS_MAX_SIZE

    for name in dir(openpyxl_numbers):
        if name.startswith("FORMAT_"):
            assert getattr(wolfxl_numbers, name) == getattr(openpyxl_numbers, name)

    for fmt_id, fmt in openpyxl_numbers.BUILTIN_FORMATS.items():
        assert wolfxl_numbers.builtin_format_code(fmt_id) == fmt
        assert wolfxl_numbers.builtin_format_id(fmt) == fmt_id
        assert wolfxl_numbers.is_builtin(fmt) is True


def test_builtin_number_formats_round_trip_without_custom_numfmts(
    tmp_path: Path,
) -> None:
    import openpyxl.styles.numbers as openpyxl_numbers

    path = tmp_path / "wolfxl-builtin-number-formats.xlsx"
    wb = wolfxl.Workbook()
    ws = wb.active

    for row_idx, (fmt_id, fmt) in enumerate(openpyxl_numbers.BUILTIN_FORMATS.items(), start=1):
        ws.cell(row=row_idx, column=1, value=1234.5)
        ws.cell(row=row_idx, column=1).number_format = fmt
        ws.cell(row=row_idx, column=2, value=fmt_id)
        ws.cell(row=row_idx, column=3, value=fmt)

    wb.save(path)

    openpyxl_wb = openpyxl.load_workbook(path)
    openpyxl_ws = openpyxl_wb.active
    wolfxl_wb = wolfxl.load_workbook(path)
    wolfxl_ws = wolfxl_wb.active

    for row_idx, fmt in enumerate(openpyxl_numbers.BUILTIN_FORMATS.values(), start=1):
        assert openpyxl_ws.cell(row_idx, 1).number_format == fmt
        assert wolfxl_ws.cell(row_idx, 1).number_format == fmt

    with zipfile.ZipFile(path) as zf:
        styles_xml = zf.read("xl/styles.xml").decode()
    assert "<numFmts" not in styles_xml


def test_wolfxl_writes_1904_epoch_dates_like_openpyxl(tmp_path: Path) -> None:
    expected_datetime = dt.datetime(2020, 1, 2, 3, 4, 5)
    expected_date = dt.datetime(1904, 1, 2)

    path = tmp_path / "wolfxl-1904-date-epoch.xlsx"
    wb = wolfxl.Workbook()
    wb.epoch = WOLFXL_CALENDAR_MAC_1904
    ws = wb.active
    ws["A1"] = expected_datetime
    ws["A2"] = expected_date.date()
    wb.save(path)

    openpyxl_wb = openpyxl.load_workbook(path)
    openpyxl_ws = openpyxl_wb.active
    assert openpyxl_wb.epoch == CALENDAR_MAC_1904
    assert openpyxl_ws["A1"].value == expected_datetime
    assert openpyxl_ws["A2"].value == expected_date

    wolfxl_wb = wolfxl.load_workbook(path)
    wolfxl_ws = wolfxl_wb.active
    assert wolfxl_wb.epoch == WOLFXL_CALENDAR_MAC_1904
    assert wolfxl_ws["A1"].value == expected_datetime
    assert wolfxl_ws["A2"].value == expected_date


def test_wolfxl_reads_openpyxl_1904_epoch_cell_values(tmp_path: Path) -> None:
    expected_datetime = dt.datetime(2020, 1, 2, 3, 4, 5)
    expected_date = dt.datetime(1904, 1, 2)

    path = tmp_path / "openpyxl-1904-date-epoch.xlsx"
    wb = openpyxl.Workbook()
    wb.epoch = CALENDAR_MAC_1904
    ws = wb.active
    ws["A1"] = expected_datetime
    ws["A2"] = expected_date.date()
    wb.save(path)

    wolfxl_wb = wolfxl.load_workbook(path)
    wolfxl_ws = wolfxl_wb.active
    assert wolfxl_wb.epoch == WOLFXL_CALENDAR_MAC_1904
    assert wolfxl_ws["A1"].value == expected_datetime
    assert wolfxl_ws["A2"].value == expected_date
