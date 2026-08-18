"""Focused openpyxl workflow parity for names, named styles, and comments."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import wolfxl
from wolfxl.comments import Comment
from wolfxl.styles import NamedStyle
from wolfxl.workbook.defined_name import DefinedName


def test_defined_name_pop_removes_pending_write_before_save(tmp_path: Path) -> None:
    wb = wolfxl.Workbook()
    wb.defined_names["MetricCell"] = DefinedName(
        name="MetricCell",
        attr_text="Sheet!$A$1",
    )

    removed = wb.defined_names.pop("MetricCell")
    assert removed.attr_text == "Sheet!$A$1"

    out = tmp_path / "defined-name-pop.xlsx"
    wb.save(out)

    assert "MetricCell" not in openpyxl.load_workbook(out).defined_names


def test_defined_name_delete_queues_modify_mode_removal(tmp_path: Path) -> None:
    src = tmp_path / "source.xlsx"
    dst = tmp_path / "deleted.xlsx"
    op_wb = openpyxl.Workbook()
    op_wb.defined_names.add(
        openpyxl.workbook.defined_name.DefinedName(
            name="MetricCell",
            attr_text="Sheet!$A$1",
        )
    )
    op_wb.save(src)

    wb = wolfxl.load_workbook(src, modify=True)
    del wb.defined_names["MetricCell"]
    wb.save(dst)

    assert "MetricCell" not in openpyxl.load_workbook(dst).defined_names


def test_add_named_style_rejects_duplicate_like_openpyxl() -> None:
    wb = wolfxl.Workbook()
    wb.add_named_style(NamedStyle(name="Metric"))

    with pytest.raises(ValueError, match="Style Metric exists already"):
        wb.add_named_style(NamedStyle(name="Metric"))

    assert wb.named_styles == ["Normal", "Metric"]


def test_comment_reassignment_copies_bound_comment_like_openpyxl() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    comment = Comment("Needs review", "Wolf")

    ws["A1"].comment = comment
    ws["B1"].comment = comment

    assert ws["A1"].comment is comment
    assert ws["B1"].comment is not comment
    assert ws["A1"].comment.parent is ws["A1"]
    assert ws["B1"].comment.parent is ws["B1"]
    assert ws["B1"].comment.text == "Needs review"
    assert ws["B1"].comment.author == "Wolf"


def test_comment_remove_unbinds_cell_comment_like_openpyxl() -> None:
    wb = wolfxl.Workbook()
    ws = wb.active
    assert ws is not None
    comment = Comment("Temporary", "Wolf")

    ws["A1"].comment = comment
    ws["A1"].comment = None

    assert ws["A1"].comment is None
    assert comment.parent is None
