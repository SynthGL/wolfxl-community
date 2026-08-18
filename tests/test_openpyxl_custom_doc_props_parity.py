"""Focused openpyxl parity for workbook custom document properties."""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest

import wolfxl
from wolfxl.packaging.custom import (
    BoolProperty,
    DateTimeProperty,
    FloatProperty,
    IntProperty,
    LinkProperty,
    StringProperty,
)

openpyxl = pytest.importorskip("openpyxl")
from openpyxl.packaging.custom import (  # noqa: E402
    BoolProperty as OpenpyxlBoolProperty,
)
from openpyxl.packaging.custom import (  # noqa: E402
    DateTimeProperty as OpenpyxlDateTimeProperty,
)
from openpyxl.packaging.custom import (  # noqa: E402
    FloatProperty as OpenpyxlFloatProperty,
)
from openpyxl.packaging.custom import IntProperty as OpenpyxlIntProperty  # noqa: E402
from openpyxl.packaging.custom import LinkProperty as OpenpyxlLinkProperty  # noqa: E402
from openpyxl.packaging.custom import StringProperty as OpenpyxlStringProperty  # noqa: E402


EXPECTED_PROPS = [
    ("Client", "StringProperty", "ACME"),
    ("Count", "IntProperty", 42),
    ("Ratio", "FloatProperty", 2.5),
    ("Reviewed", "BoolProperty", True),
    ("AsOf", "DateTimeProperty", dt.datetime(2024, 1, 2, 3, 4, 5)),
    ("LinkedCell", "LinkProperty", "Sheet!A1"),
]


def _wolfxl_custom_props() -> list[object]:
    return [
        StringProperty(name="Client", value="ACME"),
        IntProperty(name="Count", value=42),
        FloatProperty(name="Ratio", value=2.5),
        BoolProperty(name="Reviewed", value=True),
        DateTimeProperty(name="AsOf", value=dt.datetime(2024, 1, 2, 3, 4, 5)),
        LinkProperty(name="LinkedCell", value="Sheet!A1"),
    ]


def _openpyxl_custom_props() -> list[object]:
    return [
        OpenpyxlStringProperty(name="Client", value="ACME"),
        OpenpyxlIntProperty(name="Count", value=42),
        OpenpyxlFloatProperty(name="Ratio", value=2.5),
        OpenpyxlBoolProperty(name="Reviewed", value=True),
        OpenpyxlDateTimeProperty(name="AsOf", value=dt.datetime(2024, 1, 2, 3, 4, 5)),
        OpenpyxlLinkProperty(name="LinkedCell", value="Sheet!A1"),
    ]


def _snapshot_props(workbook: object) -> list[tuple[str, str, object]]:
    return [
        (prop.name, type(prop).__name__, prop.value)
        for prop in workbook.custom_doc_props
    ]


def _save_openpyxl_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()
    for prop in _openpyxl_custom_props():
        wb.custom_doc_props.append(prop)
    wb.save(path)
    wb.close()


def test_wolfxl_appends_custom_doc_props_and_openpyxl_reopens(tmp_path: Path) -> None:
    wb = wolfxl.Workbook()
    for prop in _wolfxl_custom_props():
        wb.custom_doc_props.append(prop)

    out = tmp_path / "wolfxl-custom-doc-props.xlsx"
    wb.save(out)

    openpyxl_reopened = openpyxl.load_workbook(out)
    try:
        assert _snapshot_props(openpyxl_reopened) == EXPECTED_PROPS
    finally:
        openpyxl_reopened.close()

    wolfxl_reopened = wolfxl.load_workbook(out)
    assert _snapshot_props(wolfxl_reopened) == EXPECTED_PROPS


def test_wolfxl_reads_openpyxl_custom_doc_props_and_extends_them(tmp_path: Path) -> None:
    src = tmp_path / "openpyxl-custom-doc-props.xlsx"
    _save_openpyxl_fixture(src)

    wb = wolfxl.load_workbook(src)
    assert _snapshot_props(wb) == EXPECTED_PROPS

    wb.custom_doc_props.append(StringProperty(name="Stage", value="Draft"))
    out = tmp_path / "wolfxl-extended-custom-doc-props.xlsx"
    wb.save(out)

    reopened = openpyxl.load_workbook(out)
    try:
        assert _snapshot_props(reopened) == [
            *EXPECTED_PROPS,
            ("Stage", "StringProperty", "Draft"),
        ]
    finally:
        reopened.close()


def test_custom_doc_props_duplicate_name_matches_openpyxl() -> None:
    wb = wolfxl.Workbook()
    wb.custom_doc_props.append(StringProperty(name="Client", value="ACME"))
    with pytest.raises(ValueError, match="Property with name Client already exists"):
        wb.custom_doc_props.append(StringProperty(name="Client", value="Other"))

    opx = openpyxl.Workbook()
    try:
        opx.custom_doc_props.append(OpenpyxlStringProperty(name="Client", value="ACME"))
        with pytest.raises(ValueError, match="Property with name Client already exists"):
            opx.custom_doc_props.append(
                OpenpyxlStringProperty(name="Client", value="Other")
            )
    finally:
        opx.close()
